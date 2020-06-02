from django.utils import timezone,formats
from django.template import RequestContext, Context, Template
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.shortcuts import ( 
	render_to_response , 
	render,
	get_object_or_404, 
	reverse
)
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.views.decorators.csrf import csrf_protect
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.contrib.auth.models import User
from .models import *
from django.db import connection
from datetime import datetime , timedelta
import PIL
from PIL import Image
import os
from os.path import basename
from django.conf import settings
from django.db.models import Q
from django.db import connection
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Max
from decimal import Decimal
from django.contrib import messages
from django.core.mail import send_mail
import time
from django.urls import reverse
import json
from django.template.loader import render_to_string
from django.conf.urls import include, url
#from pushjack import GCMClient
#from authorizenet import apicontractsv1
#from authorizenet.apicontrollers import*
from decimal import*
from django.conf import settings
from django.views.decorators.csrf import csrf_protect,csrf_exempt
import calendar
import stripe
from enum import Enum

stripe.api_key = settings.STRIPE_SECRET_KEY
#from googleapiclient.discovery import build
#from googleapiclient.errors import HttpError

# Create your views here.
def dictfetchall(cursor):
		"Return all rows from a cursor as a dict"
		columns = [col[0] for col in cursor.description]
		return [
				dict(zip(columns, row))
				for row in cursor.fetchall()
		]

def admin_login(request):
	if 'admin_id' in request.session:
		return HttpResponseRedirect('/admin/profile/')
	else:
		msg_error = ""
		username = password = ''
		if request.POST:
			username = request.POST.get('username')
			password = request.POST.get('password')
			#print(request.POST.get('remember_me'))
			user = authenticate(username=username, password=password)
			if user is not None:
				if user.is_staff :
					if user.is_active:
						login(request, user)
						menudict = {'usersection':None ,'organizations':None }
						request.session['admin_id'] = user.id
						current_user = UserProfile.objects.get(user_id=user.id)
						request.session['admin_type'] = current_user.role_id
						if current_user.role_id == 5 :
							if Subadminmenulist.objects.filter(userprofile=current_user ).exists():
								submenulistObj = Subadminmenulist.objects.get(userprofile=current_user )
								if submenulistObj.pagename != None :
									menulist = list(submenulistObj.pagename.split(","))
									for menu in menulist:
										if menu == 'usersection':
											menudict['usersection'] = 1
										elif menu == 'organizations':
											menudict['organizations'] = 1
									request.session['subadminmenu'] = menudict
						return HttpResponseRedirect('/admin/dashboard/')
					else:
					  msg_error = "Your account is not active, please contact the site admin."
				else:
					msg_error = "you are not admin!"
			else:
				msg_error = "Your username and/or password were incorrect."
	return render(request,'adminpanel/login.html',{'msg_error': msg_error})


def admin_dashboard(request):
	if 'admin_id' in request.session:
		user_id = request.session['admin_id']
		current_user = UserProfile.objects.get(user_id=user_id)
		current_user_profile = User.objects.get(id=user_id)
		print(dir(request))
		# print(current_user_profile)
		return render(request,'adminpanel/dashboard.html',{'current_user' : current_user,'current_user_profile' : current_user_profile })
	else :
		return HttpResponseRedirect('/admin/login/')


def admin_profile(request):
	if 'admin_id' in request.session:
		user_id = request.session['admin_id']
		current_user = UserProfile.objects.get(user_id=user_id)
		current_user_profile = User.objects.get(id=user_id)
		return render(request,'adminpanel/admin-profile.html',{'current_user' : current_user,'current_user_profile' : current_user_profile })
	else :
		return HttpResponseRedirect('/admin/login/')


def edit_logo(request):
	if 'admin_id' in request.session:
		logo = Logo.objects.get(pk=1)
		if request.method == 'POST':
			logo_id = request.POST.get('logo_id')
			logo_edit = Logo.objects.get(pk=int(logo_id))
			if len(request.FILES) != 0:
				if logo_edit.logo_image :
					image_path = settings.MEDIA_ROOT+"/"+logo_edit.logo_image.name
					base = basename(logo_edit.logo_image.url)
					fname_concat = os.path.splitext(base)[0]
					imgthumb_path=settings.MEDIA_ROOT+"/thumbnail/"+fname_concat+".jpg"
					imgmedium_path=settings.MEDIA_ROOT+"/medium/"+fname_concat+".jpg"
					#os.unlink(image_path)
					#os.unlink(imgthumb_path)
					#os.unlink(imgmedium_path)
					logo_edit.logo_image = request.FILES['logo_image']
				else :
					logo_edit.logo_image = request.FILES['logo_image']
										
				logo_edit.save()
				base = basename(logo_edit.logo_image.url)
				fname_concat = os.path.splitext(base)[0]

				imgthumb = Image.open(settings.MEDIA_ROOT+"/"+logo_edit.logo_image.name)
				imgthumb = imgthumb.resize((200,200), PIL.Image.ANTIALIAS)
				#imgthumb.thumbnail(size, Image.ANTIALIAS)

				imgthumb.save(settings.MEDIA_ROOT+"/thumbnail/"+fname_concat+".jpg")
				logo_edit.thumbnail_pic = "/thumbnail/"+fname_concat+".jpg"

				img = Image.open(settings.MEDIA_ROOT+"/"+logo_edit.logo_image.name)
				img = img.resize((365,309), PIL.Image.ANTIALIAS)
				img.save(settings.MEDIA_ROOT+"/medium/"+fname_concat+".jpg")
				logo_edit.medium_pic = "/medium/"+fname_concat+".jpg"

				logo_edit.save()
			
			logo_edit.logo_title = request.POST['logo_title']
			logo_edit.save()
			return HttpResponseRedirect('/admin/edit-logo/') 
		return render(request,'adminpanel/edit-logo.html',{'logo':logo,'admin':1})
	else :
	 return HttpResponseRedirect('/admin/login/')  


def banner_list(request):
	if 'admin_id' in request.session:
		banner = Banner.objects.all()
		return render(request, 'adminpanel/bannerlisting.html',{'banner':banner,'admin':1})
	else :
		return HttpResponseRedirect('/admin/login/')


def add_banner(request):
	if 'admin_id' in request.session:
		return render(request, 'adminpanel/add-banner.html')
	else :
		return HttpResponseRedirect('/admin/login/')


def submit_banner(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			banner_title = request.POST.get('banner_title')
			banner_text = request.POST.get('banner_text')
			banner_image = request.FILES['banner_image']
			banner_logo = request.FILES['banner_logo']
			banner_section = request.FILES['banner_section']

			banner = Banner(
			banner_title=banner_title,
			banner_text=banner_text,
			banner_image=banner_image,
			banner_order = 1,
			banner_logo=banner_logo,
			banner_section = banner_section
			)
			banner.save()

			base = basename(banner.banner_image.url)
			fname_concat = os.path.splitext(base)[0]
			imgthumb = Image.open(settings.MEDIA_ROOT+"/"+banner.banner_image.name)
			imgthumb.save(settings.MEDIA_ROOT+"/banners/"+fname_concat+".jpg")
			banner.banner_image = "/banners/"+fname_concat+".jpg"
			banner.save()

			basel = basename(banner.banner_logo.url)
			fname_concatl = os.path.splitext(basel)[0]
			imgthumbl = Image.open(settings.MEDIA_ROOT+"/"+banner.banner_logo.name)
			imgthumbl.save(settings.MEDIA_ROOT+"/banners/"+fname_concatl+".png")
			banner.banner_logo = "/banners/"+fname_concatl+".png"
			banner.save()

			return HttpResponseRedirect('/admin/banner-list/')
		else:
			return HttpResponseRedirect('/admin/add-banner/')
	else :
		return HttpResponseRedirect('/admin/login/')


def edit_banner(request, bannerId):
	if 'admin_id' in request.session:
		banner = Banner.objects.get(id=bannerId)
		return render(request, 'adminpanel/edit-banner.html', {'banner':banner})
	else :
		return HttpResponseRedirect('/admin/login/')


def submit_editbanner(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			banner_id = request.POST.get('banner_id')
			banner_title = request.POST.get('banner_title')
			banner_text = request.POST.get('banner_text')
			#banner_image = request.FILES['banner_image']
			#banner_logo = request.FILES['banner_logo']
			banner_edit = Banner.objects.get(pk=int(banner_id))

			if len(request.FILES) != 0:
				if banner_edit.banner_image :
					image_path = settings.MEDIA_ROOT+"/"+banner_edit.banner_image.name
					base = basename(banner_edit.banner_image.url)
					fname_concat = os.path.splitext(base)[0]
					imgthumb_path=settings.MEDIA_ROOT+"/banners/"+fname_concat+".jpg"
					banner_edit.banner_image = request.FILES['banner_image']
				else :
					banner_edit.banner_image = request.FILES['banner_image']

				if banner_edit.banner_logo :
					image_pathl = settings.MEDIA_ROOT+"/"+banner_edit.banner_logo.name
					basel = basename(banner_edit.banner_logo.url)
					fname_concatl = os.path.splitext(basel)[0]
					imgthumb_pathl=settings.MEDIA_ROOT+"/banners/"+fname_concatl+".png"
					banner_edit.banner_logo = request.FILES['banner_logo']
				else :
					banner_edit.banner_logo = request.FILES['banner_logo']
				banner_edit.save()

				img = Image.open(settings.MEDIA_ROOT+"/"+banner_edit.banner_image.name)
				img.save(settings.MEDIA_ROOT+"/banners/"+fname_concat+".jpg")
				banner_edit.banner_image = "/banners/"+fname_concat+".jpg"

				imgl = Image.open(settings.MEDIA_ROOT+"/"+banner_edit.banner_logo.name)
				imgl.save(settings.MEDIA_ROOT+"/banners/"+fname_concat+".png")
				banner_edit.banner_logo = "/banners/"+fname_concat+".png"

				banner_edit.save()
				return HttpResponseRedirect('/admin/banner-list/')
			else:
				data=0
			banner_edit.banner_title = request.POST['banner_title']
			banner_edit.banner_text = request.POST['banner_text']
			banner_edit.banner_section = request.POST['banner_section']
			banner_edit.save()
			return HttpResponseRedirect('/admin/banner-list/')
		else:
			return HttpResponseRedirect('/admin/add-banner/')
	else :
		return HttpResponseRedirect('/admin/login/')


def cms_list(request):
	if 'admin_id' in request.session:
		cms = Cms.objects.all()
		return render(request, 'adminpanel/cmslisting.html',{'cms':cms})
	else :
		return HttpResponseRedirect('/admin/login/')


def add_cms(request):
	if 'admin_id' in request.session:
		return render(request, 'adminpanel/add-cms.html')
	else :
		return HttpResponseRedirect('/admin/login/')


def submit_cms(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			title = request.POST.get('title')
			text = request.POST.get('text')
			short_description = request.POST.get('short_description')
			banner_image = request.FILES['banner_image']

			cms = Cms(
			title=title,
			text=text,
			short_description=short_description,
			banner_image=banner_image,
			user_id_id=1
			)
			cms.save()

			base = basename(cms.banner_image.url)
			fname_concat = os.path.splitext(base)[0]
			imgthumb = Image.open(settings.MEDIA_ROOT+"/"+cms.banner_image.name)
			imgthumb.save(settings.MEDIA_ROOT+"/banners/"+cms.banner_image.name)
			cms.banner_image = "/banners/"+cms.banner_image.name 
			cms.save()
			return HttpResponseRedirect('/admin/cms-list/')
		else:
			return HttpResponseRedirect('/admin/add-cms/')
	else :
		return HttpResponseRedirect('/admin/login/')


def edit_cms(request, cmsId):
	if 'admin_id' in request.session:
		cms = Cms.objects.get(id=cmsId)
		return render(request, 'adminpanel/edit-cms.html', {'cms':cms})
	else :
		return HttpResponseRedirect('/admin/login/')


def submit_editcms(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			cmsid = request.POST.get('id')
			title = request.POST.get('title')
			text = request.POST.get('text')
			short_description = request.POST.get('short_description')

			#banner_image = request.FILES['banner_image']
			#banner_logo = request.FILES['banner_logo']
			cms_edit = Cms.objects.get(pk=int(cmsid))
			if len(request.FILES) != 0:
				if cms_edit.banner_image :
					image_path = settings.MEDIA_ROOT+"/"+cms_edit.banner_image.name
					base = basename(cms_edit.banner_image.url)
					fname_concat = os.path.splitext(base)[0]
					imgthumb_path=settings.MEDIA_ROOT+"/banners/"+cms_edit.banner_image.name
					cms_edit.banner_image = request.FILES['banner_image']
				else :
					cms_edit.banner_image = request.FILES['banner_image']
				cms_edit.save()
				img = Image.open(settings.MEDIA_ROOT+"/"+cms_edit.banner_image.name)
				img.save(settings.MEDIA_ROOT+"/banners/"+cms_edit.banner_image.name)
				cms_edit.banner_image = "/banners/"+cms_edit.banner_image.name
				cms_edit.save()
			else:
				data=0
			cms_edit.title = request.POST['title']
			cms_edit.text = request.POST['text']
			cms_edit.short_description = request.POST['short_description']
			cms_edit.save()
			return HttpResponseRedirect('/admin/cms-list/')
		else:
			return HttpResponseRedirect('/admin/add-cms/')
	else :
		return HttpResponseRedirect('/admin/login/')


def add_social(request):   
	if 'admin_id' in request.session:
		if request.method == 'POST':
					service_title = request.POST['title']
					service_slug = request.POST['title']
					import re
					if service_slug:
					  create_slug = service_slug.lower()
					  create_slug = create_slug.replace (" ", "-")
					  create_slug = re.sub('[^a-zA-Z0-9 \n\.]', '-', create_slug)
					else :
					  create_slug = service_title.lower()
					  create_slug = create_slug.replace (" ", "-")
					  create_slug = re.sub('[^a-zA-Z0-9 \n\.]', '-', create_slug)


					if SocialMedias.objects.filter(slug=create_slug).exists():
						return HttpResponse("0")  # slug exists
					else :    
								
								cover_pic=request.FILES['cover_pic']
								urls = request.POST['urls']
								

								services_save = SocialMedias(
								title=service_title,
								slug=create_slug,
								urls=urls,
								cover_pic=cover_pic,
								)
								services_save.save()
								
							   
								base = basename(services_save.cover_pic.url)
								fname_concat = os.path.splitext(base)[0]

								imgthumb = Image.open(settings.MEDIA_ROOT+"/"+services_save.cover_pic.name)
								imgthumb = imgthumb.resize((200,200), PIL.Image.ANTIALIAS)
										#imgthumb.thumbnail(size, Image.ANTIALIAS)
								imgthumb.save(settings.MEDIA_ROOT+"/thumbnail/"+services_save.cover_pic.name)
								services_save.thumbnail_pic = "/thumbnail/"+services_save.cover_pic.name

								img = Image.open(settings.MEDIA_ROOT+"/"+services_save.cover_pic.name)
								img = img.resize((141,181), PIL.Image.ANTIALIAS)
								img.save(settings.MEDIA_ROOT+"/medium/"+services_save.cover_pic.name)
								services_save.medium_pic = "/medium/"+services_save.cover_pic.name
								return HttpResponse("1")
		return render_to_response('adminpanel/add-social.html')
	else :
	 return HttpResponseRedirect('/admin/login/')  

def list_social(request): 
	if 'admin_id' in request.session:
	   return render_to_response('adminpanel/list-social.html')
	else :
	 return HttpResponseRedirect('/admin/login/')

def sociallist(request): 
		cursor = connection.cursor()
		order = request.GET['order']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 's.title'
		 srch = request.GET['search']
		 cursor.execute("select s.id ,s.title,s.urls from adminpanel_socialmedias as s where lower(s.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			if srch :
			  cursor.execute("select s.id ,s.title,s.urls from adminpanel_socialmedias as s where lower(s.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select s.id ,s.title,s.urls from adminpanel_socialmedias as s where lower(s.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			 
		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "servicename":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			cursor.execute("select s.id ,s.title,s.urls from adminpanel_socialmedias as s where lower(s.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			
		else :
		 column_name = 's.id'
		 order="DESC"
		 cursor.execute("select * from adminpanel_socialmedias where '1'  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_services = dictfetchall(cursor)
		docs_dict = {
			'total': 2,
			'rows': [{'sid': all_service['id'],
					  'title': all_service['title'],
					  'urls': all_service['urls'],
					  'id': all_service['id'],
					  } for all_service in all_services]
		}
		return JsonResponse(docs_dict) 

def socials_details(request,sid):   
	if 'admin_id' in request.session:
		if sid :  
			services=SocialMedias.objects.get(pk=sid)
			return render_to_response('adminpanel/socials_details.html',{'services' : services })
		else:
			 return HttpResponseRedirect('/admin/list-social/')
	else :
	  return HttpResponseRedirect('/admin/login/')



def edit_social(request): 
	if request.method == 'POST':
								
								service_title=request.POST['title']
								service_id=int(request.POST['social_id'])
								service_slug = request.POST['title']
								urls = request.POST['urls']
								
								import re
								if service_slug:
								  create_slug = service_slug.lower()
								  create_slug = create_slug.replace (" ", "-")
								  create_slug = re.sub('[^a-zA-Z0-9 \n\.]', '-', create_slug)
								else :
								  create_slug = service_title.lower()
								  create_slug = create_slug.replace (" ", "-")
								  create_slug = re.sub('[^a-zA-Z0-9 \n\.]', '-', create_slug)
								#if Partners.objects.filter(slug=create_slug).exclude(id=service_id).exists():
									#return HttpResponse("0")  # slug exists    
								
								services_save=SocialMedias.objects.get(pk=service_id)    
								services_save.title=service_title
								services_save.slug=create_slug
								services_save.urls=urls
								
								services_save.save()
								
								if len(request.FILES) != 0:
									if services_save.cover_pic :
											image_path = settings.MEDIA_ROOT+"/"+services_save.cover_pic.name
											base = basename(services_save.cover_pic.url)
											fname_concat = os.path.splitext(base)[0]
											imgthumb_path=settings.MEDIA_ROOT+"/thumbnail/"+services_save.cover_pic.name
											imgmedium_path=settings.MEDIA_ROOT+"/medium/"+services_save.cover_pic.name
											os.unlink(image_path)
											os.unlink(imgthumb_path)
											os.unlink(imgmedium_path)
											services_save.cover_pic=request.FILES['cover_pic']
									else :
											services_save.cover_pic=request.FILES['cover_pic']
										
									services_save.save()
									base = basename(services_save.cover_pic.url)
									fname_concat = os.path.splitext(base)[0]

									imgthumb = Image.open(settings.MEDIA_ROOT+"/"+services_save.cover_pic.name)
									imgthumb = imgthumb.resize((200,200), PIL.Image.ANTIALIAS)
												#imgthumb.thumbnail(size, Image.ANTIALIAS)
									imgthumb.save(settings.MEDIA_ROOT+"/thumbnail/"+services_save.cover_pic.name)
									services_save.thumbnail_pic = "/thumbnail/"+services_save.cover_pic.name

									img = Image.open(settings.MEDIA_ROOT+"/"+services_save.cover_pic.name)
									img = img.resize((141,181), PIL.Image.ANTIALIAS)
									img.save(settings.MEDIA_ROOT+"/medium/"+services_save.cover_pic.name)
									services_save.medium_pic = "/medium/"+services_save.cover_pic.name

									services_save.save()
									
								
								
								return HttpResponse(service_id) 
	else:
	   return HttpResponse("0") 



def delete_social(request): 
	if 'admin_id' in request.session:
		if request.method == 'POST':  
			delete_id = request.POST.get('del_id')   
			
			SocialMedias.objects.filter(pk=int(delete_id)).delete()
			c = SocialMedias.objects.filter(pk=int(delete_id)).count()
			if c:
			   return HttpResponse("0")
			else:
			   return HttpResponse("1")
		else :
			   return HttpResponse("0")      
	else :
	  return HttpResponseRedirect('/admin/login/') 


def add_users(request):
	errormsg = ""
	manglogin = ""
	if request.method == 'POST':
		username=request.POST['username']
		email_id=request.POST['email']
		if User.objects.filter(username=username).exists():    
			errormsg = "username Exists"
			return render(request, 'adminpanel/add-user.html',{'errormsg': errormsg})
		elif User.objects.filter(email=email_id).exists(): 
			errormsg = "Email Exists"
			return render(request, 'adminpanel/add-user.html',{'errormsg': errormsg})
		else :
			if request.POST['user_role'] == '4' :
				manglogin = 1
			else:
				manglogin = 0
				user = User.objects.create_user(
					username=request.POST['username'],
					password=request.POST["password"],
					first_name=request.POST["fname"],
					last_name=request.POST["lname"],
					email=request.POST['email'],
					is_staff=manglogin
				)
				user.save()

				#Save userinfo record
				new_profile = UserProfile(
					user=user
				)
				new_profile.save()
				userRole=request.POST['user_role']
				user_roles = UserWithroles(
					user=user,
					userroles_id=userRole
				)
				user_roles.save()
				
				# userRoles=request.POST.getlist('user_role')
				# for userRole in userRoles :
				# 	user_roles = UserWithroles(
				# 			user=user,
				# 			userroles_id=userRole
				# 	)
				# 	user_roles.save() 
				if user.id != "":
					messages.add_message(request, messages.SUCCESS, 'User Added Successfully!')   
					return HttpResponseRedirect('/admin/add-user/')
				else :
					messages.add_message(request, messages.SUCCESS, 'User cant be added')
					return HttpResponseRedirect('/admin/add-user/')
	return render(request, 'adminpanel/add-user.html',{'errormsg': errormsg})

def get_userdetails(request,uid):
	if 'admin_id' in request.session:
		if uid :
			userdetails = UserProfile.objects.get(user_id=uid)
			current_user = User.objects.get(pk=uid)
			return render_to_response('adminpanel/usersprofile.html', {'userdetails': userdetails, 'current_user' : current_user , 'uid' : uid })
		else :
			return HttpResponseRedirect('/admin/profile/')
	else :
		return HttpResponseRedirect('/admin/login/')
 
def user_edit(request,user_id):
	if 'admin_id' in request.session:
		if user_id :
			userdetails = UserProfile.objects.get(user_id=user_id)
			current_user = User.objects.get(pk=user_id)
			userroles = UserWithroles.objects.filter(user_id=user_id)
			'''
			c = UserWithroles.objects.filter(user_id=int(user_id)).count()
			if c == 1 :
				for userrole in  userroles : 
					if userrole.userroles_id == 2 :
						  flag_role = "buyer"
					else : 
						 flag_role = "vendor"
			else :     
				flag_role = "both"  '''
				
			user_role = UserWithroles.objects.get(user_id=int(user_id))
			flag_role = user_role.userroles_id
			return render_to_response('adminpanel/edit-user.html', {'userdetails': userdetails, 'current_user' : current_user , 'user_id' : user_id, 'userroles' : userroles, 'flag_role' : flag_role })
		else :
		   return render_to_response('adminpanel/usersprofile.html', {'userdetails': userdetails, 'current_user' : current_user , 'uid' : user_id })
	else :
	   return HttpResponseRedirect('/admin/login/')
   
def user_editdetails(request):
	if request.method == 'POST' and 'btn_edit' in request.POST :
		edit_id=request.POST['edituser_id']
		uprofile_update = UserProfile.objects.get(user_id=int(edit_id))
		user_edit = User.objects.get(pk=int(edit_id))
		user_edit.first_name=request.POST.get('fname', "NONE")
		user_edit.last_name=request.POST.get('lname', "NONE")
		user_edit.save()
		uprofile_update.date=request.POST.get('birthdate', "NONE")
		uprofile_update.month=request.POST.get('birthmonth', "NONE")
		uprofile_update.year=request.POST.get('birthyear', "NONE")
		uprofile_update.gender=request.POST.get('gender', "NONE")
		uprofile_update.phone=request.POST.get('phone', "NONE")
		uprofile_update.address=request.POST.get('address', "NONE")
		uprofile_update.city=request.POST.get('city', "NONE")
		uprofile_update.state=request.POST.get('state', "NONE")
		uprofile_update.country=request.POST.get('country', "NONE")
		uprofile_update.save()
		
		if UserWithroles.objects.filter(user_id=int(edit_id)).exists():
			user_role=UserWithroles.objects.get(user_id=int(edit_id))
			user_role.userroles_id = request.POST['user_role']
			user_role.save()
			'''        
			userRoles=request.POST.getlist('user_role')
			for userRole in userRoles :
				user_roles = UserWithroles(
								user_id=edit_id,
								userroles_id=userRole
							)
				user_roles.save()  '''
			
			if 'is_featured' in request.POST:
				uprofile_update.isfeatured=request.POST['is_featured']
				uprofile_update.save()  
			else :
				uprofile_update.isfeatured=0
				uprofile_update.save()     
			
			if len(request.FILES) != 0:
				if uprofile_update.picture :
					image_path = settings.MEDIA_ROOT+"/"+uprofile_update.picture.name
					base = basename(uprofile_update.picture.url)
					fname_concat = os.path.splitext(base)[0]
					imgthumb_path=settings.MEDIA_ROOT+"/thumbnail/"+fname_concat+".jpg"
					imgmedium_path=settings.MEDIA_ROOT+"/medium/"+fname_concat+".jpg"
					os.unlink(image_path)
					os.unlink(imgthumb_path)
					os.unlink(imgmedium_path)
					uprofile_update.picture = request.FILES['vendor_logo']
				else :
					uprofile_update.picture = request.FILES['vendor_logo']
								
				uprofile_update.save()
				base = basename(uprofile_update.picture.url)
				fname_concat = os.path.splitext(base)[0]

				imgthumb = Image.open(settings.MEDIA_ROOT+"/"+uprofile_update.picture.name)
				imgthumb = imgthumb.resize((200,200), PIL.Image.ANTIALIAS)
										#imgthumb.thumbnail(size, Image.ANTIALIAS)
				imgthumb.save(settings.MEDIA_ROOT+"/thumbnail/"+fname_concat+".jpg")
				uprofile_update.thumbnail_pic = "/thumbnail/"+fname_concat+".jpg"

				img = Image.open(settings.MEDIA_ROOT+"/"+uprofile_update.picture.name)
				img = img.resize((365,309), PIL.Image.ANTIALIAS)
				img.save(settings.MEDIA_ROOT+"/medium/"+fname_concat+".jpg")
				uprofile_update.medium_pic = "/medium/"+fname_concat+".jpg"

				uprofile_update.save()    
			return HttpResponseRedirect('/admin/useredit/'+edit_id+'/')
	elif request.method == 'POST' and 'btn_delete' in request.POST : 
		delete_id=request.POST['edituser_id']
		if User.objects.filter(pk=int(delete_id)).exists():
			User.objects.filter(pk=int(delete_id)).delete()
			c = User.objects.filter(pk=int(delete_id)).count()
			if c:
			   return HttpResponseRedirect('/admin/useredit/'+edit_id+'/')   
			else:
				return HttpResponseRedirect('/admin/userslist/')
		else:
			return HttpResponseRedirect('/admin/useredit/'+edit_id+'/')
	else :    
	  return HttpResponseRedirect('/admin/useredit/'+edit_id+'/')  
 

 
def block_users(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			block_id = request.POST.get('block_id', "NONE")
			reason = request.POST.get('reason', "NONE")
			if User.objects.filter(pk=block_id).exists(): 
					user_block=User.objects.get(pk=int(block_id))
					user_block.is_active = False
					user_block.save()
					user_profileblock =UserProfile.objects.get(user_id=block_id)
					user_profileblock.reason = reason
					user_profileblock.save()
					return HttpResponse(1) # 1 means user is blocked 
			else:
				return HttpResponse(0) # 0 means user not exists
		else :
			return HttpResponseRedirect(3) # 3 means data is not in post
	else :
		  return HttpResponse(2) # 2 means admin is not logged in

def unblock_users(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			blocked_id = request.POST.get('blocked_id', "NONE")
			if User.objects.filter(pk=blocked_id).exists(): 
					user_block=User.objects.get(pk=int(blocked_id))
					user_block.is_active = True
					user_block.save()
					user_profileblock =UserProfile.objects.get(user_id=blocked_id)
					user_profileblock.reason = ""
					user_profileblock.save()
					activate_mail=EmailTemplates.objects.get(pk=3) 
					t = Template(activate_mail.templatebody)
					c = Context({'name': user_block.first_name, 'msg_description': "Your Account is activated" })
					msg_html = t.render(c)
					send_mail("Account Activation", 'hello world again', 'nits.nawed@gmail.com', ['nits.nawed@gmail.com',user_block.email], html_message=msg_html)
					return HttpResponse(1) # 1 means user is active now 
			else:
				return HttpResponse(0) # 0 means user not exists
		else :
			return HttpResponseRedirect(3) # 3 means data is not in post
	else :
		  return HttpResponse(2) # 2 means admin is not logged in 

def admin_logout(request):
	logout(request)
	return HttpResponseRedirect('/admin/login/')


def add_how_it_works(request):
	if 'admin_id' in request.session:
		#errormsg = "";
		if request.method == 'POST':
				title=request.POST['title']
				text=request.POST['text']
				pic = request.FILES['pic']
				howitworks = HowitWorks(
				title=title,
				text=text,
				pic=pic
				)
				howitworks.save()
				return HttpResponseRedirect('/admin/how-it-works/')
			
		return render(request,'adminpanel/howitworks.html')
	else :
	 return HttpResponseRedirect('/admin/login/')       
	
def how_it_works_list(request):    
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/howitworkslisting.html')
	else :
	 return HttpResponseRedirect('/admin/login/') 
 
def get_how_it_works_list(request): 
		totalrows = HowitWorks.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 'adminpanel_howitworks.title'
		 srch = request.GET['search']
		 cursor.execute("select * from adminpanel_howitworks where lower(adminpanel_howitworks.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 'adminpanel_howitworks.title'
			else :
			 column_name = 'adminpanel_howitworks.id'

			if srch :
			  cursor.execute("select * from adminpanel_howitworks where lower(adminpanel_howitworks.title) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select * from adminpanel_howitworks ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "title":
				column_name = 'adminpanel_howitworks.title'
			else :
			 column_name = 'adminpanel_howitworks.id'

			cursor.execute("select * from adminpanel_howitworks ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		else :
		 column_name = 'adminpanel_howitworks.id'
		 cursor.execute("select * from adminpanel_howitworks ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_howitworks = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'id': all_howitwork['id'],
					  'title': all_howitwork['title'],
					  'pic': all_howitwork['pic'],
					  'created_date': formats.date_format( all_howitwork['created_date'], "Y-m-d H:i:s"), 
					  'editid': all_howitwork['id'],
					  'deleteid': all_howitwork['id'],
					} for all_howitwork in all_howitworks]
		}
		return JsonResponse(docs_dict)   
	
def delete_how_it_work(request):
	if request.method == 'POST':
	 import json
	 delete_id = request.POST.get('del_id')
	 if HowitWorks.objects.filter(pk=int(delete_id)).exists():
			HowitWorks.objects.filter(pk=int(delete_id)).delete()
			c = HowitWorks.objects.filter(pk=int(delete_id)).count()
			if c:
			   return HttpResponse("0")  
			else:
				return HttpResponse("1")  
	 else:
			   return HttpResponse("0")  

def edit_how_it_work(request):
	if request.method == 'POST':
		_id = request.POST.get('edit_id')
		obj_edit = HowitWorks.objects.get(pk=int(_id))
		if len(request.FILES) != 0:
			if obj_edit.pic :
				image_path = settings.MEDIA_ROOT+"/"+obj_edit.pic.name
				os.unlink(image_path)
				obj_edit.pic = request.FILES['pic']
				obj_edit.title = request.POST['title']
				obj_edit.text = request.POST['text']  
				obj_edit.save() 
		else :
			  obj_edit.title = request.POST['title']  
			  obj_edit.text = request.POST['text']  
			  obj_edit.save()  
			  
		return HttpResponseRedirect('/admin/how-it-work-details/'+_id+'/')
	
def how_it_work_details(request,editId):
   if 'admin_id' in request.session:
			 if editId :
						 if HowitWorks.objects.filter(pk=editId).exists():
							   obj_edit=HowitWorks.objects.get(pk=editId)
						 else :
						   return HttpResponseRedirect('/admin/how-it-works/')
			 else :
				  return HttpResponseRedirect('/admin/how-it-works/')


			 return render(request,'adminpanel/edit-howitworks.html',{'obj_edit': obj_edit})
   else :
	   return HttpResponseRedirect('/admin/login/')

def add_templates(request):
	if 'admin_id' in request.session:
	   if request.POST:
		   template_title = request.POST.get('template_title')
		   template_subject = request.POST.get('template_subject')
		   template_body = request.POST.get('template_body')
		   etemplates = EmailTemplates(
			templatename = template_title,
			templatebody = template_subject,
			subject = template_body
			)
		   etemplates.save()
		   
		   messages.add_message(request, messages.SUCCESS, 'Category Added Successfully!')   
		   return HttpResponseRedirect('/admin/email-templates/')      
	   return render(request,'adminpanel/add-templates.html') 
	else:
	   return HttpResponseRedirect('/admin/login/')

def email_templates(request):
	if 'admin_id' in request.session:
	   emailtemplates = EmailTemplates.objects.all()
	   return render(request,'adminpanel/email-templates.html',{'emailtemplates' : emailtemplates })
	else :
		return HttpResponseRedirect('/admin/login/') 

def edit_emailtemplates(request,templateID):
	if 'admin_id' in request.session:
		emailtemplate = EmailTemplates.objects.get(pk=templateID)
		return render(request,'adminpanel/email-templatesdetails.html',{'emailtemplate' : emailtemplate })
	else :
	 return HttpResponseRedirect('/admin/login/') 
 
def edit_mailtemplate(request):
	if request.POST :
		template_id = request.POST.get('template_id')
		template_title = request.POST.get('template_title')
		template_subject = request.POST.get('template_subject')
		template_body = request.POST.get('template_body')
		
		emailtemplate = EmailTemplates.objects.get(pk=template_id)
		emailtemplate.templatename=template_title
		emailtemplate.templatebody=template_body
		emailtemplate.subject=template_subject
		emailtemplate.save() 
		return HttpResponseRedirect('/admin/edit-emailtemplate/'+template_id+'/')
	else :
		return HttpResponseRedirect('/admin/login/')
	

def add_seopages(request):
	if 'admin_id' in request.session:
	   if request.POST:
		   pagename = request.POST.get('pagename')
		   meta_title = request.POST.get('meta_title')
		   meta_keyword = request.POST.get('meta_keyword')
		   meta_description = request.POST.get('meta_description')
		   seopages = SeoPages(
			pagename = pagename,
			meta_title = meta_title,
			meta_keyword = meta_keyword,
			meta_description = meta_description
			)
		   seopages.save()
		   
		   messages.add_message(request, messages.SUCCESS, 'Meta Added Successfully!')   
		   return HttpResponseRedirect('/admin/seo-pages/')      
	   return render(request,'adminpanel/add-seopages.html') 
	else:
	   return HttpResponseRedirect('/admin/login/')     

def seo_pages(request):
	if 'admin_id' in request.session:
		emailtemplates = SeoPages.objects.all()
		return render(request,'adminpanel/seo-pages.html',{'emailtemplates' : emailtemplates })
	else :
		return HttpResponseRedirect('/admin/login/') 

def edit_seopages(request,templateID):
	if 'admin_id' in request.session:
		emailtemplate = SeoPages.objects.get(pk=templateID)
		return render(request,'adminpanel/seo-pagessdetails.html',{'emailtemplate' : emailtemplate })
	else :
	 return HttpResponseRedirect('/admin/login/') 

def edit_pages(request):
	if request.POST :
		template_id = request.POST.get('template_id')
		pagename = request.POST.get('pagename')
		meta_title = request.POST.get('meta_title')
		meta_keyword = request.POST.get('meta_keyword')
		meta_description = request.POST.get('meta_description')
		
		emailtemplate = SeoPages.objects.get(pk=template_id)
		emailtemplate.pagename=pagename
		emailtemplate.meta_title=meta_title
		emailtemplate.meta_keyword=meta_keyword
		emailtemplate.meta_description=meta_description
		emailtemplate.save() 
		return HttpResponseRedirect('/admin/edit-seopages/'+template_id+'/')
	else :
		return HttpResponseRedirect('/admin/login/') 


def contact_messages(request):
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/contactmessages.html')
	else :
	 return HttpResponseRedirect('/admin/login/')

def get_contact_messages(request):
		totalrows = ContactUs.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 'adminpanel_contactus.fromemail'
		 srch = request.GET['search']
		 cursor.execute("select * from adminpanel_contactus"
		 " where lower(adminpanel_contactus.fromemail) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "message_from":
				column_name = 'adminpanel_contactus.fromemail'
			else :
			 column_name = 'adminpanel_contactus.id'    
			
			if srch :
			  cursor.execute("select * from adminpanel_contactus"
			  " where lower(adminpanel_contactus.fromemail) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select * from adminpanel_contactus"
			 " ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "message_from":
				column_name = 'adminpanel_contactus.fromemail'
			else :
			 column_name = 'adminpanel_contactus.id'

			cursor.execute("select * from adminpanel_contactus"
			" ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		else :
		 column_name = 'adminpanel_contactus.id'
		 cursor.execute("select * from adminpanel_contactus"
		 " ORDER BY "+column_name+" "+order+" LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_contactmessages = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'message_id': all_contactmessage['id'],
					  'subject': all_contactmessage['subject'],
					  'message_from': all_contactmessage['fromemail'],
					  'entry_date': formats.date_format(all_contactmessage['entry_date'], "Y-m-d H:i:s"),
					  'is_replied': all_contactmessage['is_replied'],
					  'actionid': all_contactmessage['id'],
					} for all_contactmessage in all_contactmessages]
		}
		return JsonResponse(docs_dict)
	

def reply_message(request,msgid):
	if 'admin_id' in request.session:
	   message=ContactUs.objects.get(pk=msgid) 
	   return render(request,'adminpanel/reply-messages.html',{'message' : message})
	else :
	 return HttpResponseRedirect('/admin/login/')


def send_contactmail(request): 
	if 'admin_id' in request.session:
		if request.POST :
			msgid = request.POST.get('message_id') 
			msg_send = request.POST.get('message') 
			subject = request.POST.get('msg_subject')
			msg_to = request.POST.get('msg_to')
			name = request.POST.get('from_name')
			message=ContactUs.objects.get(pk=msgid) 
			message.is_replied = 1
			message.save()
			contact_mail=EmailTemplates.objects.get(pk=3) 
			t = Template(contact_mail.templatebody)
			c = Context({ 'msg_description': msg_send, 'name' : name })
			msg_html = t.render(c)
			send_mail(subject, 'hello world again', 'nits.nawed@gmail.com', ['nits.nawed@gmail.com',msg_to], html_message=msg_html)
			messages.add_message(request, messages.SUCCESS, 'Mail Send Successfully!')
			return HttpResponseRedirect('/admin/contact-messages/')
	else :
	 return HttpResponseRedirect('/admin/login/')


def view_contactmessage(request,msgid): 
	if 'admin_id' in request.session:
	   message=ContactUs.objects.get(pk=msgid) 
	   return render(request,'adminpanel/message-detail.html',{'message' : message})
	else :
	 return HttpResponseRedirect('/admin/login/')


def add_organization(request):   
	if 'admin_id' in request.session:
		#all_categories = recursive_category("0")
		if request.method == 'POST':
					organization_name = request.POST['organization_name']
					email=request.POST['email']
					phone=request.POST['phone']
					web_url=request.POST['web_url']
					tax_id=request.POST['tax_id']
					cause=request.POST['cause']
					about_us=request.POST['about_us']
					our_activity=request.POST['our_activity']
					why_us=request.POST['why_us']
					address=request.POST['city2']
					cover_pic=request.FILES['photo']
					
					services_save = Organization(
					organization_name=organization_name,
					email=email,
					phone=phone,
					web_url=web_url,
					tax_id=tax_id,
					cause=cause,
					about_us=about_us,
					our_activity=our_activity,
					why_us=why_us,
					address=address,
					photo=cover_pic
					)
					services_save.save()
								
					
					img = Image.open(settings.MEDIA_ROOT+"/"+"No_Image_Available.png")
					img = img.resize((432,260), PIL.Image.ANTIALIAS)
					img.save(settings.MEDIA_ROOT+"/medium/"+services_save.photo.name+"")
					services_save.photo = "/medium/"+services_save.photo.name+""

								
					return HttpResponse("1")
		return render(request,'adminpanel/add-organization.html')
	else :
	 return HttpResponseRedirect('/admin/login/')



def list_organization(request): 
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/list-organization.html')
	else :
	 return HttpResponseRedirect('/admin/login/') 


def getorglistt(request): 
		totalrows = Organization.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']
		adminid=request.session['admin_id']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 's.organization_name'
		 srch = request.GET['search']
		 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			if srch :
			  cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			 
		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "servicename":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			
		else :
		 column_name = 's.id'
		 order="DESC"
		 cursor.execute("select * from adminpanel_organization where id!=0 LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_services = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'sid': all_service['id'],
					  'title': all_service['organization_name'],
					  'phone': all_service['phone'],
					  'email': all_service['email'],
					  'id': all_service['id'],
					  } for all_service in all_services]
		}
		return JsonResponse(docs_dict)



def organization_details(request,sid):   
	if 'admin_id' in request.session:
		if sid :  
			#all_categories = recursive_category("0")
			services=Organization.objects.get(pk=sid)
			return render(request,'adminpanel/organization-detail.html',{'services' : services })
		else:
			 return HttpResponseRedirect('/admin/list-artical/')
	else :
	  return HttpResponseRedirect('/admin/login/')


def organization_edit(request): 
	if request.method == 'POST':
								
								organization_name=request.POST['organization_name']
								service_id=int(request.POST['service_id'])
								email=request.POST['email']
								phone=request.POST['phone']
								web_url=request.POST['web_url']
								tax_id=request.POST['tax_id']
								cause=request.POST['cause']
								about_us=request.POST['about_us']
								our_activity=request.POST['our_activity']
								why_us=request.POST['why_us']
								address=request.POST['city2']
							   
								services_save=Organization.objects.get(pk=service_id)    
								services_save.organization_name=organization_name
								services_save.email=email
								services_save.phone=phone
								services_save.web_url=web_url
								services_save.tax_id=tax_id
								services_save.cause=cause
								services_save.about_us=about_us
								services_save.our_activity=our_activity
								services_save.why_us=why_us
								services_save.address=address
								
								services_save.save()
								
								if len(request.FILES) != 0:
									if services_save.photo :
											image_path = settings.MEDIA_ROOT+"/"+services_save.photo.name
											imgthumb_path=settings.MEDIA_ROOT+"/thumbnail/"+services_save.photo.name+""
											imgmedium_path=settings.MEDIA_ROOT+"/medium/"+services_save.photo.name+""
											#os.unlink(image_path)
											#os.unlink(imgthumb_path)
											#os.unlink(imgmedium_path)
											services_save.photo=request.FILES['cover_pic']
									else :
											services_save.photo=request.FILES['cover_pic']
										
									services_save.save()
									#base = basename(services_save.cover_pic.url)
									#fname_concat = os.path.splitext(base)[0]

									imgthumb = Image.open(settings.MEDIA_ROOT+"/"+"No_Image_Available.png")
									imgthumb = imgthumb.resize((200,200), PIL.Image.ANTIALIAS)
												#imgthumb.thumbnail(size, Image.ANTIALIAS)
									imgthumb.save(settings.MEDIA_ROOT+"/thumbnail/"+services_save.photo.name+"")
									services_save.photo = "/medium/"+services_save.photo.name+""

									services_save.save()
									
								
								
								return HttpResponse(service_id) 
	else:
	   return HttpResponse("0")



def delete_organization(request): 
	if 'admin_id' in request.session:
		if request.method == 'POST':  
			delete_id = request.POST.get('del_id')   
			
			Organization.objects.filter(pk=int(delete_id)).delete()
			c = Organization.objects.filter(pk=int(delete_id)).count()
			if c:
			   return HttpResponse("0")
			else:
			   return HttpResponse("1")
		else :
			   return HttpResponse("0")      
	else :
	  return HttpResponseRedirect('/admin/login/')



def pendingrequest(request): 
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/pendingrequest.html')
	else :
	 return HttpResponseRedirect('/admin/login/')


def getpendinglist(request): 
		totalrows = CordinatorRequest.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']
		adminid=request.session['admin_id']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 's.organization_name'
		 srch = request.GET['search']
		 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			if srch :
			  cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			 
		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "servicename":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			
		else :
		 column_name = 's.id'
		 order="DESC"
		 cursor.execute("select s.id as reid,s.user_id_id,s.org_id_id,s.status,u.id,u.first_name,u.last_name,u.email,o.id,o.organization_name,o.phone from adminpanel_cordinatorrequest as s INNER JOIN auth_user as u ON s.user_id_id=u.id INNER JOIN adminpanel_organization as o on s.org_id_id=o.id where s.status='Pending' LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_services = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'sid': all_service['reid'],
					  'title': all_service['organization_name'],
					  'name': all_service['first_name']+" "+all_service['last_name'],
					  'email': all_service['email'],
					  'phone': all_service['phone'],
					  'id': all_service['reid'],
					  'user_id': all_service['user_id_id'],
					  'org_id': all_service['org_id_id'],
					  } for all_service in all_services]
		}
		return JsonResponse(docs_dict)

def get_approvaldetails(request,uid):
	if 'admin_id' in request.session:
		if uid :
			crequest = CordinatorRequest.objects.get(id=uid)
			current_user = User.objects.get(pk=crequest.user_id_id)
			current_org = Organization.objects.get(pk=crequest.org_id_id)
			user_profile = UserProfile.objects.get(user_id_id=crequest.user_id_id)
			return render(request,'adminpanel/approvaldetails.html', {'crequest': crequest, 'current_user' : current_user , 'uid' : uid , 'current_org' : current_org , 'user_profile' : user_profile })
		else :
			return HttpResponseRedirect('/admin/profile/')
	else :
		return HttpResponseRedirect('/admin/login/')



def approve_coordinator(request): 
	if request.method == 'POST':
								
		user_id=request.POST['user_id']
		organization_id=request.POST['organization_id']
		
		cursor = connection.cursor()
		cursor.execute("update adminpanel_cordinatorrequest set status = 'Approved' where user_id_id='"+str(user_id)+"' and org_id_id='"+str(organization_id)+"'")
		
		cursor1 = connection.cursor()
		cursor1.execute("update adminpanel_userprofile set role_id = '3' where user_id_id='"+str(user_id)+"'")
		
		return HttpResponse(1) 
	else:
	   return HttpResponse("0")


def listcoordinator(request): 
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/listcoordinator.html')
	else :
	 return HttpResponseRedirect('/admin/login/')

def getcordlist(request): 
		totalrows = UserProfile.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']
		adminid=request.session['admin_id']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 's.organization_name'
		 srch = request.GET['search']
		 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			if srch :
			  cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			 
		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "servicename":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			
		else :
		 column_name = 's.id'
		 order="DESC"
		 cursor.execute("select s.id ,s.first_name,s.last_name,s.email,p.phone_number,p.user_id_id,p.role_id from auth_user as s INNER JOIN adminpanel_userprofile as p ON s.id=p.user_id_id where p.role_id=3 LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_services = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'sid': all_service['id'],
					  'name': all_service['first_name']+" "+all_service['last_name'],
					  'phone': all_service['phone_number'],
					  'email': all_service['email'],
					  'id': all_service['id'],
					  } for all_service in all_services]
		}
		return JsonResponse(docs_dict)

def listvolunteers(request): 
	if 'admin_id' in request.session:
	   return render(request,'adminpanel/listvolunteers.html')
	else :
	 return HttpResponseRedirect('/admin/login/')


def getvollist(request): 
		totalrows = UserProfile.objects.count()
		cursor = connection.cursor()
		order = request.GET['order']
		adminid=request.session['admin_id']

		if 'search' in request.GET and 'sort' not in request.GET :
		 column_name = 's.organization_name'
		 srch = request.GET['search']
		 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		elif 'sort' in request.GET and 'search' in request.GET :
			srch = request.GET['search']
			sort = request.GET['sort']
			if sort == "title":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			if srch :
			  cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			else:
			 cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			 
		elif 'sort' in request.GET and 'search' not in request.GET:
			sort = request.GET['sort']
			if sort == "servicename":
				column_name = 's.name'
			else :
			 column_name = 'c.category_name'

			cursor.execute("select s.id ,s.organization_name,s.email,s.phone from adminpanel_organization as s where s.id!=0 and lower(s.organization_name) like lower('%"+srch+"%') ORDER BY "+column_name+" "+order+"  LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])
			
		else :
		 column_name = 's.id'
		 order="DESC"
		 cursor.execute("select s.id ,s.first_name,s.last_name,s.email,p.phone_number,p.user_id_id,p.role_id from auth_user as s INNER JOIN adminpanel_userprofile as p ON s.id=p.user_id_id where p.role_id=2 LIMIT "+request.GET['limit']+" OFFSET "+ request.GET['offset'])

		all_services = dictfetchall(cursor)
		docs_dict = {
			'total': totalrows,
			'rows': [{'sid': all_service['id'],
					  'name': all_service['first_name']+" "+all_service['last_name'],
					  'phone': all_service['phone_number'],
					  'email': all_service['email'],
					  'id': all_service['id'],
					  } for all_service in all_services]
		}
		return JsonResponse(docs_dict)

def admin_management(request): 
	if 'admin_id' in request.session:
		if UserProfile.objects.filter(Q(role_id=5)).exists(): 
			return render(request,'adminpanel/adminmanagement.html',{"UserProfileCollections": 1})
		else:
			return render(request,'adminpanel/adminmanagement.html',{"UserProfileCollections": 0})
		#    for p in pp:
		#        print(p.role.desc)
		# print(dir(UserProfile))
	else :
		return HttpResponseRedirect('/admin/login/')

def jsonsubadminmanagement(request):
	if 'admin_id' in request.session:
		if UserProfile.objects.filter(Q(role_id=5)).exists(): 
			columns = [               # atatable column index  => database column name
				'up.id',
				'u.first_name',
				'u.last_name',
				'u.email',
				'up.phone_number',
				'opertaions',
			]
			columns1 = [               # atatable column index  => database column name
				'id',
				'user_id__first_name',
				'user_id__last_name',
				'user_id__email',
				'phone_number',
			]
			cursor = connection.cursor()
			requestData = request.GET
			total = UserProfile.objects.filter(Q(role_id=5)).count()
			sql = "SELECT up.id, u.first_name, u.last_name, u.email, up.phone_number,up.role_id "
			sql =sql + " FROM adminpanel_userprofile  AS up  INNER JOIN auth_user AS u ON up.user_id_id=u.id WHERE up.role_id=5"

			# test = UserProfile.objects.select_related('user_id').filter(Q(role_id=5)).count()
			test = UserProfile.objects.select_related('user_id').filter(Q(role_id=5))
			# p =   test.filter(Q(user_id__first_name__startswith='a') | Q(user_id__last_name__startswith='a') | Q(user_id__email__contains='a') | Q(phone_number__contains='a') | Q(id__contains='a'))
			# p =   test.filter(Q(user_id__first_name__startswith='a') | Q(user_id__last_name__startswith='a') | Q(user_id__email__contains='a') | Q(phone_number__contains='a') | Q(id__contains='a')).count()
			# print(p.query)
			# print(p)
			# print(test.query)
			if requestData.get('search[value]') :
				sql =sql + " AND ( u.first_name LIKE  '" + str(requestData.get('search[value]')) + "%'"
				sql =sql + " OR u.last_name LIKE '" + str(requestData.get('search[value]')) + "%'"
				sql =sql + " OR u.email LIKE '%" + str(requestData.get('search[value]')) + "%'"
				sql =sql + " OR CAST(up.phone_number AS text) LIKE '%" + str(requestData.get('search[value]')) + "%'"
				sql =sql + " OR CAST(up.id AS text) LIKE '%" + str(requestData.get('search[value]')) + "%' )"
				
			cursor.execute(sql)
			totalFiltered = cursor.rowcount
					   
			sql = sql + " ORDER BY  "+str( columns[int(requestData.get('order[0][column]')) ] ) +" "+ str(requestData.get('order[0][dir]')) +" LIMIT " + str(requestData.get('length')) + " OFFSET " + str(requestData.get('start')) + "  " 
			cursor.execute(sql)
			# print(sql)
			ordering_direction = str( columns1[int(requestData.get('order[0][column]')) ] )
			if(str(requestData.get('order[0][dir]'))=='desc'):
				ordering_direction='-'+ordering_direction
			# print(test.query)
			# print(ordering_direction)
			# print(requestData.get('start'))
			# print(requestData.get('length'))
			# uu = test.order_by(ordering_direction)[int(requestData.get('start')) : (int(requestData.get('length')) + int(requestData.get('start')) ) ]  #offset : (offset+limit)
			# print(uu)
			# datag=[]
			# for singleVal in uu :
			#     print(singleVal)
			#     singeListy = []
			#     name = singleVal.user_id.first_name +" "+ singleVal.user_id.last_name
			#     singeListy.append(singleVal.id)
			#     singeListy.append(name)
			#     singeListy.append(singleVal.user_id.email)
			#     singeListy.append(singleVal.phone_number)
			#     singeListy.append('<a href="'+reverse("edit_subadmin", args=[singleVal.id])+'">Edit</a> <a href="'+reverse("delete_subadmin", args=[singleVal.id])+'" href="JavaScript:void(0);" onclick="confirmationmsg(this)">Delete</a>')
			#     datag.append(singeListy)
			# print(datag)
			UserProfileCollections = cursor.fetchall() 
			data = []
			for singleVal in UserProfileCollections :
				print(singleVal)
				singeList = []
				name = singleVal[1] +" "+ singleVal[2]
				singeList.append(singleVal[0])
				singeList.append(name)
				singeList.append(singleVal[3])
				singeList.append(singleVal[4] )
				singeList.append('<a href="'+reverse("edit_subadmin", args=[singleVal[0]])+'">Edit</a> <a href="'+reverse("delete_subadmin", args=[singleVal[0]])+'" href="JavaScript:void(0);" onclick="confirmationmsg(this)">Delete</a>')
				data.append(singeList)
			usercollection_dict = {}
			usercollection_dict['draw'] = requestData.get('draw')
			usercollection_dict['recordsTotal'] = total
			usercollection_dict['recordsFiltered'] = totalFiltered
			usercollection_dict['data'] = data
			return JsonResponse(usercollection_dict)
	   
def add_subadmin(request):
	if 'admin_id' in request.session:
		if request.method == 'POST':
			# try:
				if User.objects.filter(Q(username=request.POST['username']) | Q(email=request.POST['email'])).exists():
					messages.error(request, 'username or email already exist!.')
					return render(request, 'adminpanel/addsubadmin.html',
					{
						'username'        : request.POST.get ('username', ''),
						'firstname'       : request.POST.get ('firstname', ''),
						'lastname'        : request.POST.get ('lastname', ''),
						'email'           : request.POST.get ('email', ''),
						'password'        : request.POST.get ('pwd', ''),
						'phone_number'    : request.POST.get ('phone_number', ''),
						'address'         : request.POST.get ('address', ''),
						
					})
				else:
					activate_no = int(time.time())

					user = User.objects.create_user(
						username=request.POST['username'],
						password=request.POST["pwd"],
						first_name=request.POST["firstname"],
						last_name=request.POST["lastname"],
						email=request.POST['email'],
						is_staff=1 # it checks if user is admin or not
					)
					user.save()
					if len(request.FILES) != 0:
						profile_image = request.FILES['profile_image']
						profile_image_name = str(profile_image).replace(' ', '-')
						profile_image_name = 'inteer-'+str(user.id)+'-'+ str(profile_image_name)
						handle_uploaded_file(profile_image, profile_image_name)
						profile_image = profile_image_name
					else:
						profile_image=None

					new_profile = UserProfile(
						user_id=user,
						address=request.POST['address'],
						phone_number=request.POST['phone_number'],
						about_me=request.POST['about_me'],
						latitude=22.57,
						longitude=88.36,
						is_verified=1,
						interest_id_id=1,
						role_id=5,
						profile_image = profile_image,
						# user_id_id=user_id,
						activate_token=activate_no
					)
					new_profile.save()
					
					if 'menu' in request.POST:
						menulist = request.POST.getlist('menu')
						concatmenu=None
						for singlemenu in menulist:
							if concatmenu != None:
								concatmenu = concatmenu + ',' + str(singlemenu)
							else:
								concatmenu = str(singlemenu)
						# concatmenu.strip(',')
						Subadminmenulist(
							user =  new_profile.user_id,
							userprofile =  new_profile,
							pagename    =  concatmenu,

						).save()
					   

					messages.add_message(request, messages.SUCCESS, 'Sub Admin Added Successfully!') 
					return HttpResponseRedirect(reverse('admin_management'))
			# except:
			#     print('vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv')
			#     return render(request, 'adminpanel/addsubadmin.html', {
			#         # 'question': question,
			#         'error_message': "You didn't select a choice.",
			#     })
		else :
			return render(request,'adminpanel/addsubadmin.html')
	else :
		return HttpResponseRedirect('/admin/login/') 

def edit_subadmin(request, subadminId):
	if 'admin_id' in request.session:
		if request.method == 'POST' :
				print(request.POST)
				uprofile_update        = UserProfile.objects.get(id=subadminId)
				uprofile_update.user_id.first_name   = request.POST['firstname']
				uprofile_update.user_id.last_name   = request.POST['lastname']
				uprofile_update.user_id.email   = request.POST['email']
				uprofile_update.user_id.username   = request.POST['username']

				if User.objects.filter(Q(username=request.POST['username']) | Q(email=request.POST['email']) ).exclude(id=uprofile_update.user_id.id).exists():
					messages.error(request, 'Email or Username already exist!.')
					return render(request, 'adminpanel/editsubadmin.html',
					{
						'userProfileObj':uprofile_update
						
					})
				uprofile_update.user_id.save()

				if len(request.FILES) != 0:
					profile_image = request.FILES['profile_image']
					profile_image_name = str(profile_image).replace(' ', '-')
					profile_image_name = 'inteer-'+str(uprofile_update.user_id.id)+'-'+ str(profile_image_name)
					handle_uploaded_file(profile_image, profile_image_name)
					profile_image = profile_image_name
					uprofile_update.profile_image  = profile_image

				uprofile_update.address  = request.POST['address']
				uprofile_update.phone_number   = request.POST['phone_number']
				uprofile_update.latitude = request.POST['latitude']
				uprofile_update.longitude  = request.POST['longitude']
				uprofile_update.about_me= request.POST['about_me']
				uprofile_update.physical_ability   = request.POST['physical_ability']
				if 'is_verified' in request.POST  and request.POST['is_verified'] == 'verified':
				   verified=1
				else:
					verified=0
				uprofile_update.is_verified  = verified
				uprofile_update.save()
				
				concatmenu=None
				if 'menu' in request.POST:
					menulist = request.POST.getlist('menu')
					for singlemenu in menulist:
						if concatmenu != None:
							concatmenu = concatmenu + ',' + str(singlemenu)
						else:
							concatmenu = str(singlemenu)
					# concatmenu.strip(',')
				if Subadminmenulist.objects.filter(userprofile=uprofile_update ).exists():
					subadminmenulistObj = Subadminmenulist.objects.get(userprofile=uprofile_update )
					subadminmenulistObj.pagename = concatmenu
					subadminmenulistObj.save()
				else:
					Subadminmenulist(
						user =  uprofile_update.user_id,
						userprofile =  uprofile_update,
						pagename    =  concatmenu,
					).save()
					
				messages.add_message(request, messages.SUCCESS, 'Sub Admin Updated Successfully!') 
				return HttpResponseRedirect(reverse('admin_management'))
		else:
			userProfileObj = UserProfile.objects.get(id=subadminId)
			menudict = {'usersection':None ,'organizations':None }
			if Subadminmenulist.objects.filter(userprofile=userProfileObj ).exists():
				submenulistObj = Subadminmenulist.objects.get(userprofile=userProfileObj )
				if submenulistObj.pagename != None :
					menulist = list(submenulistObj.pagename.split(","))
					for menu in menulist:
						if menu == 'usersection':
							menudict['usersection'] = 'checked'
						elif menu == 'organizations':
							menudict['organizations'] = 'checked'
			else:
				menulist=None
			return render(request, 'adminpanel/editsubadmin.html', {'userProfileObj':userProfileObj,'menudict' : menudict })
	else :
		return HttpResponseRedirect('/admin/login/')

def delete_subadmin(request,subadminId):
	userProfileObj = UserProfile.objects.get(id=subadminId)
	userProfileObj.user_id.is_active=False
	userProfileObj.user_id.save()
	userProfileObj.delete()
	# del_id = int(userProfileObj.user_id.id)
	# userProfileObj.user_id.delete()
	# print(userProfileObj.user_id)
	# u = User.objects.filter(id=771)
	# u.is_active = False
	# print(dir(u))
	# u.save()


	# print(u)
	# User.objects.filter(id=del_id).delete()
	# User.objects.filter(id=userProfileObj.user_id.id).delete()
	messages.add_message(request, messages.SUCCESS, 'Sub Admin deleted Successfully!') 
	return HttpResponseRedirect(reverse('admin_management'))

def handle_uploaded_file(file, filename):
	if not os.path.exists('media/profileimage'):
		path = os.path.join('/media', 'profileimage') 
		os.mkdir(path)

	with open('media/profileimage/' + filename, 'wb+') as destination:
		for chunk in file.chunks():

			destination.write(chunk)


def show_product(request, product_id, product_slug):
	product = get_object_or_404(Product, id=product_id)

	if request.method == 'POST':
		add_item_to_cart(request)
	#     form = CartForm(request, request.POST)
	#     if form.is_valid():
	#         request.form_data = form.cleaned_data
	#         cart.add_item_to_cart(request)
	#         return redirect('show_cart')

	# form = CartForm(request, initial={'product_id': product.id})
	return render(request, 'adminpanel/stripe_product_detail.html', {'product': product,})

def add_item_to_cart(request):
	# cart_id = _cart_id(request)

	product_id = request.form_data['product_id']
	quantity = request.form_data['quantity']

	p = get_object_or_404(Product, id=product_id)

	price = p.price

	cart_items = get_all_cart_items(request)

	item_in_cart = False

	for cart_item in cart_items:
		if cart_item.product_id == product_id:
			cart_item.update_quantity(quantity)
			# cart_item.save()
			item_in_cart = True

	if not item_in_cart:
		item = CartItem(
			cart_id = _cart_id(request),
			price = price,
			quantity = quantity,
			product_id = product_id,
		)

		# item.cart_id = cart_id
		item.save()

# def stripehome(request):
# 	if request.user.is_authenticated:
# 		key = settings.STRIPE_PUBLISHABLE_KEY
# 		o = datetime.fromtimestamp(1574407637)
# 		q= '{}-{}-{}'.format(o.year, o.month, o.day)
# 		print(q)
# 		userObj = User.objects.get(id=373)
# 		user_profile = UserProfile.objects.get(user_id=userObj)
# 		user_profile.subscribed = True
# 		subscription_end_date = datetime.fromtimestamp(int(return_dict['subs']['current_period_end']))
# 		user_profile.subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
# 		user_profile.save()
# 		print(user_profile)
# 		return render(request,'adminpanel/stripehome.html',{'key': key})

# 	else:
# 		messages.error(request, 'Please Login!') 
# 		return HttpResponseRedirect(reverse('admin_login'))

# def charge(request): # new
# 	planList = settings.STRIPE_PLAN_ID_DICT
# 	if request.method == 'POST':
# 		if request.user.is_authenticated:
# 			try:
# 				user = PayementInformations.objects.get(user=request.user,plan_id=planList[0][0])
# 			except PayementInformations.DoesNotExist:
# 				user = None
# 			if user == None:
# 				return_dict = payment_processing_execution(request)
# 				pay = return_dict['pay']
				
# 				if pay['status'] == 'paid' and pay['paid']==True:
# 					insert_to_payment_table = PayementInformations(
# 						user        = request.user,
# 						userprofile = UserProfile.objects.get(user_id=request.user),
# 						volunteer_number    = return_dict['subs']['quantity'],
# 						customer_id  = return_dict['cust']['id'],
# 						email =request.user.email,
# 						token =request.POST['stripeToken'],
# 						plan_id =return_dict['plan_response']['id'],
# 						plan_type = 0,
# 						package = 0,
# 						trial_period =return_dict['trial_period_days'],
# 						subscription_id = return_dict['subs']['id'],
# 						subscription_start_date = return_dict['subs']['current_period_start'],
# 						subscription_end_date = return_dict['subs']['current_period_end'],
# 						subscribed = True,
# 						subscription_response = return_dict['subs'],
# 						plan_response = return_dict['plan_response'],
# 						product_response = return_dict['product_response'],
# 					)
# 					insert_to_payment_table.save()
# 					messages.add_message(request, messages.SUCCESS, 'you have paid $'+str(request.POST['card-amount'])) 
# 					return HttpResponseRedirect(reverse('stripehome'))
# 				else:
# 					return render(request, 'adminpanel/charge.html',context={"paid": 0})
# 					# pay['status'],pay['paid'],paid['invoice_pdf'],pay['']
# 			else:
# 				dt_object = datetime.fromtimestamp(int(user_related_plan_info.subscription_end_date))

# 				if(dt_object <= datetime.now()  ):
# 					messages.error(request, 'Repayment to the same plan, same subscribe') 
# 					return HttpResponseRedirect(reverse('stripehome'))
# 				else:
# 					messages.error(request, 'You have already subscribe to the plan') 
# 					return HttpResponseRedirect(reverse('stripehome'))
# 		else:
# 			messages.error(request, 'Please Login!') 
# 			return HttpResponseRedirect(reverse('admin_login'))
# 	else:
# 		return render(request, 'adminpanel/charge.html',context={"paid": 0})

def testfn(request):
	import json
	return HttpResponse(1)


@csrf_exempt
def stripewebhook(request):
	import json
	payload = request.body
	print('bbbbbbbbbbbbbbbbb')
	# print(payload)
	# my_json = payload.decode('utf8').replace("'", '"')
	my_json = payload.decode('utf8')
	# print(my_json)
	# data = json.loads(my_json)
	# print(data)
	event = None
	# return render(request, 'adminpanel/stripewebhook.html',context={"response": 400})
	try:
		event = stripe.Event.construct_from(json.loads(my_json), stripe.api_key)
		if event.type == 'customer.subscription.updated':
			sub_evt = event.data.object
			sub_evt_json =  json.loads(json.dumps(sub_evt))
			subscription_status = 0
			package = SubscriptionPlanDetails.objects.filter(plan_id = sub_evt_json['plan']['id']).first()
			# package = SubscriptionPlanDetails.objects.filter(plan_id = 'plan_GHnoTOwhdFrBMp').first()
			product_response  = stripe.Product.retrieve(sub_evt_json['plan']['product'])
			subscribedObjList  = PayementInformations.objects.filter(subscription_id=sub_evt_json['id'])
			print('package---------------')
			print(sub_evt_json['plan']['id'])
			print(package)
			print(len(subscribedObjList))
			if len(subscribedObjList) >= 1 and (package)!=None:
				chk_subscribed_exist = PayementInformations.objects.filter(subscription_id=sub_evt_json['id'],subscription_start_date=sub_evt_json['current_period_start'],subscription_end_date=sub_evt_json['current_period_end'])
				print(len(chk_subscribed_exist))
				if len(chk_subscribed_exist) == 0:
					userObj = 0
					token = 0
					plan_type = 0
					volunteer_number = 0
					for obj in subscribedObjList :
						obj.subscription_deleted = 1
						obj.subscription_ended = True
						obj.subscribed = False
						obj.save()
						userObj = obj.user
						token = obj.token
						plan_type = obj.plan_type
						volunteer_number = obj.volunteer_number
						print('ok')
					
					subscription_renew = PayementInformations(
						user        = userObj,
						userprofile = UserProfile.objects.get(user_id=userObj),
						volunteer_number    = volunteer_number,
						customer_id  = sub_evt_json['customer'],
						email =userObj.email,
						token =token,
						plan_id =sub_evt_json['plan']['id'],
						# plan_id ='plan_GHnoTOwhdFrBMp',
						plan_type = package.plan,
						package = package.package,
						subscription_id = sub_evt_json['id'],
						subscription_start_date = sub_evt_json['current_period_start'],
						subscription_end_date = sub_evt_json['current_period_end'],
						subscribed = True,
						subscription_response = sub_evt_json,
						plan_response = sub_evt_json['plan'],
						product_response = product_response,
						subscription_plan_details_id = package
					)
					subscription_renew.save()

					subscription_start_date = datetime.fromtimestamp(int(sub_evt_json['current_period_start']))
					subscription_end_date = datetime.fromtimestamp(int(sub_evt_json['current_period_end']))
					user_profile = UserProfile.objects.get(user_id=userObj)
					user_profile.subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
					user_profile.subscribed = True
					user_profile.save()

					if subscription_renew.subscription_deleted == SubcriptionDeletion.ACTIVE:
						subscription_status = 0
					elif subscription_renew.subscription_deleted == SubcriptionDeletion.DELETED:
						subscription_status = 1
					elif subscription_renew.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
						subscription_status = 3
					request.session['package'] = subscription_renew.package
					request.session['plan_type'] = subscription_renew.plan_type
					request.session['trial_period'] = subscription_renew.trial_period
					request.session['subscription_ended'] = subscription_renew.subscription_ended
					request.session['subscription_deleted'] = subscription_status
					
		
		return render(request, 'adminpanel/stripewebhook.html',context={"response": 200})
		# return HttpResponse(status=200)
	except ValueError as e:
		
		return HttpResponse(status=400)
	return HttpResponseRedirect('/admin/profile/')
	# Handle the event
	if event.type == 'customer.subscription.updated':
		payment_intent = event.data.object # contains a stripe.PaymentIntent
		print('finally webhook done')
		return render(request, 'adminpanel/stripewebhook.html',context={"response": 200})
	else:
		return render(request, 'adminpanel/stripewebhook.html',context={"response": 400})
		

def upload_all_csv(request):
	organizations = Organization.objects.filter(user_id=42)
	if 'admin_id' in request.session:
		return render(request, 'adminpanel/excel_organisation.html', {'organization': organizations,})
	return render(request, 'adminpanel/excel_organisation.html', {'organization': organizations,})
	


def  upload_organization_opportunity(request):
	import csv 
	import json
	import io
	import openpyxl
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST' and 'admin_id' in request.session :
		data = []
		data_result = []
		all_volunteer ={}
		# user = User.objects.get(id=42)
		# user = User.objects.get(id=request.session['admin_id'])
		getfile = request.FILES['csv_file']
		splitted_value = getfile.content_type.split('/')[1].split(".")
		sheet_type = splitted_value[len(splitted_value)-1]
		print('kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
		print(sheet_type)
		print('kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk') 
		if sheet_type == 'sheet':
			wb = openpyxl.load_workbook(getfile)
			for sheetName in wb.sheetnames:
				if sheetName == 'Organisation':
					ws = wb.get_sheet_by_name('Organisation')
					for index, row in enumerate(ws.iter_rows(),start=2):
						c=1 
						cell_dict={}
						for cell in row:
							cell_dict[ws.cell(row=1, column=c).value] = ws.cell(row=index, column=c).value
							c+=1 
						# print(cell_dict)  
						if 'Organization Name(Youth development)' in cell_dict and 'Location' in cell_dict :
							if cell_dict['Organization Name(Youth development)'] != None and cell_dict['Location'] != None :
								organization_name = cell_dict['Organization Name(Youth development)']
								address 		  = cell_dict['Location']
								  
								org_obj = Organization.objects.filter(organization_name = organization_name,address=address).first()
								if org_obj != None:
									excel_org_save(org_obj,cell_dict,0)
								else:
									new_org_obj = Organization.objects.create(organization_name = organization_name,address=address,parent_id=0)
									excel_org_save(new_org_obj,cell_dict,1)
				if sheetName == 'Opportunity':
					org_obj = None
					ws = wb.get_sheet_by_name('Opportunity')
					for index, row in enumerate(ws.iter_rows(),start=2):
						c=1 
						cell_dict={}
						for cell in row:
							cell_dict[ws.cell(row=1, column=c).value] = ws.cell(row=index, column=c).value
							c+=1
						print(cell_dict)
						print('balllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllllll')
						if 'Organization Name(Youth development)' in cell_dict and 'Coordinator Name' in cell_dict:
							if cell_dict['Organization Name(Youth development)'] != None :
								organization_name = cell_dict['Organization Name(Youth development)']
								address 		  = cell_dict['Location']
								user_email 		  = cell_dict['Coordinator Name']
								cause 		  	  = cell_dict['Cause']
							
								org_obj = Organization.objects.filter(organization_name = organization_name,address=address).first()

								if org_obj == None :
									org_obj = Organization.objects.create(organization_name = organization_name,address=address,parent_id=0,cause = cause)
								print(org_obj)
								print('kokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokokok')
								if org_obj != None :
									## chk user existance ##
									try:
										validate_email(user_email)
										user_collection = User.objects.filter(email=user_email)
										if len(user_collection) > 0 :
											user = user_collection.last()
										else:  # create user and user profile
											user = new_user_create(user_email)
										# chk if the user is already a co-ordinator ##
										already_co_ordinator_user = CordinatorRequest.objects.filter(user_id=str(user.id),role=3,status='Approved')
										if len(already_co_ordinator_user) > 0 :
											print('heeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee')
											co_ordinator_user = already_co_ordinator_user.last()
										else :
											print('theeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee')
											UserProfile.objects.filter(user_id_id=str(user.id)).update(role_id=3)
											co_ordinator_user = CordinatorRequest(user_id=user,org_id_id=str(org_obj.id),role=3,status='Approved',is_request='coordinator',oppurtunity_id=0).save()
										
										## check if the co ordinator is the co ordinator of current organization, if yes then add opportunity else pass
										if co_ordinator_user.org_id.id == org_obj.id:
											add_opportunity = add_opportunity_by_coordinator(user,org_obj,cell_dict) ## add opportunity
									
										# chk if the user is already a co-ordinator ##
									except ValidationError :
										pass 
									## chk user existance ##
			organization = Organization.objects.all()
			messages.add_message(request, messages.SUCCESS, 'Import successfully') 
			return render(request, 'adminpanel/excel_organisation.html', {'organization': organization,'msg':'success'})
		else:
			organization = Organization.objects.all()
			messages.add_message(request, messages.ERROR, 'xcel file only to be imported') 
			return render(request, 'adminpanel/excel_organisation.html', {'organization': organization,'msg':'error'})

def add_opportunity_by_coordinator(user,org_obj,cell_dict):
	author_name = user.first_name+' '+ user.last_name
	no_of_volunteers = 1
	if type(cell_dict['Number of Volunteers Needed']) == int and cell_dict['Number of Volunteers Needed'] > 0 :
		no_of_volunteers = cell_dict['Number of Volunteers Needed']

	opportunitiy = Opportunities(
		user_id = user,
		org_id = org_obj,
		opportunity_name = cell_dict['Opportunity Name'],
		description = cell_dict['Opportunity Name'],
		author_name = author_name,
		address = cell_dict['Location'],
		parent_id = 0,
		lon = 88.4567,
		lat = 22.2345,
		no_ofyear = 5,
		parent_opportunity = 0,
		start_date = timezone.now(),
		end_date = timezone.now(),
		no_of_volunteers= no_of_volunteers
	)
	opportunitiy.save()

def new_user_create(user_email):
	user = User.objects.create_user(
		password='123456',
		is_superuser=False,
		username=user_email,
		first_name='first name',
		last_name='last name',
		email=user_email,
		is_staff=False,
		is_active=True
	)
	user.save()
	user_id=user.id

	user_info = UserProfile(
		address='',
		latitude=22.57,
		longitude=88.36,
		is_verified=0,
		interest_id_id=1,
		role_id=2, # volunteer 2
		user_id_id=user_id,
		activate_token='123456',
	)
	user_info.save() 
	return user



def json_upload_organization(request):
	columns = [               # atatable column index  => database column name
				'cause',
				'organization_name',
				'email',
				'phone',
			]
			# atatable column index  => database column name
	columns1 = [               
		'cause',
		'organization_name',
		'email',
		'phone',
	]
	cursor = connection.cursor()
	requestData = request.GET
	total = Organization.objects.all().count()
	sql = "SELECT cause, organization_name, email, phone "
	sql =sql + " FROM adminpanel_organization  "
	
	if requestData.get('search[value]') :
		sql =sql + " WHERE ( cause LIKE  '%" + str(requestData.get('search[value]')) + "%'"
		sql =sql + " OR organization_name LIKE '%" + str(requestData.get('search[value]')) + "%'"
		sql =sql + " OR email LIKE '%" + str(requestData.get('search[value]')) + "%'"
		sql =sql + " OR CAST(phone AS text) LIKE '%" + str(requestData.get('search[value]')) + "%')"
	print(sql)	
	cursor.execute(sql)
	totalFiltered = cursor.rowcount
				
	sql = sql + " ORDER BY  "+str( columns[int(requestData.get('order[0][column]')) ] ) +" "+ str(requestData.get('order[0][dir]')) +" LIMIT " + str(requestData.get('length')) + " OFFSET " + str(requestData.get('start')) + "  " 
	cursor.execute(sql)
	# print(sql)
	ordering_direction = str( columns1[int(requestData.get('order[0][column]')) ] )
	if(str(requestData.get('order[0][dir]'))=='desc'):
		ordering_direction='-'+ordering_direction

	UserProfileCollections = cursor.fetchall() 
	data = []
	for singleVal in UserProfileCollections :
		print(singleVal)
		singeList = []
		singeList.append(singleVal[0])
		singeList.append(singleVal[1])
		singeList.append(singleVal[2])
		singeList.append(singleVal[3] ) 
		data.append(singeList)
	usercollection_dict = {}
	usercollection_dict['draw'] = requestData.get('draw')
	usercollection_dict['recordsTotal'] = total
	usercollection_dict['recordsFiltered'] = totalFiltered
	usercollection_dict['data'] = data
	return JsonResponse(usercollection_dict)

def excel_org_save(org_obj,cell_dict,new_reqest):
	# org_obj.parent_id = org_obj.parent_id

	if cell_dict['Cause'] != None and cell_dict['Cause'] != 'NA' :
		org_obj.cause = cell_dict['Cause']

	if cell_dict['Location'] != None and cell_dict['Location'] != 'NA' :
		org_obj.address = cell_dict['Location']
	
	if cell_dict['Email'] != None and cell_dict['Email'] != 'NA' :
		org_obj.email = cell_dict['Email']
	
	if cell_dict['IRS rank'] != None and cell_dict['IRS rank'] != 'NA' :
		org_obj.irs_rank = cell_dict['IRS rank']
	
	if cell_dict['Website Link'] != None and cell_dict['Website Link'] != 'NA' :
		org_obj.web_url = cell_dict['Website Link']

	if cell_dict['Location1'] != None and cell_dict['Location1'] != 'NA' :
		org_obj.address1 = cell_dict['Location1']

	if cell_dict['Contact Number2'] != None and cell_dict['Contact Number2'] != 'NA' :
		org_obj.phone2 = cell_dict['Contact Number2']
	
	if cell_dict['FB Page'] != None and cell_dict['FB Page'] != 'NA' :
		org_obj.fb_url = cell_dict['FB Page']

	if cell_dict['EIN'] != None and cell_dict['EIN'] != 'NA' :
		org_obj.tax_id = cell_dict['EIN']
	
	if cell_dict['Who we are?'] != None and cell_dict['Who we are?'] != 'NA' :
		org_obj.about_us = cell_dict['Who we are?']
	
	if cell_dict['Event 1'] != None and cell_dict['Event 1'] != 'NA' :
		org_obj.event1 = cell_dict['Event 1']
	
	if cell_dict['Event 2'] != None and cell_dict['Event 2'] != 'NA' :
		org_obj.event2 = cell_dict['Event 2']
	
	if cell_dict['Event 3'] != None and cell_dict['Event 3'] != 'NA' :
		org_obj.event3 = cell_dict['Event 3']
	
	if cell_dict['Event 3'] != None and cell_dict['Event 3'] != 'NA' :
		org_obj.event3 = cell_dict['Event 3']

	if cell_dict['Submission Date'] != None and cell_dict['Submission Date'] != 'NA' :
		org_obj.submission_date = cell_dict['Submission Date']

	if cell_dict['why should we valunteer with you?'] != None and cell_dict['why should we valunteer with you?'] != 'NA' :
		org_obj.why_us = cell_dict['why should we valunteer with you?']

	if cell_dict['What we do?'] != None and cell_dict['What we do?'] != 'NA' :
		org_obj.our_activity = cell_dict['What we do?'] 

	org_obj.save() 

def usersubscription_list(request):
	# subscription_collection = PayementInformations.objects(~Q(package='unlock'))
	if 'admin_id' in request.session:
		return render(request,'adminpanel/subscriptioncollection.html',{"subscription_collection": 1})
	else :
		return HttpResponseRedirect('/admin/login/')

def json_subscription_list(request):
	
	from datetime import datetime
	columns = [               # atatable column index  => database column name
		'email',
		'subscription_start_date',
		'subscription_end_date',
		'admin_chane_subcription_date',
		'operations',
	]
	# atatable column index  => database column name
	columns1 = [               
		'email',
		'subscription_start_date',
		'subscription_end_date',
		'admin_chane_subcription_date',
		'operations'
	]
	cursor = connection.cursor()
	requestData = request.GET
	total = Organization.objects.all().count()
	distinct_emails  = PayementInformations.objects.values('email').distinct()
	# print(distinct_emails)
	print('hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh')
	deleted_ids =''
	for single_email in distinct_emails :
		single_email_collection = PayementInformations.objects.filter(Q(email=single_email['email']) , ~Q(package='unlock') ,Q(subscription_ended=True) ).order_by('-id')
		for emailObj in single_email_collection :
			if emailObj.subscription_deleted == SubcriptionDeletion.DELETED:
				deleted_ids+=str(emailObj.id)+','
				break
	deleted_ids = deleted_ids.rstrip(',')
	print(deleted_ids)
	sql = "SELECT id,email, subscription_start_date,subscription_end_date,admin_chane_subcription_date "
	if deleted_ids != '':
		sql =sql + " FROM adminpanel_payementinformations  Where id in (" +deleted_ids+ ") "
	else:
		sql =sql + " FROM adminpanel_payementinformations  Where package!='unlock' and subscription_deleted = '3' "

	print(sql)
	print('mmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmm')

	# sql =sql + " FROM adminpanel_payementinformations  Where package!='unlock' and subscription_deleted = '3' "
	
	if requestData.get('search[value]') :
		sql =sql + " AND ( email LIKE  '%" + str(requestData.get('search[value]')) + "%'"
		sql =sql + " OR subscription_start_date LIKE '%" + str(requestData.get('search[value]')) + "%'"
		sql =sql + " OR subscription_end_date LIKE '%" + str(requestData.get('search[value]')) + "%')"
	print(sql)	
	cursor.execute(sql)
	totalFiltered = cursor.rowcount
				
	sql = sql + " ORDER BY  "+str( columns[int(requestData.get('order[0][column]')) ] ) +" "+ str(requestData.get('order[0][dir]')) +" LIMIT " + str(requestData.get('length')) + " OFFSET " + str(requestData.get('start')) + "  " 
	cursor.execute(sql)
	# print(sql)
	ordering_direction = str( columns1[int(requestData.get('order[0][column]')) ] )
	if(str(requestData.get('order[0][dir]'))=='desc'):
		ordering_direction='-'+ordering_direction

	UserProfileCollections = cursor.fetchall() 
	data = []
	for singleVal in UserProfileCollections :
		print(singleVal[2])
		subscription_start_date = datetime.fromtimestamp(int(singleVal[2]))
		subscription_end_date = datetime.fromtimestamp(int(singleVal[3]))
		final_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
		final_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
					
		singeList = []

		singeList.append(singleVal[1])
		singeList.append(final_start_date)
		singeList.append(final_end_date)
		singeList.append(singleVal[4])
		singeList.append('<a href="'+reverse("edit_subscription", args=[singleVal[0]])+'">Edit</a>')
		data.append(singeList)
	usercollection_dict = {}
	usercollection_dict['draw'] = requestData.get('draw')
	usercollection_dict['recordsTotal'] = total
	usercollection_dict['recordsFiltered'] = totalFiltered
	usercollection_dict['data'] = data
	return JsonResponse(usercollection_dict)

def getTimestamp(getDateString):
	from datetime import datetime
	final_st_dt = datetime.strptime(getDateString, "%Y-%m-%d")
	final_st_dt_timestamp = datetime.timestamp(final_st_dt)
	return str(final_st_dt_timestamp).rsplit('.', 2)[0] 

def edit_subscription(request,subscriptionId):

	if 'admin_id' in request.session:
		from datetime import datetime
		paymentObj = PayementInformations.objects.get(id=subscriptionId)
		
		if request.method == 'POST' :
			startdate = request.POST['startdate']
			# enddate   = request.POST['enddate']
			
			if startdate.find('/') != -1 :
				start_array = startdate.split('/')
				st_dt = start_array[1]+'-'+start_array[0]+'-01' #year-month-day
				final_st_dt_timestamp = getTimestamp(st_dt)
				paymentObj.subscription_start_date =  final_st_dt_timestamp
				paymentObj.admin_chane_subcription_date =  True 
				paymentObj.save()
			
			# if enddate.find('/') != -1 :
			# 	start_array = enddate.split('/')
			# 	mnth = start_array[0]
			# 	yr   = start_array[1]
			# 	end_day = calendar.monthrange(int(yr),int(mnth))[1]
			# 	st_dt = start_array[1]+'-'+start_array[0]+'-'+str(end_day) #year-month-day
			# 	final_st_dt_timestamp = getTimestamp(st_dt)
			# 	paymentObj.subscription_end_date =  final_st_dt_timestamp
			# 	paymentObj.admin_chane_subcription_date =  True 
			# 	paymentObj.save()


				
			messages.add_message(request, messages.SUCCESS, 'User subscription Admin Updated Successfully!') 
			return HttpResponseRedirect(reverse('usersubscription_list'))
		else:
			subscription_start_date = datetime.fromtimestamp(int(paymentObj.subscription_start_date))
			subscription_end_date = datetime.fromtimestamp(int(paymentObj.subscription_end_date))
			final_subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
			final_subscription_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
			print(type(subscription_start_date))
			print(final_subscription_start_date)
			print(final_subscription_end_date)
			return render(request, 'adminpanel/edit_subscription.html', { 'paymentObj':paymentObj, 'subscription_start_date' : final_subscription_start_date , 'subscription_end_date' : final_subscription_end_date })
	else :
		return HttpResponseRedirect('/admin/login/')


''' def subadmincheck(request):
	if request.is_ajax():
		print(request.POST)
		message = 1
	else:
		message = 0
	return HttpResponse(message)'''
'''def csv_import_export_test(request):
	import csv
	import psycopg2
	if request.method == 'POST':
		f = request.FILES['csv_input']
		decoded_file = f.read().decode('utf-8').splitlines()
		reader = csv.DictReader(decoded_file)
		data = []
		for row in reader:
			data.append(row)
		return render(request,'adminpanel/csv_import_export_test.html', {"filedata": data})
		conn = psycopg2.connect("host=138.68.12.41 dbname=db_inter user=social_django_user password=Host@123456")
		cur = conn.cursor()
		with open('/var/www/html/inteer/djangogirls/webservice/Master_Sheet_All_new.csv', 'r') as f:
			reader = csv.reader(f)
			next(reader)  # Skip the header row.
			for row in reader:
				cur.execute("insert into adminpanel_organization (organization_name, parent_id, address, tax_id, email, web_url, phone, cause, about_us, our_activity, why_us, photo, address1, event1, event2, event3, fb_url, irs_rank, phone2, submission_date) values ('"+row[0]+"','"+row[1]+"','"+row[2]+"','"+row[3]+"','"+row[4]+"','"+row[5]+"','"+row[6]+"','"+row[7]+"','"+row[8]+"','"+row[9]+"','"+row[10]+"','"+row[11]+"','"+row[12]+"','"+row[13]+"','"+row[14]+"','"+row[15]+"','"+row[16]+"','"+row[17]+"','"+row[18]+"','"+row[19]+"')")
		conn.commit()
		return render(request,'adminpanel/csv_import_export_test.html')'''
