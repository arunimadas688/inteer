from django.shortcuts import render
from django.utils import timezone,formats
from django.template import RequestContext,Context, Template, Context, loader
from django.shortcuts import render_to_response
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.views.decorators.csrf import csrf_protect,csrf_exempt
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.contrib.auth.models import User
from adminpanel.models import *
from django.db import connection
from django.db.models import Avg
from datetime import datetime, timedelta, date
import PIL
from PIL import Image
import os
from os.path import basename
from django.conf import settings
from django.db.models import Q
from django.db import connection
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Max
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.contrib import messages
import re
from django.core.files.storage import FileSystemStorage
from xlsxwriter.workbook import Workbook
from django.utils.html import strip_tags
from decimal import Decimal
ACCESS_CONTROL_ALLOW_ORIGIN = '*'
import _thread
# import sched, time 
import json
import time
import stripe
from django.core import serializers
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import ast 

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser,FileUploadParser
from rest_framework.decorators import api_view
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.utils.urls import replace_query_param, remove_query_param
from rest_framework.parsers import JSONParser
from rest_framework.pagination import PageNumberPagination
from rest_framework import mixins
from rest_framework import generics
from rest_framework import status
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_jwt.settings import api_settings
from rest_framework_simplejwt.tokens import RefreshToken
stripe.api_key = settings.STRIPE_SECRET_KEY

# Define a function for the thread
# def print_time( threadName, delay):
#    count = 0
#    while count < 5:
#       time.sleep(delay)
#       count += 1
#       print( threadName, time.ctime(time.time()) ) 

# # Create two threads as follows
# try:
#    _thread.start_new_thread( print_time, ("Thread-1", 2, ) )
# except:
#    print("Error: unable to start thread") 

# while 1:
#    pass 
class HelloView(APIView):
	permission_classes = (IsAuthenticated,)

	def get(self, request) :
		content = {'message': 'Hello, World!'}
		return Response(content)

	def post(self,request) :
		content = {'message': 1}
		return HttpResponse(1)

class MyTokenObtainPairView(TokenObtainPairView):

	def post(self,request):
		import json
		import base64
		from django.core.serializers.json import DjangoJSONEncoder
		# if request.method == 'POST':
		rtn_obj = {}
		username = request.POST.get('username', None)
		password = request.POST.get('password', None)
		first_decode = base64.b64decode(password)
		# return HttpResponse(str(first_decode))
		second_step = str(first_decode, 'utf-8')
		splitdata = second_step.split('----')
		second_decode = base64.b64decode(splitdata[0])
		password_decrypted = str(second_decode, 'utf-8')

		try:
			user = authenticate(username=username, password=password_decrypted)
			jwt_payload_handler = api_settings.JWT_PAYLOAD_HANDLER
			jwt_encode_handler = api_settings.JWT_ENCODE_HANDLER

			payload = jwt_payload_handler(user)
			token = jwt_encode_handler(payload)
			refresh = RefreshToken.for_user(user)
			return JsonResponse( {'refresh': str(refresh),'access': str(refresh.access_token) } )
		except:
			return JsonResponse({"detail": "No active account found with the given credentials"})
	
##########Done5#############
def dictfetchall(cursor):
	"Return all rows from a cursor as a dict"
	columns = [col[0] for col in cursor.description]
	return [
		dict(zip(columns, row))
		for row in cursor.fetchall()
	]
#####################Done#####################
def home_index(request):
	import json
	imageurl =settings.IMAGE_URL
	banner_9 = {}
	banner_10 = {}
	banner_12 = {}
	banner_13 = {}
	all_banner = []
	# cursor = connection.cursor()
	# cursor.execute("select banner_image,banner_title,banner_text,banner_logo from adminpanel_banner where id=8")
	# banner = dictfetchall(cursor)
	# cursor.execute("select banner_image,title,short_description from adminpanel_cms where id IN(1,2,3)")
	# cms = dictfetchall(cursor)
	# cursor.execute("select banner_image,title,short_description from adminpanel_cms where id=5")
	# whyinter = dictfetchall(cursor)
	# cursor.execute("select banner_image from adminpanel_cms where id IN(6,7,8)")
	# bannerlowercms = dictfetchall(cursor)

	banner_9_fetch = Banner.objects.filter(id=9).values('banner_image','banner_title', 'banner_text','banner_logo')
	banner_9['banner_image'] = banner_9_fetch[0]['banner_image']
	banner_9['banner_title'] = banner_9_fetch[0]['banner_title']
	banner_9['banner_text'] = banner_9_fetch[0]['banner_text']
	banner_9['banner_logo'] = banner_9_fetch[0]['banner_logo']

	banner_10_fetch = Banner.objects.filter(id=10).values('banner_image','banner_title', 'banner_text','banner_logo')
	banner_10['banner_image'] = banner_10_fetch[0]['banner_image']
	banner_10['banner_title'] = banner_10_fetch[0]['banner_title']
	banner_10['banner_text'] = banner_10_fetch[0]['banner_text']
	banner_10['banner_logo'] = banner_10_fetch[0]['banner_logo']

	banner_12_fetch = Banner.objects.filter(id=12).values('banner_image','banner_title', 'banner_text','banner_logo')
	banner_12['banner_image'] = banner_12_fetch[0]['banner_image']
	banner_12['banner_title'] = banner_12_fetch[0]['banner_title']
	banner_12['banner_text'] = banner_12_fetch[0]['banner_text']
	banner_12['banner_logo'] = banner_12_fetch[0]['banner_logo']

	banner_13_fetch = Banner.objects.filter(id=13).values('banner_image','banner_title', 'banner_text','banner_logo')
	banner_13['banner_image'] = banner_13_fetch[0]['banner_image']
	banner_13['banner_title'] = banner_13_fetch[0]['banner_title']
	banner_13['banner_text'] = banner_13_fetch[0]['banner_text']
	banner_13['banner_logo'] = banner_13_fetch[0]['banner_logo']

	home_banner =  Banner.objects.filter(banner_section = 'Home').values('banner_image','banner_title', 'banner_text','banner_logo')
	print(home_banner)
	if len(home_banner) > 0 :
		for singleBanner in home_banner :
			print(singleBanner)
			banner_dict = {}
			banner_dict['banner_image'] = singleBanner['banner_image']
			banner_dict['banner_title'] = singleBanner['banner_title']
			banner_dict['banner_text'] = singleBanner['banner_text']
			banner_dict['banner_logo'] = singleBanner['banner_logo']
			if len(banner_dict) > 0 :
				all_banner.append(banner_dict)

	if banner_9 and  banner_10 and  banner_12 and  banner_13 :
		data = json.dumps({"Ack":"1", "banner_10":banner_10,"banner_9":banner_9, 'banner_12':banner_12,'banner_13':banner_13,'imageurl':imageurl,'all_banner':all_banner})
		return HttpResponse(data)
	else:
		data = json.dumps({"Ack":"0", "banner_10":'','banner_9':'','banner_12':'','banner_13':'','imageurl':'','all_banner':all_banner})
		return HttpResponse(data)

###############Done################
# def login_user_submit(request):
class login_user_submit(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		import base64
		from django.core.serializers.json import DjangoJSONEncoder
		# if request.method == 'POST':
		rtn_obj = {}
		username = request.POST.get('username', None)
		password = request.POST.get('password', None)
		first_decode = base64.b64decode(password)
		# return HttpResponse(str(first_decode))
		second_step = str(first_decode, 'utf-8')
		splitdata = second_step.split('----')
		second_decode = base64.b64decode(splitdata[0])
		password_decrypted = str(second_decode, 'utf-8')
		# return HttpResponse(1)
		# return HttpResponse(second_decode)
		user = authenticate(username=username, password=password_decrypted)
		# print(user)
		# print('kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
		# return HttpResponse(1)

		if user is not None: 
			login(request, user)
			
			#  check if user subscribed to any plan or not
			user_payment_obj = PayementInformations.objects.filter(user = user).last()
			subscription_status = 0
			if user_payment_obj:
				if user_payment_obj.subscription_deleted == SubcriptionDeletion.ACTIVE:
					subscription_status = 0
				elif user_payment_obj.subscription_deleted == SubcriptionDeletion.DELETED:
					subscription_status = 1
				elif user_payment_obj.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
					subscription_status = 3

				subscription_start_date = datetime.fromtimestamp(int(user_payment_obj.subscription_start_date))
				subscription_end_date = datetime.fromtimestamp(int(user_payment_obj.subscription_end_date))
				
				request.session['package'] = user_payment_obj.package
				request.session['plan_type'] = user_payment_obj.plan_type
				request.session['trial_period'] = user_payment_obj.trial_period
				request.session['subscription_ended'] = user_payment_obj.subscription_ended
				request.session['subscription_deleted'] = subscription_status

				rtn_obj['package'] = user_payment_obj.package
				rtn_obj['plan_type'] = user_payment_obj.plan_type
				rtn_obj['trial_period'] = user_payment_obj.trial_period
				rtn_obj['subscription_ended'] = user_payment_obj.subscription_ended
				rtn_obj['subscription_deleted'] = subscription_status
				rtn_obj['subscription_end_date'] = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
				rtn_obj['subscription_start_date']  = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)




			request.session['member_id'] = user.id
			userData = User.objects.get(id = user.id)
			userpro = UserProfile.objects.get(user_id = user.id)
			
			userprofiles = UserProfile.objects.filter(
				user_id_id=user.id).values('id', 'address', 'phone_number', 'profile_image', 'latitude', 'longitude', 'about_me', 'physical_ability', 'is_verified', 'interest_id_id', 'role_id')
			
			rtn_obj['ack'] = "1"
			rtn_obj['user_id'] = str(user.id)
			rtn_obj['first_name'] = userData.first_name
			rtn_obj['last_name'] = userData.last_name
			rtn_obj['email'] = userData.email
			rtn_obj['role_id'] = userprofiles[0]['role_id']
			rtn_obj['username'] = userData.username
			rtn_obj['user_image'] = userprofiles[0]['profile_image']
			rtn_obj['msg_error'] = " Log In Successfull! "
			rtn_obj['address'] = userprofiles[0]['address']
			# rtn_obj['subscription_id'] = userprofiles[0]['id']
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
		else:
			rtn_obj['ack'] = "0"
			rtn_obj['msg_error'] = " Your username and/or password were incorrect. "
			data = json.dumps(rtn_obj)
			return HttpResponse(data)

	def get(self,request):
		rtn_obj = {}
		rtn_obj['ack'] = "2"
		rtn_obj['msg_error'] = "Error 404"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

	def put(self,request):
		rtn_obj = {}
		rtn_obj['ack'] = "2"
		rtn_obj['msg_error'] = "Error 404"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

	def delete(self,request):
		rtn_obj = {}
		rtn_obj['ack'] = "2"
		rtn_obj['msg_error'] = "Error 404"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done################
def facebook_login(request):
	import json
	rtn_obj = {}

	if request.method == 'POST':
		fbid = request.POST['fbid']
		fname = request.POST['first_name']
		lname = request.POST['last_name']
		username = request.POST['username']
		email = request.POST['email']
		if User.objects.filter(email=email).exists():
			UserData = User.objects.get(email=email)
			fbprofile = UserProfile.objects.get(user_id = UserData.id)
			fbprofile.facebook_id = fbid
			fbprofile.save()
			rtn_obj['ack'] = "1"
			rtn_obj['user_id'] = str(UserData.id)
			rtn_obj['first_name'] = fname
			rtn_obj['last_name'] = lname
			rtn_obj['email'] = email
			rtn_obj['fb_id'] = fbid
			rtn_obj['msg_error'] = " Log In Successfull! "
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
		else :
			if UserProfile.objects.filter(facebook_id=fbid).exists():
				fbprofile = UserProfile.objects.get(facebook_id=fbid)
				UserData = User.objects.get(email=email)
				
				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(UserData.id)
				rtn_obj['first_name'] = fname
				rtn_obj['last_name'] = lname
				rtn_obj['email'] = email
				rtn_obj['fb_id'] = fbid
				rtn_obj['msg_error'] = " Log In Successfull! "
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else :
				user = User.objects.create_user(
					username = username,
					email = email,
					first_name = fname,
					last_name = lname
				)
				user.save()

				userID = user.id

				new_profile = UserProfile(
					user = user,
					facebook_id = fbid
				)
				new_profile.save()
				user_withroles = UserWithroles(
					user = user,
					userroles_id = 2
				)
				user_withroles.save()

				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(userID)
				rtn_obj['first_name'] = fname
				rtn_obj['last_name'] = lname
				rtn_obj['email'] = email
				rtn_obj['fb_id'] = fbid
				rtn_obj['msg_error'] = " Log In Successfull! "
				data = json.dumps(rtn_obj)
				return HttpResponse(data)

###############Done################
def google_login(request):
	import json
	rtn_obj = {}

	if request.method == 'POST':
		gpid = request.POST['gpid']
		fname = request.POST['first_name']
		lname = request.POST['last_name']
		username = request.POST['username']
		email = request.POST['email']
		if User.objects.filter(email = email).exists():
			UserData = User.objects.get(email=email)
			fbprofile = UserProfile.objects.get(user_id = UserData.id)
			fbprofile.google_id = gpid
			fbprofile.save()
			rtn_obj['ack'] = "1"
			rtn_obj['user_id'] = str(UserData.id)
			rtn_obj['first_name'] = fname
			rtn_obj['last_name'] = lname
			rtn_obj['email'] = email
			rtn_obj['gp_id'] = gpid
			rtn_obj['msg_error'] = " Log In Successfull! "
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
		else :
			if UserProfile.objects.filter(google_id = gpid).exists():
				fbprofile = UserProfile.objects.get(google_id = gpid)
				UserData = User.objects.get(email=email)
				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(UserData.id)
				rtn_obj['first_name'] = fname
				rtn_obj['last_name'] = lname
				rtn_obj['email'] = email
				rtn_obj['gp_id'] = gpid
				rtn_obj['msg_error'] = " Log In Successfull! "
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else :
				user = User.objects.create_user(
					username = username,
					email = email,
					first_name = fname,
					last_name = lname
				)
				user.save()

				userID = user.id

				new_profile = UserProfile(
					user = user,
					google_id = gpid
				)
				new_profile.save()
				user_withroles = UserWithroles(
					user = user,
					userroles_id = 2
				)
				user_withroles.save()

				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(userID)
				rtn_obj['first_name'] = fname
				rtn_obj['last_name'] = lname
				rtn_obj['email'] = email
				rtn_obj['gp_id'] = gpid
				rtn_obj['msg_error'] = " Log In Successfull! "
				data = json.dumps(rtn_obj)
				return HttpResponse(data)

###############Done################
class forgot_password_submit(ObtainAuthToken):
	
	def post(self,request):
		import json
		import base64
		rtn_obj = {}
		forgot_mail = request.POST['forgot_mail']
		if User.objects.filter(email=forgot_mail).exists() :
			forgotuser = User.objects.get(email=forgot_mail)
			# return HttpResponse(forgotuser.id)
			# userprofile = UserProfile.objects.get(user_id=forgotuser.id)
			if forgotuser:
				forgot_link = settings.PATH_URL+"/resetpassword/"+str(forgot_mail)
				mailfname = forgotuser.first_name
				forgetpassword_mail=EmailTemplates.objects.get(pk=3)
				t = forgetpassword_mail.templatebody
				t1=t.replace('[NAME]',mailfname,1) 
				t2=t1.replace('[LINK]',forgot_link,1)
				msg_html = t2
				send_mail(forgetpassword_mail.subject, 'hello world again', 'mail@cbnits.com', [forgotuser.email], html_message=msg_html)
				rtn_obj['ack'] = "1"
				rtn_obj['msg_error'] = "Email Send successfully"
			else:
				rtn_obj['ack'] = "0"
				rtn_obj['msg_error'] = "User Does not exist"
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
		else :
			rtn_obj['ack'] = "0"
			rtn_obj['msg_error'] = "Email not Send successfully"
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
	def put(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Email not Send successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def delete(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Email not Send successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def get(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Email not Send successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def patch(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Email not Send successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done################
class reset_password_submit(ObtainAuthToken):
	
	# permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		import base64
		rtn_obj = {}
		forgot_mail = request.POST['forgot_mail']
		newpassword=request.POST['password']
		confirmpassword=request.POST['cpassword']

		first_decode = base64.b64decode(newpassword)
		second_step = str(first_decode, 'utf-8')
		splitdata = second_step.split('----')
		second_decode = base64.b64decode(splitdata[0])
		password_decrypted = str(second_decode, 'utf-8')

		first_decode_confirmpassword = base64.b64decode(confirmpassword)
		second_step_confirmpassword = str(first_decode_confirmpassword, 'utf-8')
		splitdata_confirmpassword = second_step_confirmpassword.split('----')
		second_decode_confirmpassword = base64.b64decode(splitdata_confirmpassword[0])
		confirmpassword_decrypted = str(second_decode_confirmpassword, 'utf-8')

		if password_decrypted != confirmpassword_decrypted :
			rtn_obj['ack'] = "0"
			rtn_obj['msg_error'] = "confirmpassword is not same!"
			data = json.dumps(rtn_obj)
			return HttpResponse(data)
		else :
			user=User.objects.get(email=forgot_mail)
			user.set_password(password_decrypted)
			user.save()
			user_details=User.objects.get(email=forgot_mail)
			user = authenticate(username=user_details.username, password=password_decrypted)
			if user is not None:
				login(request, user)
				request.session['member_id'] = user.id
				userData = User.objects.get(id = user.id)
				userpro = UserProfile.objects.get(user_id = user.id)

				userprofiles = UserProfile.objects.filter(
					user_id_id=user.id).values('id', 'address', 'phone_number', 'profile_image', 'latitude', 'longitude', 'about_me', 'physical_ability', 'is_verified', 'interest_id_id', 'role_id')

				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(user.id)
				rtn_obj['first_name'] = userData.first_name
				rtn_obj['last_name'] = userData.last_name
				rtn_obj['email'] = userData.email
				rtn_obj['username'] = userData.username
				rtn_obj['role_id'] = userprofiles[0]['role_id']
				rtn_obj['user_image'] = userprofiles[0]['profile_image']
				rtn_obj['msg_error'] = " Log In Successfull! "
				rtn_obj['address'] = userprofiles[0]['address']
				# rtn_obj['subscription_id'] = userprofiles[0]['id']
				
				rtn_obj['ack'] = "1"
				rtn_obj['msg_error'] = "Password was succesfully reset"
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else:
				rtn_obj['ack'] = "0"
				rtn_obj['msg_error'] = " Your username and/or password were incorrect. "
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
	def get(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Password Not change successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def put(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Password Not change successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def patch(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Password Not change successfully"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	

###############Done################
def logout_page(request):
	import json
	rtn_obj = {}
	logout(request)
	rtn_obj['ack'] = "1"
	rtn_obj['msg_error'] = "Logout Successfully"
	data = json.dumps(rtn_obj)
	return HttpResponse(data)

###############Done################
class  register_user(ObtainAuthToken):
	
	
	def post(self, request):
		import json
		import base64
		
		rtn_obj = {}
		first_name = request.POST.get('first_name', None)
		last_name = request.POST.get('last_name', None)
		email = request.POST.get('email', None)
		location = request.POST.get('location', None)
		phonenumber = request.POST.get('phonenumber', None)
		cityLat = request.POST.get('cityLat', None)
		role_selected = request.POST.get('role_selected',None)

		 
		try:
			validate_email( email )
			if role_selected == '':
				role_selected = 2
				# return HttpResponse(role_selected)
			else:
				role_selected = 3
				# return HttpResponse(role_selected)
			if cityLat:
				cityLat = cityLat
			else:
				cityLat = 22.572646
			cityLng = request.POST.get('cityLng', None)
			if cityLng:
				cityLng = cityLng
			else:
				cityLng = 88.36389499999996
			password = request.POST.get('password', None)

			first_decode = base64.b64decode(password)
			second_step = str(first_decode, 'utf-8')
			splitdata = second_step.split('----')
			second_decode = base64.b64decode(splitdata[0])
			password_decrypted = str(second_decode, 'utf-8')
			if User.objects.filter(email=email).exists():
				rtn_obj['msg_error'] = "User Exists!"
				rtn_obj['ack'] = "0"
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else :
				import time
				activate_no = int(time.time())
				user = User.objects.create_user(
				password=password_decrypted,
				is_superuser=False,
				username=email,
				first_name=first_name,
				last_name=last_name,
				email=email,
				is_staff=False,
				is_active=True
				)
				user.save()

				#Save userinfo record
				user_info = UserProfile(
					address=location,
					phone_number=phonenumber,
					latitude=cityLat,
					longitude=cityLng,
					is_verified=0,
					interest_id_id=1,
					role_id=role_selected,
					user_id_id=user.id,
					activate_token=activate_no
				)
				user_info.save()
				activation_link = settings.PYTHON_URL+"api/activation/"+str(activate_no)
				mailpwd = request.POST.get('password', None)
				mailfname = request.POST.get('first_name', None)
				register_mail=EmailTemplates.objects.get(pk=1)
				t = register_mail.templatebody
				t1=t.replace('[NAME]',mailfname,1)
				t2=t1.replace('[LINK]',activation_link,1)
				msg_html = t2
				send_mail(register_mail.subject, 'InterApp', 'mail@cbnits.com', [user.email], html_message=msg_html)

				rtn_obj['ack'] = "1"
				rtn_obj['user_id'] = str(user.id)
				rtn_obj['first_name'] = first_name
				rtn_obj['last_name'] = last_name
				rtn_obj['email'] = email
				rtn_obj['role_id'] = role_selected
				rtn_obj['username'] = email
				rtn_obj['user_image'] = ''
				rtn_obj['msg_error'] = " Log In Successfull! "
				rtn_obj['msg_error'] = "Registerd!"
				rtn_obj['ack'] = "1"
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
		except ValidationError:
			rtn_obj['msg_error'] = "Enter a valid email"
			rtn_obj['ack'] = "0"
			data = json.dumps(rtn_obj)
			return HttpResponse(data)

		
	def get(self, request) :
		rtn_obj={}
		rtn_obj['msg_error'] = "Registerd Not Successful!"
		rtn_obj['ack'] = "2"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def put(self, request) :
		rtn_obj={}
		rtn_obj['msg_error'] = "Registerd Not Successful!"
		rtn_obj['ack'] = "2"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def patch(self, request) :
		rtn_obj={}
		rtn_obj['msg_error'] = "Registerd Not Successful!"
		rtn_obj['ack'] = "2"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def delete(self, request) :
		rtn_obj={}
		rtn_obj['msg_error'] = "Registerd Not Successful!"
		rtn_obj['ack'] = "2"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done################
def activate_link(request,base64string):
	import json
	import base64
	rtn_obj = {}
	rtnvalue = activate_account(base64string)
	if rtnvalue == "0":
		rtn_obj['msg_error'] = "Account is already Activated!"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponseRedirect(settings.PATH_URL+"login")
		
	else :
		   rtn_obj['msg_error'] = "Account is Activated!"
		   rtn_obj['Ack'] = "1"
		   data = json.dumps(rtn_obj)
		   return HttpResponseRedirect(settings.PATH_URL+"login")

###############Done################
def activate_account(base64string):
	if UserProfile.objects.filter(activate_token=int(base64string)).exists():
		profile=UserProfile.objects.get(activate_token=int(base64string))
		user = User.objects.get(pk=profile.user_id_id)
		user.is_active = True
		user.save()
		value_to_rtn = "1"
		return value_to_rtn

class user_profile(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		import datetime 
		rtn_obj = {}
		imageurl =settings.IMAGE_URL
		new_user_profile = {}
		subcription_details_list = []
		final_subcription_details_list = []
		mm = []
		final_subcription_details_dict = {}
		user_id = request.POST.get('user_id',None)
		# return HttpResponse(user_id)
		cursor11 = connection.cursor()
		sql1 = "select id,email,first_name,last_name from auth_user where id='"+user_id+"'"
		cursor11.execute(sql1)
		user = cursor11.fetchone()

		cursor = connection.cursor()
		user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
		cursor.execute(user_details_sql)
		user_details = dictfetchall(cursor)

		cursor12 = connection.cursor()
		sql2 = "select id, address, phone_number, profile_image, latitude, longitude, about_me, physical_ability, is_verified, interest_id_id , subscribed, subscription_end_date,subscription_start_date from adminpanel_userprofile where user_id_id='"+user_id+"'"
		## NEW DECORATED USER PROFILE ##
		try :
			user_profile_obj = UserProfile.objects.get(user_id=int(user_id))
			new_user_profile['id'] = user_profile_obj.id
			new_user_profile['address'] = user_profile_obj.address
			new_user_profile['phone_number'] = user_profile_obj.phone_number
			new_user_profile['profile_image'] = user_profile_obj.profile_image
			new_user_profile['latitude'] = user_profile_obj.latitude
			new_user_profile['longitude'] = user_profile_obj.longitude
			new_user_profile['physical_ability'] = user_profile_obj.physical_ability
			new_user_profile['is_verified'] = user_profile_obj.is_verified
			new_user_profile['about_me'] = user_profile_obj.about_me
			new_user_profile['interest_id'] = user_profile_obj.interest_id.id
			new_user_profile['subscribed'] = user_profile_obj.subscribed
			new_user_profile['subscription_end_date'] = user_profile_obj.subscription_end_date
			new_user_profile['subscription_start_date'] = user_profile_obj.subscription_start_date
		except :
			pass
		## NEW DECORATED USER PROFILE ##
		cursor12.execute(sql2)
		userprofiles = cursor12.fetchone()
		volunteerexists ={}
		coordinatorexist = {}
		org_list = {}
		cursor14 = connection.cursor()
		
		sql4 ="select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest left join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org = 0 and adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.status ='Approved' and adminpanel_cordinatorrequest.is_request='volunteer'"
		cursor14.execute(sql4)
		volunteerexist = dictfetchall(cursor14)

		idArray = []
		for key, value in enumerate(volunteerexist):
			if value['org_id'] not in idArray:
				idArray.append(value['org_id'])
		
		cursor13 = connection.cursor()
		sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest left join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org = 0 and adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.is_request='coordinator'"
		cursor13.execute(sql3)
		coordinatorexist = dictfetchall(cursor13)

		idArray1 = []
		for key, value in enumerate(coordinatorexist):
			if value['org_id'] not in idArray1:
				idArray1.append(value['org_id'])

		cursor14 = connection.cursor()
		sql5 = "select adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.status as astatus, adminpanel_organization.address as address from adminpanel_organization where adminpanel_organization.affiliated_org = 0 and adminpanel_organization.user_id='"+user_id+"'"
		cursor14.execute(sql5)
		org_list = dictfetchall(cursor14)
		## user subscription details  ##
		# 0,1 for active subscription, 3 -> for unsubscribe
		subcription_details_informations = PayementInformations.objects.filter(user_id=int(user_id)).order_by('-id')
		for subscription_obj in subcription_details_informations :
			deleted_flag = 0
			if subscription_obj.subscription_deleted == SubcriptionDeletion.DEACTIVATED :
				print('here')
				break
			else:
				from datetime import datetime
				if subscription_obj.subscription_deleted == SubcriptionDeletion.ACTIVE :
					try:
						str_to_dict_result = ast.literal_eval(subscription_obj.subscription_response)
					except Exception as ex:
						# if type(ex).__name__ == 'ValueError':
						str_to_dict_result = json.loads(subscription_obj.subscription_response)

				if subscription_obj.subscription_deleted == SubcriptionDeletion.DELETED :
					deleted_flag = 1
					print('nmnmnmnmnmnmnmnmnmnm')
					print(subscription_obj.id)
					print('dfsdrsdrstdxtctcccyc')
					try:
						str_to_dict_result = ast.literal_eval(subscription_obj.subscription_response)
					except Exception as ex:
						# if type(ex).__name__ == 'ValueError':
						str_to_dict_result = json.loads(subscription_obj.subscription_response)
					# str_to_dict_result = json.loads(subscription_obj.subscription_response)
				# subcription_details_dict = {}
				if deleted_flag == 1 :
					subscription_start_date = datetime.fromtimestamp(int(subscription_obj.subscription_start_date))
					subscription_end_date = datetime.fromtimestamp(int(subscription_obj.subscription_end_date))
				else:
					subscription_start_date = datetime.fromtimestamp(int(str_to_dict_result['current_period_start']))
					subscription_end_date = datetime.fromtimestamp(int(str_to_dict_result['current_period_end']))
				# mm.append(subscription_start_date)
				# mm.append(subscription_end_date)
				# subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
				# subscription_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
				# subcription_details_dict['start_date'] = subscription_start_date
				# subcription_details_dict['end_date'] = subscription_end_date
				subcription_details_list.append(subscription_start_date)
				subcription_details_list.append(subscription_end_date)
				
		
		
				

		if len(subcription_details_list) >0 :
			len_list = len(subcription_details_list)
			min_dt = min(subcription_details_list)
			max_dt = max(subcription_details_list)
			final_subcription_details_dict['start_date'] = '{}-{}-{}'.format(min_dt.year, min_dt.month, min_dt.day)
			final_subcription_details_dict['end_date'] = '{}-{}-{}'.format(max_dt.year, max_dt.month, max_dt.day)
			final_subcription_details_list.append(final_subcription_details_dict)
		## user subscription details ##
		if coordinatorexist or  volunteerexist:
			data = json.dumps({"Ack":"1","user_details":user_details, "user_id":user_id,'user':user,'userprofile':userprofiles,'coordinatorexists':coordinatorexist,'imageurl':imageurl, 'volunteers_exists' : volunteerexist,'new_user_profile':new_user_profile,'final_subcription_details_list':final_subcription_details_list})
			return HttpResponse(data)
		else:
		   data = json.dumps({"Ack":"0", "user_id":user_id,'user':user,'userprofile':userprofiles,'coordinatorexists':'','imageurl':'', 'volunteers_exists' : '','user_details':user_details,'new_user_profile':new_user_profile,'final_subcription_details_list':final_subcription_details_list})
		   return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack":"0", "user_id":'','user':'','userprofile':'','coordinatorexists':'','imageurl':'', 'volunteers_exists' : ''})
		return HttpResponse(data)
	def delete(self,request):
		data = json.dumps({"Ack":"0", "user_id":'','user':'','userprofile':'','coordinatorexists':'','imageurl':'', 'volunteers_exists' : ''})
		return HttpResponse(data)


class get_cordinator_organization(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json

		user_id = request.POST.get('user_id',None)
		cursor13 = connection.cursor()
		sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org = 0 and adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.is_request='coordinator' "
		cursor13.execute(sql3)
		coordinatorexist = dictfetchall(cursor13)
		data = json.dumps({"Ack":"1", "coordinatorexist" :coordinatorexist ,"msg": "All listed value here"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack":"0", "msg": "Only Post method allowed"})
		return HttpResponse(data)

###############Done3################
# def profile_edit(request):
class profile_edit(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		rtn_obj = {}
		user_id = request.POST.get('user_id',None)
		email = request.POST.get('email',None)
		first_name = request.POST.get('first_name',None)
		last_name = request.POST.get('last_name',None)
		phone_number = request.POST.get('phone_number',None)
		# if request.method == 'POST':
		User.objects.filter(id=str(user_id)).update(first_name=first_name, last_name=last_name, email=email)
		UserProfile.objects.filter(user_id_id=user_id).update(phone_number=phone_number)
		rtn_obj['msg_error'] = "Profile Updated successfully!"
		rtn_obj['Ack'] = "1"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
		# else :
		# 	rtn_obj['msg_error'] = "Profile not Updated successfully!"
		# 	rtn_obj['Ack'] = "0"
		# 	data = json.dumps(rtn_obj)
		# 	return HttpResponse(data)
	def get(self,request):
		rtn_obj = {}
		rtn_obj['msg_error'] = "Profile not Updated successfully!"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def put(self,request):
		rtn_obj = {}
		rtn_obj['msg_error'] = "Profile not Updated successfully!"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def delete(self,request):
		rtn_obj = {}
		rtn_obj['msg_error'] = "Profile not Updated successfully!"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	def patch(self,request):
		rtn_obj = {}
		rtn_obj['msg_error'] = "Profile not Updated successfully!"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done1################
# def change_password(request):
class change_password(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		import base64
		user_id = request.POST.get('user_id', None)
		new_password = request.POST.get('new_password', None)
		conf_password = request.POST.get('conf_password', None)
		old_password = request.POST.get('old_password', None)


		first_decode = base64.b64decode(new_password)
		second_step = str(first_decode, 'utf-8')
		splitdata = second_step.split('----')
		second_decode = base64.b64decode(splitdata[0])
		password_decrypted = str(second_decode, 'utf-8')

		first_decode_confirmpassword = base64.b64decode(conf_password)
		second_step_confirmpassword = str(first_decode_confirmpassword, 'utf-8')
		splitdata_confirmpassword = second_step_confirmpassword.split('----')
		second_decode_confirmpassword = base64.b64decode(splitdata_confirmpassword[0])
		confirmpassword_decrypted = str(second_decode_confirmpassword, 'utf-8')

		first_decode_oldpass = base64.b64decode(old_password)
		second_step_oldpass = str(first_decode_oldpass, 'utf-8')
		splitdata_oldpass = second_step_oldpass.split('----')
		second_decode_oldpass = base64.b64decode(splitdata_oldpass[0])
		oldpass_decrypted = str(second_decode_oldpass, 'utf-8')

		user_details = User.objects.get(id=user_id)

		user = authenticate(username=user_details.username, password=oldpass_decrypted)
			#return HttpResponse(user)
		if user is not None:
			if password_decrypted == confirmpassword_decrypted:
				u = User.objects.get(id=user_id)
				u.set_password(password_decrypted)
				u.save()

				data = json.dumps({"Ack": 1, "msg": "Password successfully changed"})
			else:
				data = json.dumps({"Ack": 0, "msg": "Password and confirm password not matched"})
		else:
			data = json.dumps({"Ack": 1, "msg": "Old password is wrong"})
		return HttpResponse(data)

###############Done3################
def get_activity_categories(self):
	import json
	get_cat = ActivityCategory.objects.all().order_by('id')
	categories = []
	if get_cat:
		for cat in get_cat:
			cat_array = {}
			cat_array['id'] = cat.id
			cat_array['name'] = cat.name
			cat_array['description'] = cat.description
			cat_array['category_image'] = cat.category_image
			categories.append(cat_array)

	data = json.dumps({"Ack": 1, "categories": categories})

	return HttpResponse(data)

###############Done1################
def get_organizations(self):
	import json

	org_details = []
		
	organization_details = Organization.objects.all()
	org_details = fetch_row_organization(organization_details)

	data = json.dumps({"Ack": 1, "organizations": org_details})
	return HttpResponse(data)

###############Done1################
class update_organization(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		organization_id = request.POST.get('org_id', None)
		org_details = []
		
		organization_details = Organization.objects.filter(id=str(organization_id))
		org_details = fetch_row_organization(organization_details)

		if request.FILES:
			# folder=settings.UPLOAD_URL
			# image_path = settings.UPLOAD_URL_ROOT
			
			# org_image = request.FILES['org_image']
			# fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 
			# filename = fs.save(org_image.name, org_image)
			# file_url_image = fs.url(filename)
			# file_url = request.FILES['org_image'].name

			folder=settings.ORGANISATION_IMAGE
			org_image = request.FILES['org_image']
			fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 
			
			splitted_value = org_image.name.split(".")
			img_type = splitted_value[len(splitted_value)-1]

			image_name_get = ''
			for i in range(len(splitted_value)-1):
				image_name_get+=str(splitted_value[i]).replace(" ", "_")+'_'
			
			img_name =image_name_get+ str( time.time()).split('.')[0]
			final_image = img_name+'.'+img_type

			filename = fs.save(final_image, org_image)

			imgthumb = Image.open(settings.ORGANISATION_IMAGE+final_image)
			imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
			imgthumb.save(settings.ORGANISATION_IMAGE+final_image)
			
			file_url =final_image

		else:
			file_url = org_details[0]['photo']

		tax_id=	request.POST.get('tax_id', org_details[0]['tax_id'])
		email = request.POST.get('email', org_details[0]['email'])
		phone = request.POST.get('phone', org_details[0]['phone'])
		peername = request.POST.get('peer_name', None)
		user_id = request.POST.get('user_id', None)
		
		Organization.objects.filter(id=str(organization_id)).update(tax_id = tax_id,email= email,phone=phone)
 
		cordinatorRequest = CordinatorRequest(
			user_id = User.objects.get(id=str(user_id)),
			org_id = Organization.objects.get(id=str(organization_id)),
			status = 'Pending',
			role = int(2),
			oppurtunity_id = 0,
			is_url = False,
			is_request = "coordinator"
		) 
		cordinatorRequest.save()

		CordinatorRequest.objects.filter(user_id=str(user_id),org_id =str(organization_id)).update(is_request="coordinator")
		user_details = []
		user_details_fetch = User.objects.filter(id=str(user_id)).values('first_name', 'last_name', 'email')

		full_data = {}

		full_data['first_name'] = user_details_fetch[0]['first_name']
		full_data['last_name'] = user_details_fetch[0]['last_name']
		full_data['email'] = user_details_fetch[0]['email']
		user_details.append(full_data)
		
		request_link = settings.PATH_URL+"/approve_coordinator_request/"+str(organization_id)+"/"+str(full_data['email'])
		
		msg_html = '<p>Hi Admin,</p><p>There is a request for coordinator of Organisation named '+org_details[0]['organization_name']+',<br/> tax_id:'+tax_id+', <br/> email:'+email+', <br/> phone:'+phone+'</p><a href='+request_link+'>Request Access Link:'+request_link+' </a><p>Thanks, <br/> Inteer Team</p>'

		ReceiverEmail = 'ainteer72@gmail.com'

		user_id = request.POST.get('user_id', None)
		userprofiles= User.objects.filter(id=user_id)
		msg_html2 = '<p>Thanks for applying as a coordinator. We are reviewing your information. We are in process of contacting this person and we will let you know as soon as we know more.</p> <p>Thanks,</p><p>Inteer Team</p>'

		send_mail("Notification For Coordinator Request Access", 'Inteer','ainteer72@gmail.com',[userprofiles[0].email],html_message=msg_html2 )

		if send_mail("Notification For Coordinator Request Access", 'Inteer',userprofiles[0].email,['ainteer72@gmail.com'], html_message=msg_html ) :
			data = json.dumps({"Ack": 1, "msg": "Organization successfully saved","link":request_link,'body':msg_html})
		else:
			data = json.dumps({"Ack": 0, "msg": "Mail sent failed"}) 

		return HttpResponse(data)

###############Done2################
class approve_coordinator_request(ObtainAuthToken):
	
	
	def post(self,request) :
		import json
		rtn_obj = {}
		volunteer_mail = request.POST['volunteer_mail']
		organization_id = request.POST['organization_id']
		# pp = user=User.objects.get(id=9)
		# test_user=User.objects.get(email=pp.email)
		# print(test_user)
		# print('jlhjhjhjkk')
		if organization_id:
			organization=Organization.objects.get(id=organization_id)
			if organization is not None:
				Organization.objects.filter(id=organization_id).update(status='Approved')
		if volunteer_mail:
			print(volunteer_mail)
			# user=User.objects.filter(email__icontains=volunteer_mail).last()
			user=User.objects.get(email=volunteer_mail)
			print(user.id)
			print(volunteer_mail)
			print('mmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmmm')
			if user is not None:
				userpro = UserProfile.objects.get(user_id = user.id)
				UserProfile.objects.filter(user_id_id=str(user.id)).update(role_id=3)
				CordinatorRequest.objects.filter(user_id=str(user.id),org_id_id=str(organization_id)).update(role=3,status='Approved')
				
				# user_id = request.POST.get('user_id', None)
				userprofiles= User.objects.filter(id=user.id)
				msg_html2 = '<p> Congratulations we have completed your application process and here`s what you can look forward to</p> <p>Thanks,</p><p>Inteer Team</p>'

				send_mail("Notification For Coordinator Request Access", 'Inteer','ainteer72@gmail.com',[userprofiles[0].email],html_message=msg_html2 )
				rtn_obj['ack'] = "1"
				rtn_obj['msg_error'] = "Request accepted successfully"
				rtn_obj['role_id'] = 3
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else:
				rtn_obj['ack'] = "0"
				rtn_obj['msg_error'] = " Mail id not valid. "

				data = json.dumps(rtn_obj)
				return HttpResponse(data)
	def get(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Only post method allowed"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done6################
class approve_organization_request(ObtainAuthToken):
	

	def post(self,request) :
		import json
		rtn_obj = {}
		organization_id = request.POST['organization_id']
		
		if organization_id:
			organization=Organization.objects.get(id=organization_id)
			if organization is not None:
				Organization.objects.filter(id=organization_id).update(status='Approved')
				
				rtn_obj['ack'] = "1"
				rtn_obj['msg_error'] = "Request accepted successfully"
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else:
				rtn_obj['ack'] = "0"
				rtn_obj['msg_error'] = " Organization id not valid. "

				data = json.dumps(rtn_obj)
				return HttpResponse(data)
	def get(self,request) :
		rtn_obj = {}
		rtn_obj['ack'] = "0"
		rtn_obj['msg_error'] = "Only post method allowed"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done2################
class approve_all(APIView):

	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		data = {}
		all_volunteer = request.POST.get('all_volunteer')
		all_volunteer = json.loads(all_volunteer)
		# return HttpResponse(data)
		if all_volunteer != '':
			for volunteer in all_volunteer:
				volunteers = CordinatorRequest.objects.filter(user_id_id = volunteer['user_id_id']).update(status = 'Approved')
			data = json.dumps({"Ack": 1, "msg": "All application is approved"}) 
		else:
			data = json.dumps({"Ack": 0, "msg": "volunteer list is empty"})
		return HttpResponse(data)
			 
	def get(self,request):
		data = {}
		data = json.dumps({"Ack": 0, "msg": "only post method allowed"})
		return HttpResponse(data)


###############Done1################
class save_organization(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		# print(request.FILES)
		# print(request.POST)

		if request.FILES:

			folder=settings.ORGANISATION_IMAGE
			# return HttpResponse(folder)
			# return HttpResponse(folder)
			# `image_path` = settings.ORGANISATION_IMAGE
			# image_path = settings.UPLOAD_URL_ROOT
			
			org_image = request.FILES['file']
			fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 
			
			# file_url_image = fs.url(filename)
			# print('lllllllllllllllllllllllll')
			# print(file_url_image)
			# splitted_value = file_url_image.split("media/")  
			# path = settings.UPLOAD_URL+org_image.name
			# return HttpResponse(splitted_value[1]) 
			# # path = settings.UPLOAD_URL+org_image.name
			# return HttpResponse(settings.UPLOAD_URL)

			splitted_value = org_image.name.split(".")
			img_type = splitted_value[len(splitted_value)-1]

			image_name_get = ''
			for i in range(len(splitted_value)-1):
				image_name_get+=str(splitted_value[i]).replace(" ", "_")+'_'
			
			img_name =image_name_get+ str( time.time()).split('.')[0]
			final_image = img_name+'.'+img_type

			filename = fs.save(final_image, org_image)

			imgthumb = Image.open(settings.ORGANISATION_IMAGE+final_image)
			imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
			imgthumb.save(settings.ORGANISATION_IMAGE+final_image)
			
			file_url =final_image
		else:
			file_url = ""
		
		
		organization_name = request.POST.get('organization_name', None)
		address = request.POST.get('address', None)
		print(request.POST.get('organization_name', None))
		print(request.POST)
		presence_check = Organization.objects.filter(organization_name = organization_name,address=address)
		print('hhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhhh')
		print(len(presence_check))
		print('llllllllllllllllllllllllllllllllllll')
		# return HttpResponse(presence_check)
		user_id = request.POST.get('user_id', None)
		userprofiles= User.objects.filter(id=user_id)
		if presence_check :

			data = json.dumps({"Ack": 2, "msg": "Already Present"}) 
		else:
			organization = Organization(
				organization_name = request.POST.get('organization_name', None),
				parent_id = request.POST.get('parent_id', None),
				address = request.POST.get('address', None),
				tax_id = request.POST.get('tax_id', None),
				email = request.POST.get('email', None),
				web_url = request.POST.get('web_url', None),
				phone = request.POST.get('phone', None),
				cause = request.POST.get('peer_name', None),
				about_us = request.POST.get('about_us', None),
				our_activity = request.POST.get('our_activity', None),
				why_us = request.POST.get('why_us', None),
				photo = file_url,
				address1 = request.POST.get('address1', None),
				phone2 = request.POST.get('phone2', None),
				irs_rank = request.POST.get('irs_rank', None),
				fb_url = request.POST.get('fb_url', None),
				event1 = request.POST.get('event1', None),
				event2 = request.POST.get('event2', None),
				event3 = request.POST.get('event3', None),
				status= 'Pending',
				user_id = request.POST.get('user_id', None)
			)
			organization.save() 

			user_details = []
			user_id = request.POST.get('user_id', None)
			user_details_fetch = User.objects.filter(id=int(user_id)).values('first_name', 'last_name', 'email')

			full_data = {}

			full_data['first_name'] = user_details_fetch[0]['first_name']
			full_data['last_name'] = user_details_fetch[0]['last_name']
			full_data['email'] = user_details_fetch[0]['email']
			user_details.append(full_data)

			cordinatorRequest = CordinatorRequest(
				user_id = User.objects.get(id=int(user_id)),
				org_id = Organization.objects.get(id=int(organization.id)),
				status = 'Pending',
				role = int(2),
				oppurtunity_id = 0,
				is_url = False,
				is_request = "coordinator"
			)  
			cordinatorRequest.save()

			org_id = organization.id
			request_link = settings.PATH_URL+"/approve_coordinator_request/"+str(org_id)+"/"+str(full_data['email'])
			# request_link2 = settings.PATH_URL+"/approve_coordinator_request/"+str(full_data['email'])
			
			msg_html = '<p>Hi Admin,</p><p>There is a addition of new Organisation named '+request.POST.get('organization_name', None)+',<br/> tax_id:'+request.POST.get('tax_id', None)+', <br/> email:'+request.POST.get('email', None)+', <br/> phone:'+request.POST.get('phone', None)+'</p><a href='+request_link+'>Request Access Link:'+request_link+' </a><p><p>Thanks, <br/> Inteer Team</p>'

			msg_html2 = '<p>Hello '+userprofiles[0].first_name+' '+userprofiles[0].last_name+',</p> <p>Thanks for applying as a coordinator. We are reviewing your information. We are in process of contacting this person and we will let you know as soon as we know more.</p> <p>Thanks,</p><p>Inteer Team</p>'

			send_mail("Notification For Coordinator Request Access", 'Inteer','ainteer72@gmail.com',[userprofiles[0].email],html_message=msg_html2 )

			# ReceiverEmail = 'kher.nachiket@gmail.com'
			ReceiverEmail = 'ainteer72@gmail.com'
			if send_mail('Notification Email For Organization', 'InterApp', 'ainteer72@gmail.com', [ReceiverEmail], html_message=msg_html):
				data = json.dumps({"Ack": 1, "msg": "Organization successfully saved"})
			else:
				data = json.dumps({"Ack": 0, "msg": "Mail sent failed"}) 
			
			# cordinatorRequest = CordinatorRequest(
			# 	user_id = User.objects.get(id=str(user_id)),
			# 	org_id = Organization.objects.get(id=str(org_id)),
			# 	status = 'Pending',
			# 	role = int(2),
			# 	oppurtunity_id = 0,
			# 	is_url = False,
			# 	is_request = "coordinator"
			# ) 
			# cordinatorRequest.save()

			# request_link = settings.PATH_URL+"/approve_coordinator_request/"+str(full_data['email'])
			
			# msg_html = '<p>Hi Admin,</p><p>There is a request for coordinator of Organisation named '+org_details[0]['organization_name']+',<br/> tax_id:'+tax_id+', <br/> email:'+email+', <br/> phone:'+phone+'</p><a href='+request_link+'>Request Access Link:'+request_link+' </a><p>Thanks, <br/> Inteer Team</p>'

			# ReceiverEmail = 'ainteer72@gmail.com'
			# if send_mail('Notification Email For Request Coordinator', 'InterApp', 'ainteer72@gmail.com', [ReceiverEmail], html_message=msg_html):
			# 	data = json.dumps({"Ack": 1, "msg": "Organization successfully saved","link":request_link,'body':msg_html})
			# else:
			# 	data = json.dumps({"Ack": 0, "msg": "Mail sent failed"}) 
		return HttpResponse(data)
	
###############Done1################
# def get_opportunity_by_user(request):
class get_opportunity_by_user(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json

		user_id = request.POST.get('user_id', None)
		oppor_details = []
		opportunity_details = Opportunities.objects.filter(user_id_id=str(user_id))
		oppor_details = fetch_row_opportunity(opportunity_details) 
		data = json.dumps({"Ack": 1, "opportunities": oppor_details})
		return HttpResponse(data)

###############Done1################
class add_activity_category(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):

		import json

		name = request.POST.get('name', None)
		desc = request.POST.get('description', None)

		activity_category = ActivityCategory(
			name = name,
			description = desc
		)
		activity_category.save()

		data = json.dumps({"Ack": 1, "msg": "Activity category successfully saved"})
		return HttpResponse(data)

	def get(self,request):

		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

###############Done1################
class set_end_date(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):

		import json

		start_date = request.POST.get('start_date', None)
		repeat_type = request.POST.get('repeat_type', None)
		repeat_number = request.POST.get('repeat_number', None)

class reoccur(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		import datetime
	
		rtn_obj = {}
		user_id = request.POST.get('user_id', None)
		opportunity_id = request.POST.get('opportunity_id', None)
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		repeat_type = request.POST.get('repeat_type', None)
		repeat_times = request.POST.get('repeat_times', None)
		date_1 = datetime.datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
		date_2 = datetime.datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S") 
		end_moment = request.POST.get('end_moment', None)

		cursor = connection.cursor()

		fetch_query = "select ao.id,ao.user_id_id,ao.opportunity_name, ao.org_id_id, ao.lat, ao.lon, ao.description, ao.address, ao.image,ao.no_of_volunteers, ao.start_date as start_date, ao.end_date as end_date, au.first_name, au.last_name from adminpanel_opportunities ao left join auth_user au on ao.user_id_id = au.id WHERE ao.id="+opportunity_id
		cursor.execute(fetch_query)
		opportunity = dictfetchall(cursor)

		
		if int(repeat_times) > 0:
			if repeat_type == "daily":
				i = 1
				while i <= int(repeat_times):
					new_date = date_1 + datetime.timedelta(days=i)
					end_date = date_2 + datetime.timedelta(days=i)
					opportunitiy = Opportunities(
						user_id_id = opportunity[0]['user_id_id'],
						org_id_id = opportunity[0]['org_id_id'],
						opportunity_name = opportunity[0]['opportunity_name'],
						description = opportunity[0]['description'],
						address = opportunity[0]['address'],
						parent_id = opportunity_id,
						parent_opportunity = opportunity_id,
						start_date = new_date,
						end_date = end_date,
						image = opportunity[0]['image'],
						lat = opportunity[0]['lat'],
						lon = opportunity[0]['lon']

					)

					opportunitiy.save()
					i += 1
					print(new_date)
				
			elif repeat_type == "weekly":
				j = 1
				while j <= int(repeat_times):
					new_date = date_1 + datetime.timedelta(days=(7*j))
					end_date = date_2 + datetime.timedelta(days=(7*j)) 
					opportunitiy = Opportunities(
						user_id_id = opportunity[0]['user_id_id'],
						org_id_id = opportunity[0]['org_id_id'],
						opportunity_name = opportunity[0]['opportunity_name'],
						description = opportunity[0]['description'],
						address = opportunity[0]['address'],
						parent_id = opportunity_id,
						parent_opportunity = opportunity_id,
						start_date = new_date,
						end_date = end_date,
						image = opportunity[0]['image'],
						lat = opportunity[0]['lat'],
						lon = opportunity[0]['lon']

					)

					opportunitiy.save()
					j += 1
					
			elif repeat_type == "monthly":
				k = 1
				while k <= int(repeat_times):
					new_date = date_1 + datetime.timedelta(days=(31*k))
					end_date = date_2 + datetime.timedelta(days=(31*k))
					# k += 1
					opportunitiy = Opportunities(
						user_id_id = opportunity[0]['user_id_id'],
						org_id_id = opportunity[0]['org_id_id'],
						opportunity_name = opportunity[0]['opportunity_name'],
						description = opportunity[0]['description'],
						address = opportunity[0]['address'],
						parent_id = opportunity_id,
						parent_opportunity = opportunity_id,
						start_date = new_date,
						end_date = end_date,
						image = opportunity[0]['image'],
						lat = opportunity[0]['lat'],
						lon = opportunity[0]['lon']

					)

					opportunitiy.save()
					k += 1
					print(new_date)
			
			Opportunities.objects.filter(id=str(opportunity_id)).update(repeat_number=repeat_times)
			rtn_obj['msg'] = "opportunities Updated successfully!"
			rtn_obj['Ack'] = "1"
			rtn_obj['repeat_times']= repeat_times

			data = json.dumps(rtn_obj)
			return HttpResponse(data)

	def get(self,request):
		rtn_obj = {}
		rtn_obj['msg'] = "input data by post method"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

###############Done4################
class  edit_opportunity(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self, request):
		import json
		oppo_id = request.POST.get('oppo_id', None)
		user_id = request.POST.get('user_id', None)
		
		opportunity_details = Opportunities.objects.filter(id=str(oppo_id)).values('org_id_id','opportunity_name','description','no_of_volunteers','author_name','address','parent_id','no_ofyear','no_ofyear','start_date','end_date','image')
		
		org_id = request.POST.get('org_id', opportunity_details[0]['org_id_id'])
		opportunity_name = request.POST.get('opportunity_name', opportunity_details[0]['opportunity_name'])
		description = request.POST.get('description', opportunity_details[0]['description'])
		no_of_volunteers = request.POST.get('no_of_volunteers', None)
		author_name = request.POST.get('author_name', opportunity_details[0]['author_name'])
		address = request.POST.get('address', opportunity_details[0]['address'])

		parent_id = request.POST.get('parent_id', opportunity_details[0]['parent_id'])
		no_ofyear = request.POST.get('no_ofyear', opportunity_details[0]['no_ofyear'])

		start_date = request.POST.get('start_date', opportunity_details[0]['start_date'])
		end_date = request.POST.get('end_date', opportunity_details[0]['end_date'])
		lat = request.POST.get('lat')
		lon = request.POST.get('lon')
		questions = request.POST.get('questions')
		opportunityObj = Opportunities.objects.get(id = str(oppo_id))
		

		category_ids = request.POST.get('category_id')
		category_ids = json.loads(category_ids)

		if category_ids:
			for cat in category_ids:
				search_object = OpportunityCategories.objects.filter(category_id_id =str(cat['id']),opportunity_id_id=str(oppo_id))
				if search_object:
					new_added = 0
				else:
					catObj = ActivityCategory.objects.get(id = str(cat['id']))
					opportunityCategories = OpportunityCategories(
						opportunity_id = opportunityObj,
						category_id = catObj
					)
					opportunityCategories.save()
		

		if len(request.FILES) != 0:
			image_file = request.FILES['file']
		else :
			image_file = opportunity_details[0]['image']

		org_id_id = opportunity_details[0]['org_id_id']

		user = User.objects.get(id=user_id)
		organization = Organization.objects.get(id = org_id_id)
		
		opportunitiy = Opportunities(
			user_id = user,
			org_id = organization,
			opportunity_name = opportunity_name,
			description = description,
			author_name = author_name,
			address = address,
			parent_id = 0,
			no_ofyear = no_ofyear,
			parent_opportunity = 0,
			start_date = start_date,
			end_date = end_date,
			image = image_file,
			id = oppo_id,
			no_of_volunteers= no_of_volunteers,
			lat = lat,
			lon = lon
		)

		opportunitiy.save()
		return HttpResponse(1) 
		if image_file:
			imgthumb = Image.open(settings.MEDIA_ROOT+"/"+opportunitiy.image.name)
			imgthumb.save(settings.MEDIA_ROOT + '/' + opportunitiy.image.name)
			opportunitiy.image = opportunitiy.image.name
			opportunitiy.save()

		
		for cat in category_ids:
			catObj = ActivityCategory.objects.get(id = str(cat['id']))
			opportunityCategories = OpportunityCategories(
				opportunity_id = opportunityObj,
				category_id = catObj
			)
			opportunityCategories.save()

		
		if questions != None :
			questions = json.loads(questions)
			for question in questions:
				question = OpportunityQuestions(
					opportunity = opportunityObj,
					question = question['question']
				)
				question.save()

		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved","id": opportunitiy.id,"start_date":start_date,"end_date":end_date})
		return HttpResponse(data)

	def get(self, request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

###############Done4################
def get_organization_name(request):
	import json

	if request.method == 'POST':
		user_id = request.POST.get('user_id', None)
		# return HttpResponse(user_id)
		organization = []
		if user_id:
			getorgId = CordinatorRequest.objects.filter(user_id_id = user_id).values('org_id_id')

			if len(getorgId) > 0:
				
				org_id_id = getorgId[0]['org_id_id']
				get_details_organization = Organization.objects.filter(id=str(org_id_id)).values('organization_name')

				data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved","Org_name": get_details_organization[0]['organization_name']})		

				return HttpResponse(data) 

			else:
				return HttpResponse(0)
			
	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})

	return HttpResponse(1)

###############Done4################
# def add_opportunity(request):
class add_opportunity(APIView):
	# return HttpResponse(1)

	def post(self,request):
		import json
		user_id = request.POST.get('user_id', None)
		org_id = request.POST.get('org_id', None)
		opportunity_name = request.POST.get('opportunity_name', None)
		description = request.POST.get('description', None)
		no_of_volunteers = request.POST.get('no_of_volunteers', None)
		author_name = request.POST.get('author_name', None)
		address = request.POST.get('address', None)
		lat = request.POST.get('lat', 22.2345)
		lon = request.POST.get('lon',88.4567)
		
		parent_id = request.POST.get('parent_id', None)
		no_ofyear = request.POST.get('no_ofyear', None)

		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		# return HttpResponse(end_date)
		
		category_ids = request.POST.get('category_id')
		# return HttpResponse(category_ids)
		category_ids = json.loads(category_ids)
	
		if len(request.FILES) != 0:
			image_file = request.FILES['file']

			
			# file_url = image_file.name

		else :
			image_file = ''
		
		org_id_id =116
		# cursor = connection.cursor()
		# sql = "select * from adminpanel_cordinatorrequest where user_id_id='"+ user_id +"'"
		# cursor.execute(sql)
		# getorgId = dictfetchall(cursor)
		# if org_id:
		# 	org_id_id =org_id
		# else:
		# 	org_id_id =116
		# print('yyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyy')
		# # print(org_id)
		# return HttpResponse(1)
		getorgId = CordinatorRequest.objects.filter(user_id_id = user_id,is_request = 'coordinator').values('org_id_id')
		
		if len(getorgId) > 0:
			org_id_id = getorgId[0]['org_id_id']
			

		user = User.objects.get(id=user_id)
		organization = Organization.objects.get(id = org_id_id)
		opportunitiy = Opportunities(
			user_id = user,
			org_id = organization,
			opportunity_name = opportunity_name,
			description = description,
			author_name = author_name,
			address = address,
			parent_id = 0,
			lon = lon,
			lat = lat,
			no_ofyear = no_ofyear,
			parent_opportunity = 0,
			start_date = start_date,
			end_date = end_date,
			image = image_file,
			no_of_volunteers= no_of_volunteers
		)

		opportunitiy.save()
		# return HttpResponse(24346576)
		# print(opportunitiy.image.name)
		# print(image_file)
		if len(request.FILES) != 0:

			folder=settings.UPLOAD_IMAGES_URL
			fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 
			

			splitted_value = opportunitiy.image.name.split(".")
			img_type = splitted_value[len(splitted_value)-1]

			image_name_get = ''
			for i in range(len(splitted_value)-1):
				image_name_get+=str(splitted_value[i]).replace(" ", "_")+'_'
			
			img_name =image_name_get+ str( time.time()).split('.')[0]
			final_image = img_name+'.'+img_type

			filename = fs.save(final_image, image_file)

			file_url_image = fs.url(filename)


		if image_file:
			# imgthumb = Image.open(settings.UPLOAD_IMAGES_URL+final_image)

			# imgthumb = Image.open(settings.MEDIA_ROOT+"/"+final_image)

			# imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
			# imgthumb.save(settings.UPLOAD_IMAGES_URL + final_image)

			# imgthumb.save(settings.MEDIA_ROOT + '/' + final_image)
			
			opportunitiy.image = final_image
			opportunitiy.save()

		# opportunityObj = Opportunities.objects.latest('id')
		# opportunityObj = {}
		
		if request.POST.get('opportunity_id'):
			opportunity_id = request.POST.get('opportunity_id')
			opportunityObj = Opportunities.objects.get(id = str(opportunity_id))
		else:
			opportunityObj = Opportunities.objects.latest('id')
			
		for cat in category_ids:
			catObj = ActivityCategory.objects.get(id = str(cat['id']))
			opportunityCategories = OpportunityCategories(
				opportunity_id = opportunityObj,
				category_id = catObj
			)
			opportunityCategories.save()

		questions = request.POST.get('questions')
		questions = json.loads(questions)
		
		if questions[0]['question'] != '':
			for question in questions:
				question = OpportunityQuestions(
					opportunity = opportunityObj,
					question = question['question']
				)
				question.save()

		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved","id": opportunitiy.id})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

###############Done1################
# def add_organiztion(request):
class add_organiztion(APIView):

	def post(self,request):
		import json
		organization_name = request.POST.get('organization_name', None)
		parent_id = request.POST.get('parent_id', None)
		address = request.POST.get('address', None)
		tax_id = request.POST.get('tax_id', None)
		email = request.POST.get('email', None)
		web_url = request.POST.get('web_url', None)
		phone = request.POST.get('phone', None)
		cause = request.POST.get('cause', None)
		about_us = request.POST.get('about_us', None)
		our_activity = request.POST.get('our_activity', None)
		why_us = request.POST.get('why_us', None)
		user_id = request.POST.get('user_id', None)

		userprofiles= User.objects.filter(id=user_id)

		# adminEmail = 'kher.nachiket@gmail.com'
		adminEmail = 'ainteer72@gmail.com'
		if len(request.FILES) != 0:
			image = request.FILES['file']
		else :
			image = ''

		organization = Organization(
			organization_name = organization_name,
			parent_id = parent_id,
			address = address,
			tax_id = tax_id,
			email = email,
			web_url = web_url,
			phone = phone,
			cause = cause,
			about_us = about_us,
			our_activity = our_activity,
			why_us = why_us,
			status= 'Pending',
			user_id = user_id
		)
		organization.save()

		if photo:
			imgthumb = Image.open(settings.MEDIA_ROOT+"/"+organization.image.name)
			imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
			imgthumb.save(settings.MEDIA_ROOT + '/' + organization.image.name)

			organization.photo = organization.image.name
			organization.save()

		msg_html = '<p>Hi Admin,</p><p>There is a addition of new Organisation named '+organization_name+',<br/> tax_id:'+tax_id+' <br/>, email:'+email+' <br/>, phone:'+phone+'</p><p><p>Thanks,</p>'
		msg_html2 = '<p>Hello '+userprofiles[0].first_name+' '+userprofiles[0].last_name+',</p> <p>Thanks for applying as a coordinator. We are reviewing your information. We are in process of contacting this person and we will let you know as soon as we know more.</p> <p>Thanks,</p><p>Inteer Team</p>'

		send_mail("Notification For Coordinator Request Access", 'Inteer','ainteer72@gmail.com',[userprofiles[0].email],html_message=msg_html2 )
		# if send_mail('Notification Email For Organization', 'InterApp', 'kher.nachiket@gmail.com', [adminEmail], html_message=msg_html):
		if send_mail('Notification Email For Organization', 'InteerApp', 'ainteer72@gmail.com', [adminEmail], html_message=msg_html):
			data = json.dumps({"Ack": 1, "msg": "Organization successfully saved"})
		else:
			data = json.dumps({"Ack": 0, "msg": "Mail sent failed"})
		return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)
	def put(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)
	def patch(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)


###############Done1################
class edit_organization(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):

		import json

		organization_id = request.POST.get('org_id', None)
		org_details = []
		
		organization_details = Organization.objects.filter(id=str(organization_id))
		org_details = fetch_row_organization(organization_details)

		

	# 	# organization_name=	request.POST.get('organization_name', org_details[0]['organization_name'])
		about_us = request.POST.get('about_us', org_details[0]['about_us'])
		our_activity = request.POST.get('our_activity', org_details[0]['our_activity'])
		why_us = request.POST.get('why_us', None)
		
		
		Organization.objects.filter(id=str(organization_id)).update(about_us = about_us,our_activity= our_activity,why_us=why_us)
		data = json.dumps({"Ack": 1, "msg": "Organization successfully updated"})
		return HttpResponse(data)
	

###############Done1################
class add_volunteer(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self, request):
		import json

		volunteer_name = request.POST.get('volunteer_name')
		volunteer = Volunteer(
			volunteer_name = volunteer_name
		)
		volunteer.save()
		data = json.dumps({"Ack": 1, "msg": "Volunteer Added"})
		return HttpResponse(data)

	def get(self,request):

		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

class get_volunteers(APIView):

	permission_classes = (IsAuthenticated,)

	def get(self,request):
		import json
		cursor = connection.cursor()
		sql = "select * from adminpanel_volunteer"
		cursor.execute(sql)
		volunteers = dictfetchall(cursor)
		data = json.dumps({"Ack": 1, "volunteers": volunteers})
		return HttpResponse(data)

	def post(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

def get_opportunities(self,offset):
	import json
	limit= 12
	cursor = connection.cursor()
	sql = "select AO.id, AO.opportunity_name, AO.description, AO.image, AU.first_name, AU.last_name, AO.org_id_id, AORG.web_url, AORG.photo, AORG.organization_name, AORG.phone2 as secret_message from adminpanel_opportunities AO left join auth_user AU on AO.user_id_id = AU.id left join adminpanel_organization AORG on AORG.id = AO.org_id_id WHERE AORG.affiliated_org = 0 order by AO.id desc LIMIT "+str(limit)+" OFFSET "+str(offset)
		
	cursor.execute(sql)
	opportunities = dictfetchall(cursor)
	data = json.dumps({"Ack": 1, "opportunities": opportunities, "image_url": settings.BASE_URL + "/media/"})
	return HttpResponse(data)

def get_opportunitiesfiltered(request,offset):
	import json

	if request.method == 'POST':
		name = request.POST.get('name', None).lower() 
		cat_ids = request.POST.get('categories', None)
		cat_ids = json.loads(cat_ids)
		ids = ",".join(str(x) for x in cat_ids)
		user_id = request.POST.get('user_id', None)
		limit = 12 

		lat = request.POST.get('lat',None)
		lon = request.POST.get('lon',None)
		changed_limit = int(offset)+limit
		# return HttpResponse(changed_limit)
		address = request.POST.get('address',None)
		user_details =[]
		related_organization = []
		if user_id:
			cursor = connection.cursor()
			user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
			cursor.execute(user_details_sql)
			user_details = dictfetchall(cursor)
			cursor13 = connection.cursor()
			sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org = 0 and adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.status='Approved'"
			cursor13.execute(sql3)
			related_organization = dictfetchall(cursor13)

		c = connection.cursor()
		sql = "SELECT  AO.id, AO.opportunity_name, AO.description, AO.image, AU.id as uid, AU.first_name, AU.last_name,AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo"
		if address:
			sql += ", (3959 * acos (cos ( radians("+str(lat)+") )* cos( radians( AO.lat ) )* cos( radians( AO.lon ) - radians("+str(lon)+") )+ sin ( radians("+str(lat)+") ) * sin( radians( AO.lat ) ))) AS distance "
		sql += " FROM adminpanel_opportunities AO "
		sql += "LEFT JOIN auth_user AU "
		sql += "ON AO.user_id_id = AU.id "
		sql += "LEFT JOIN adminpanel_userprofile AP "
		sql += "ON AO.user_id_id = AP.user_id_id "
		sql += "LEFT JOIN adminpanel_opportunitycategories AOC "
		sql += "ON AOC.opportunity_id_id = AO.id "
		sql += "LEFT JOIN adminpanel_activitycategory AAC "
		sql += "ON AAC.id = AOC.category_id_id "
		sql += "LEFT JOIN adminpanel_organization AOR "
		sql += "ON AO.org_id_id = AOR.id "
		
		sql += "WHERE 1=1 AND AOR.affiliated_org = 0 AND AO.start_date > current_date "
		
		if name:
			sql += "AND LOWER(AO.opportunity_name) LIKE '%"+name+"%' "

		if ids:
			sql += "AND AOC.category_id_id IN ("+ids+") "

		# if address:
		# 	sql += "AND distance < 50"

		sql += "ORDER BY AO.start_date ASC "
 
		# data = json.dumps({"Ack": 1, "sql": sql})
		# return HttpResponse(data) 
		c.execute(sql)
		allRowfetch = dictfetchall(c)
		# print(allRowfetch)
		# return HttpResponse(1)
		
		allRow =[]
		idArray = []
		opportunity = []
		i= 0
		
		for key, value in enumerate(allRowfetch):
			if address:
			
				# print(round(value['distance']),2)  

				if value['distance'] != 0.0 : 
					# if int(round(value['distance'])) <= 50:
					if value['distance']:
						if float(str(value['distance'])) <= float(str(80.46)):
							
							if value['id'] not in idArray:
								idArray.append(value['id'])
								allRow.append(value)
				if value['distance'] == 0.0 :
					if value['id'] not in idArray:
							idArray.append(value['id'])
							allRow.append(value)
			else:
				if value['id'] not in idArray:
					idArray.append(value['id'])
					allRow.append(value)

		if allRow:
			for i in range(len(allRow)):
				o_id=allRow[i]['id']
				if user_id:
					cursor2 = connection.cursor()
					sql2 = "select count(*) as save_count from adminpanel_opportunitysaved WHERE opportunity_id="+str(o_id)+" and user_id="+str(user_id)
					cursor2.execute(sql2)
					savedopportunities = dictfetchall(cursor2)
					allRow[i]['is_saved']=savedopportunities

					question_cursor = connection.cursor()
					question_sql = "select * from adminpanel_opportunityquestions where opportunity_id ="+str(o_id)
					question_cursor.execute(question_sql)
					questions = dictfetchall(question_cursor)
					allRow[i]['questions'] = questions

					shared_conn = connection.cursor()
					shared_sql = "select count(*) from adminpanel_opportunityshared where opportunity_id = " + str(o_id)
					shared_conn.execute(shared_sql)
					shared_res = dictfetchall(shared_conn)
					# print(shared_res)
					if shared_res[0]['count'] > 0:
						allRow[i]['shared'] = 1
					else:
						allRow[i]['shared'] = 0
				else:
					allRow[i]['shared'] = 0

				if user_id:
					if allRow[i]['uid'] == str(user_id):
						allRow[i]['is_apply'] = 0
					else:
						allRow[i]['is_apply'] = 1
						apply_cursor = connection.cursor()
						apply_sql = "select * from adminpanel_cordinatorrequest where oppurtunity_id ="+str(o_id)+" and user_id_id = "+str(user_id)
						apply_cursor.execute(apply_sql)
						applystatus = dictfetchall(apply_cursor)
						if applystatus:
							allRow[i]['is_apply'] = 2
							allRow[i]['application_status'] = applystatus[0]['status']
				else:
					allRow[i]['is_apply'] = 0
			
				allRow[i]['secret_message'] =''


		data = json.dumps({"Ack": 1, "opportunities": allRow, "idArray": idArray, "user_details":user_details,'related_organization':related_organization,"image_url": settings.BASE_URL + "/media/"})
	else :
		data = json.dumps({"Ack": 0, "msg": "Only POST method allowed","user_details":user_details})
	return HttpResponse(data)

class copy_opportunity(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		opportunityId = request.POST.get('opportunityId', None)
		copyAddress = request.POST.get('copyAddress', None)
		copyStartDate = request.POST.get('copyStartDate',None)
		copyEndDate = request.POST.get('copyEndDate',None)
		
		cursor = connection.cursor()
		# sql = "SELECT * FROM adminpanel_opportunities WHERE id="+opportunityId
		# cursor.execute(sql)
		# all_objs = dictfetchall(cursor)
		# return HttpResponse(opportunityId)
		all_objs = Opportunities.objects.filter(id=str(opportunityId)).values('user_id_id','org_id_id','opportunity_name','description','no_of_volunteers','author_name','address','start_date','end_date','parent_id','image','lat','lon')
		 

		user_id = all_objs[0]['user_id_id']
		org_id = all_objs[0]['org_id_id']
		opportunity_name = all_objs[0]['opportunity_name']
		description = all_objs[0]['description']
		no_of_volunteers = all_objs[0]['no_of_volunteers']
		author_name = all_objs[0]['author_name']
		if copyAddress:
			address = copyAddress
		else:
			address = all_objs[0]['address']

		if copyStartDate:
			start_date = copyStartDate
		else:
			start_date = all_objs[0]['start_date']

		if copyEndDate:
			end_date = copyEndDate
		else:
			end_date = all_objs[0]['end_date']
		

		parent_id = opportunityId 
		no_ofyear = all_objs[0]['parent_id'] 
				
		if len(request.FILES) != 0:
			image_file = ''
		else :
			image_file = all_objs[0]['image'] 


		org_id_id = all_objs[0]['org_id_id']

		user = User.objects.get(id=user_id)
		organization = Organization.objects.get(id = org_id_id)
		

		opportunitiy = Opportunities(
			user_id = user,
			org_id = organization,
			opportunity_name = opportunity_name,
			description = description,
			author_name = author_name,
			address = address,
			parent_id = parent_id,
			no_ofyear = no_ofyear,
			parent_opportunity = parent_id,
			start_date = start_date,
			end_date = end_date,
			image = image_file,
			lat = all_objs[0]['lat'],
			lon = all_objs[0]['lon']
		)

		opportunitiy.save()

		if image_file:
			imgthumb = Image.open(settings.MEDIA_ROOT+"/"+opportunitiy.image.name)
			imgthumb.save(settings.MEDIA_ROOT + '/' + opportunitiy.image.name)
			opportunitiy.image = opportunitiy.image.name
			opportunitiy.save()

		opportunityObj = Opportunities.objects.latest('id')
		#return HttpResponse(opportunityObj.id)

		sqlCategory = "SELECT * FROM adminpanel_opportunitycategories WHERE opportunity_id_id = "+str(opportunityObj.id)
		cursor.execute(sqlCategory)
		allCategories = dictfetchall(cursor)

		for cat in allCategories:
			catObj = ActivityCategory.objects.get(id = cat['category_id'])
			opportunityCategories = OpportunityCategories(
				opportunity_id = opportunityObj,
				category_id = catObj
			)
			opportunityCategories.save()

		sqlReminders = "SELECT * FROM adminpanel_reminders WHERE opportunity_id_id = "+str(opportunityObj.id)
		cursor.execute(sqlReminders)
		allReminders = dictfetchall(cursor)


		for reminder in allReminders:
			rem = Reminders(
				start_date = reminder['date'],
				before_hour = reminder['hour'],
				opportunity_id = opportunityObj
			)
			rem.save()

		sqlQuestion = "SELECT * FROM adminpanel_opportunityquestions WHERE opportunity_id = "+str(opportunityObj.id)
		cursor.execute(sqlQuestion)
		allQuestion = dictfetchall(cursor)



		for question in allQuestion:
			question = OpportunityQuestions(
				opportunity = opportunityObj,
				question = question['question']
			)
			question.save()

		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)



class delete_opportunity(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from datetime import datetime
		from django.core.serializers.json import DjangoJSONEncoder
		data={} 
		id = request.POST.get('id', None)
		#user_id = request.POST.get('user_id', None)
		cursor = connection.cursor()
		sql = ""
		is_delete = request.POST.get('is_delete', None)#1 means all 0 means single

		sql="select count(adminpanel_cordinatorrequest.id) from adminpanel_cordinatorrequest Where adminpanel_cordinatorrequest.oppurtunity_id ="+id
		cursor.execute(sql)
		check_objs = dictfetchall(cursor)
		
		if check_objs[0]['count'] != 0:
			sql = "DELETE FROM adminpanel_volunteersactivities WHERE adminpanel_volunteersactivities.opportunity_id="+str(id)
			cursor.execute(sql)

			if check_objs[0]['count'] >= 1:
				
				sql = "DELETE FROM adminpanel_cordinatorrequest WHERE adminpanel_cordinatorrequest.oppurtunity_id="+str(id)
				cursor.execute(sql)

		sql1="select count(adminpanel_opportunityshared.opportunity_id) from adminpanel_opportunityshared Where adminpanel_opportunityshared.opportunity_id ="+str(id)
		cursor.execute(sql1)
		check_objs1 = dictfetchall(cursor)
		if check_objs1[0]['count'] >= 1:
			sql3 = "DELETE FROM adminpanel_opportunityshared WHERE adminpanel_opportunityshared.opportunity_id="+str(id)
			cursor.execute(sql3)

		sql2="select count(adminpanel_opportunitysaved.opportunity_id) from adminpanel_opportunitysaved Where adminpanel_opportunitysaved.opportunity_id ="+str(id)
		cursor.execute(sql2)
		check_objs2 = dictfetchall(cursor)
		if check_objs2[0]['count'] >= 1:
			sql4 = "DELETE FROM adminpanel_opportunitysaved WHERE adminpanel_opportunitysaved.opportunity_id="+str(id)
			cursor.execute(sql4)
		
		sql5="select count(adminpanel_opportunitycategories.opportunity_id_id) from adminpanel_opportunitycategories Where adminpanel_opportunitycategories.opportunity_id_id ="+str(id)
		cursor.execute(sql5)
		check_objs5 = dictfetchall(cursor)
		if check_objs5[0]['count'] >= 1:
			sql6= "DELETE FROM adminpanel_opportunitycategories WHERE adminpanel_opportunitycategories.opportunity_id_id="+str(id)
			cursor.execute(sql6)

		sql22 = "select count(adminpanel_opportunityanswers.opportunity_id) from adminpanel_opportunityanswers where adminpanel_opportunityanswers.opportunity_id ="+str(id)
		cursor.execute(sql22)
		check_objs22 = dictfetchall(cursor)
		if check_objs22[0]['count'] >= 1:
			sql23= "DELETE FROM adminpanel_opportunityanswers WHERE adminpanel_opportunityanswers.opportunity_id="+str(id)
			cursor.execute(sql23)



		sql7 = "select count(adminpanel_opportunityquestions.opportunity_id) from adminpanel_opportunityquestions where adminpanel_opportunityquestions.opportunity_id ="+str(id)
		cursor.execute(sql7)
		check_objs7 = dictfetchall(cursor)
		if check_objs7[0]['count'] >= 1:
			sql8= "DELETE FROM adminpanel_opportunityquestions WHERE adminpanel_opportunityquestions.opportunity_id="+str(id)
			cursor.execute(sql8)

		sql9 = "select count(adminpanel_reminders.opportunity_id_id) from adminpanel_reminders where adminpanel_reminders.opportunity_id_id ="+str(id)
		cursor.execute(sql9)
		check_objs9 = dictfetchall(cursor)
		if check_objs9[0]['count'] >= 1: 
			sql10= "DELETE FROM adminpanel_reminders WHERE adminpanel_reminders.opportunity_id_id="+str(id)
			cursor.execute(sql10)

		
		if is_delete == '1':
			sql = "DELETE FROM adminpanel_opportunities WHERE adminpanel_opportunities.parent_opportunity="+str(id)
			cursor.execute(sql)
			##26.2.2020
			# child_collection = Opportunities.objects.filter(parent_opportunity=id)
			
			# if len(child_collection) == 0 :
			# 	op_obj = Opportunities.objects.get(id=id)
			# 	if op_obj.parent_opportunity != 0 :
			# 		delete_childs = Opportunities.objects.filter(parent_opportunity=op_obj.parent_opportunity).delete()
			# 		try:
			# 			Opportunities.objects.get(id=op_obj.parent_opportunity).delete()
			# 		except:
			# 			pass

			##26.2.2020
		sql = "DELETE FROM adminpanel_opportunities WHERE adminpanel_opportunities.id="+str(id)
		cursor.execute(sql)

		data= json.dumps({"Ack":1, "message":"Successfully Deleted"})
		return HttpResponse(data)
		
	def get(self,request):
		data = json.dumps({"Ack":0},cls=DjangoJSONEncoder)	
		return HttpResponse(data)

class get_opportunity_by_id(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from datetime import datetime
		from django.core.serializers.json import DjangoJSONEncoder
		id = request.POST.get('id', None)
		user_id = request.POST.get('user_id', None)

		related_organization = []
		if user_id:
			cursor13 = connection.cursor()
			sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org = 0 and adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.status='Approved'"
			cursor13.execute(sql3)
			related_organization = dictfetchall(cursor13)

		cursor = connection.cursor()
		user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
		cursor.execute(user_details_sql)
		user_details = dictfetchall(cursor)

		sql = "select ao.id,ao.user_id_id,ao.opportunity_name, ao.org_id_id, ao.description, ao.address, ao.image,ao.no_of_volunteers, ao.start_date as start_date, ao.end_date as end_date, au.first_name, au.last_name, ausr.profile_image from adminpanel_opportunities ao left join auth_user au on ao.user_id_id = au.id left join adminpanel_userprofile ausr on ao.user_id_id = ausr.user_id_id WHERE ao.id="+id
		
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		if opportunity:

			start_date=opportunity[0]['start_date']
			end_date=opportunity[0]['end_date']
			formatted_start_date = start_date.strftime("%b %d %Y %H:%M:%S")
			formatted_end_date = end_date.strftime("%b %d %Y %H:%M:%S")
			opportunity[0]['start_date'] = start_date.strftime("%Y-%m-%d %H:%M:%S")
			opportunity[0]['end_date'] = end_date.strftime("%Y-%m-%d %H:%M:%S") 
				
			c = connection.cursor()
			
			catSql = "SELECT aac.id, aac.name, aac.description, aac.category_image FROM adminpanel_opportunitycategories aoc LEFT JOIN adminpanel_activitycategory aac ON aac.id = aoc.category_id_id WHERE aoc.opportunity_id_id = '"+ id +"'"
			c.execute(catSql)
			cats = dictfetchall(c)

			orgid = str(opportunity[0]['org_id_id'])

			organization_cursor = connection.cursor()
			organization_sql = "select * from adminpanel_organization WHERE affiliated_org=0 and id='"+ orgid +"'"
			organization_cursor.execute(organization_sql)
			organization = dictfetchall(organization_cursor)

			# org_name = organization[0]['organization_name'] 

			question_cursor = connection.cursor()
			question_sql = "select * from adminpanel_opportunityquestions where opportunity_id ="+id
			question_cursor.execute(question_sql)
			questions = dictfetchall(question_cursor)

			applied = {"status":"","is_apply":0}
			saved = {"save_count": 0}
			shared = 0


			if user_id != None:
				conn = connection.cursor()
				check_sql = "select * from adminpanel_cordinatorrequest where user_id_id = " + user_id + " and oppurtunity_id = " + id
				conn.execute(check_sql)
				user_objs = dictfetchall(conn)
				if len(user_objs) > 0:
					applied['status'] = user_objs[0]['status']
					applied['is_apply']= 1

				conn1 = connection.cursor()
				check_saved_sql = "select count(*) as save_count from adminpanel_opportunitysaved where user_id = " + user_id + " AND opportunity_id = " + id
				conn1.execute(check_saved_sql)
				check_objs = dictfetchall(conn1)
				saved = check_objs

				shared_conn = connection.cursor()
				shared_sql = "select count(*) from adminpanel_opportunityshared where opportunity_id = " + str(id)
				shared_conn.execute(shared_sql)
				shared_res = dictfetchall(shared_conn)
				# print(shared_res)
				if shared_res[0]['count'] > 0:
					shared = 1
				else:
					shared = 0
				
				if opportunity[0]['user_id_id'] == user_id:
					applied['is_apply'] = 0
				else:
					# applied['is_apply']= 1
					apply_cursor = connection.cursor()
					apply_sql = "select * from adminpanel_cordinatorrequest where oppurtunity_id ="+str(id)+" and user_id_id = "+str(user_id)
					apply_cursor.execute(apply_sql)
					applystatus = dictfetchall(apply_cursor)
					if applystatus:
						applied['is_apply'] = 2
						application_status = applystatus[0]['status']
				
			reminder_conn = connection.cursor()
			reminders_sql = "select id, start_date as date, before_hour as hour from adminpanel_reminders where opportunity_id_id = " + id
			reminder_conn.execute(reminders_sql)
			reminders = dictfetchall(reminder_conn)
			o_id = id
			
			sql2 = "select adminpanel_userprofile.profile_image, auth_user.date_joined, auth_user.id,auth_user.first_name, auth_user.last_name,adminpanel_cordinatorrequest.status from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id= adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
			
			# print(sql2)
			# return HttpResponse(1)
			cursor.execute(sql2)
			volunteers = dictfetchall(cursor)

			data = json.dumps({"Ack": 1,"start_date": formatted_start_date, "end_date" : formatted_end_date,"organization":organization, "opportunity": opportunity, "image_url": settings.BASE_URL + "/media/", "cats": cats, "questions": questions, "applied": applied, "saved": saved, "shared":shared, "reminders": reminders, "user_details":user_details, "volunteers": volunteers, "related_organization" : related_organization},cls=DjangoJSONEncoder)
		else:
			data = json.dumps({"Ack": 0})
		return HttpResponse(data)

###############Done4################
# def submit_volenteer_request(request):
class submit_volenteer_request(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		id = request.POST.get('id', None)
		user_id = request.POST.get('user_id', None)
		role = request.POST.get('role',None)
		opportunity_details = Opportunities.objects.filter(id=id).values('org_id_id','opportunity_name','description','no_of_volunteers','author_name','address','parent_id','no_ofyear','no_ofyear','start_date','end_date','image')

		getorgId=opportunity_details[0]['org_id_id']

		apply_coordinator = CordinatorRequest(
			user_id_id=user_id,
			org_id_id=getorgId,
			status="Pending",
			address='N/A',
			employee_number=0,
			oppurtunity_id=id,
			role=role
			)
		apply_coordinator.save()
		data = json.dumps({"Ack": 1, "msg": "Volunteer Applied"})
	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

###############Done4################
class get_opportunity_by_user_date(APIView):

	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		import datetime
		user_id = request.POST.get('user_id', None)
		event_date = request.POST.get('event_date', None)
		opportunity_details = Opportunities.objects.filter(user_id_id=user_id, start_date = event_date)
		opportunity = fetch_row_opportunity(opportunity_details) 

		data = json.dumps({"Ack": 1, "opportunities": opportunity})
		return HttpResponse(data)

###############Done2################
class add_activity(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		activity_name = request.POST.get('activity_name', None)
		description = request.POST.get('description', None)
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		opportunity_id = request.POST.get('opportunity_id', None)

		activities = Activities(
			opportunity = Opportunities.objects.get(id = opportunity_id),
			activity_name = activity_name,
			author_name = '',
			address = '',
			start_date = start_date,
			end_date = end_date,
			description = description
		)
		activities.save()
		data = json.dumps({"Ack": 1, "msg": "Activity successfully saved"})
		return HttpResponse(data)

###############Done4################
class edit_activity(APIView):
	permission_classes = (IsAuthenticated,)
	def post(self,request):
		import json
		id = request.POST.get('id', None)
		activity_name = request.POST.get('activity_name', None)
		description = request.POST.get('description', None)
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)

		# cursor1 = connection.cursor()
		# cursor1.execute("update adminpanel_activities SET activity_name='"+activity_name+"', description='"+description+"', start_date='"+start_date+"', end_date='"+end_date+"' WHERE id='"+id+"'")
		
		Activities.objects.filter(id=str(id)).update(activity_name = str(activity_name), description= str(description), start_date= str(start_date), end_date= str(end_date))
			
		data = json.dumps({"Ack": 1, "msg": "Activity updated successfully"})
		return HttpResponse(data)

	

def get_activities(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		opportunity_id = request.POST.get('opportunity_id', None)

		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_activities WHERE opportunity_id="+opportunity_id
		cursor.execute(sql)
		activities = dictfetchall(cursor)

		data = json.dumps({"Ack": 1, "activities": activities}, cls=DjangoJSONEncoder)
	else :
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

###############Done4################
def get_activity(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		activity_id = request.POST.get('activity_id', None)

		activities = Activities.objects.filter(opportunity_id = opportunity_id).values('start_date','end_date','id','activity_name','description')

		created_date = str(activities[0]['start_date'])
		dt = created_date.split(' ')
		start_date = dt[0]
		tm = dt[1].split('+')
		start_time = tm[0]

		end_date = str(activities[0]['end_date'])
		dt1 = end_date.split(' ')
		tm1 = dt1[1].split('+')
		end_time = tm1[0]
		
		activity = {}
		activity['id'] = activities[0]['id']
		activity['activity_name'] = activities[0]['activity_name']
		activity['description'] = activities[0]['description']
		activity['start_date'] = start_date
		activity['start_time'] = start_time
		activity['end_time'] = end_time

		data = json.dumps({"Ack": 1, "activity": activity}, cls = DjangoJSONEncoder)
	else :
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

###############Done4################
# def add_volunteersactivities(request):
class add_volunteersactivities(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json

		user_id = request.POST.get('user_id', None)
		opportunity_id = request.POST.get('opportunity_id', None)
		activity_id = request.POST.get('activity_id', None)
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)

		volunteersactivities = Volunteersactivities(
			user = User.objects.get(id = user_id),
			opportunity = Opportunities.objects.get(id = opportunity_id),
			activity = Activities.objects.get(id = activity_id),
			start_date = start_date,
			end_date = end_date
		)

		volunteersactivities.save()
		data = json.dumps({"Ack": 1, "msg": "Volunteers activity saved successfully"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
		return HttpResponse(data)

def get_volunteersactivities(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		activity_id = request.POST.get('activity_id', None)

		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_volunteersactivities WHERE activity_id="+activity_id
		cursor.execute(sql)
		activities = dictfetchall(cursor)

		data = json.dumps({"Ack": 1, "activities": activities}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

class get_statistics(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):

		import json
		from django.core.serializers.json import DjangoJSONEncoder
		totalData = []
		opportunity_id = request.POST.get('opportunity_id', None)

		opportunityCursor = connection.cursor()
		opportunitySql = "SELECT * FROM adminpanel_opportunities WHERE id="+opportunity_id
		opportunityCursor.execute(opportunitySql)
		opportunityObjs = dictfetchall(opportunityCursor)

		for obj in opportunityObjs:
			opportunityObj = obj


		cursor = connection.cursor()
		sql = "SELECT id, activity_name FROM adminpanel_activities WHERE opportunity_id="+opportunity_id+" ORDER BY id ASC"
		cursor.execute(sql)
		activities = dictfetchall(cursor)
		activityHeader = ['Genre']

		for act in activities:
			activityHeader.append(act['activity_name'])
		totalData.append(activityHeader)

		useridsCursor = connection.cursor()
		useridsSql = "SELECT DISTINCT user_id FROM adminpanel_volunteersactivities WHERE activity_id IN (SELECT id FROM adminpanel_activities WHERE opportunity_id="+opportunity_id+")"
		useridsCursor.execute(useridsSql)
		userids = dictfetchall(useridsCursor)
		# print(userids)

		for userid in userids:
			vactivitiesCursor = connection.cursor()
			vactivitiesSql = "SELECT DISTINCT va.activity_id, CONCAT(au.first_name, ' ', au.last_name) full_name, EXTRACT(epoch FROM va.end_date - va.start_date)/3600 tothr FROM adminpanel_volunteersactivities va LEFT JOIN auth_user au ON va.user_id = au.id WHERE va.activity_id IN (SELECT id FROM adminpanel_activities WHERE opportunity_id="+opportunity_id+") AND va.user_id="+str(userid['user_id'])+" ORDER BY va.activity_id ASC"
			vactivitiesCursor.execute(vactivitiesSql)
			vactivities = dictfetchall(vactivitiesCursor)

			isFullname = False
			activityBody = []

			for item in activities:
				exist = None

				for index, vact in enumerate(vactivities):
					if isFullname == False:
						isFullname = True
						activityBody.append(vact['full_name'])

					if vact['activity_id'] == item['id']:
						exist = index

				if exist != None:
					activityBody.append(vactivities[exist]['tothr'])
				else:
					activityBody.append(0)

			totalData.append(activityBody)

		data = json.dumps({"Ack": 1, "data": totalData, "opportunity": opportunityObj}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
		return HttpResponse(data)

###############Done5################
class get_opportunities_by_user(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		from django.core.serializers.json import DjangoJSONEncoder
		import json

		user_id = request.POST.get('user_id', None)
		opportunity_details = Opportunities.objects.filter(user_id_id=str(user_id))
		opportunityObjs = fetch_row_opportunity(opportunity_details) 
		data = json.dumps({"Ack": 1, "opportunities": opportunityObjs}, cls=DjangoJSONEncoder)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
		return HttpResponse(data)

class get_organization_by_user(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id', None)
		organizationCursor = connection.cursor()
		organizationSql = "select adminpanel_organization.*,adminpanel_cordinatorrequest.* from adminpanel_organization inner join adminpanel_cordinatorrequest on adminpanel_organization.id=adminpanel_cordinatorrequest.org_id_id where  adminpanel_organization.affiliated_org=0 and adminpanel_cordinatorrequest.user_id_id="+user_id
		organizationCursor.execute(organizationSql)
		organizationObjs = dictfetchall(organizationCursor)
		data = json.dumps({"Ack": 1, "organizations": organizationObjs}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
		return HttpResponse(data)

###############Done2################
# def apply_volunteer(request):
class apply_volunteer(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json

		user_id = request.POST.get('user_id', None)
		opportunity_id = request.POST.get('opportunity_id', None)
		description = request.POST.get('description', None)
		
		volunteerrequirement = VolunteerRequirement(
			user_id_id = user_id,
			opportunity_id_id = opportunity_id,
			description_json = description
		)
		
		volunteerrequirement.save()
		data = json.dumps({"Ack": 1, "msg": "Volunteer Applied"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

class get_volunteersbyopportunity(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder

		oppurtunity_id = request.POST.get('opportunity_id', None)

		# ###Count Approved voluteer list
		# cursor1 = connection.cursor()
		# sql1  = "select count(*) from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id='"+str(oppurtunity_id)+"' and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
		# cursor1.execute(sql1)
		# Volunteer_count = dictfetchall(cursor1)
		# data = json.dumps({"Ack": 0, "msg": Volunteer_count})
		# return HttpResponse(data)

		cursor = connection.cursor()
		opportunity_details = Opportunities.objects.filter(id=oppurtunity_id).values('org_id')
		org_id_id = opportunity_details[0]['org_id']
		sql = "select DISTINCT(auth_user.email) AS email, RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name)))) AS full_Name, adminpanel_cordinatorrequest.org_id_id from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.org_id_id="+str(org_id_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'" 

		



		cursor.execute(sql)
		activities = dictfetchall(cursor)
		data = json.dumps({"Ack": 1, "voluntererArr": activities, "oppurtunity_id":oppurtunity_id,"org_id_id":org_id_id}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):

		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
		return HttpResponse(data)

###############Done2################
class update_request(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		rtn_obj = {}
		msg_html2 = ""

		id = request.POST.get('id',None)
		status = request.POST.get('status',None)		
		CordinatorRequest.objects.filter(id=str(id)).update(status = str(status)) 
		updatedetails = CordinatorRequest.objects.filter(id=id)
		opportunity_id = updatedetails[0].oppurtunity_id

		if opportunity_id:
			link = settings.PATH_URL+"/opportunity-details/"+str(opportunity_id)

		cursor = connection.cursor()
		sql = "select AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and AO.id ="+str(opportunity_id) 
		cursor.execute(sql)
		opportunity_details = dictfetchall(cursor)

		userprofiles= User.objects.filter(id=updatedetails[0].user_id_id) 
		# userprofiles = User.objects.filter(id=str(updatedetails[0].user_id_id))
		if status == 'Approved':
			# msg_html2 = '<p>Congratulations!</p> <p>Your application has been approved.</p> <p>Thanks,</p><p>Inteer Team</p>'
			msg_html2 += '<p>Congratulations!</p><p>Your application has been approved.</p> <a href="'+link+'"></p><div style="width: 300px; border: 1px solid #d8d8d8;"><div style="width: 300px; height: 180px;">'
			if opportunity_details[0]['image']: 
				msg_html2 += '<img src="'+settings.UPLOAD_IMAGES_URL_ROOT+opportunity_details[0]['image']+'" style="width: 100%; height: 100%;">'
				# msg_html2 += '<img src="'+settings.IMAGE_URL+opportunity_details[0]['image']+'" style="width: 100%; height: 100%;">'
				print(msg_html2)
			else:
				msg_html2 +='<img src="'+settings.APP_URL+'"assets/img/default.png" style="width: 100%; height: 100%;">' 
			
			msg_html2 +='</div><div style="padding: 0px 20px 20px;"><div style="display: flex; align-items: center;"><div style="margin-right: 10px;"><p style="line-height: 35px;background: #fff;border: 1px solid #d8d8d8; width: 35px; height: 35px; text-align: center; border-radius: 50%;">TA</p></div><div><p style="margin: 0px; font-family: "Open Sans", sans-serif; ">'+opportunity_details[0]['organization_name']+'</p><small>'+opportunity_details[0]['first_name']+' '+opportunity_details[0]['last_name']+'</small></div></div><div><h3 style="font-family: Open Sans, sans-serif; margin: 0px; font-size: 13px;">'+opportunity_details[0]['opportunity_name']+'</h3><p style="font-family: Open Sans, sans-serif; margin: 0px; margin-top: 5px; font-size: 13px; color: #676767;">'+opportunity_details[0]['description']+'</p></div></div></div></a><p>Thanks,</p><p>Inteer Team</p>'

			send_mail("Inteer Opportunity Application Approved", 'Inteer','ainteer72@gmail.com',[userprofiles[0].email],html_message=msg_html2 )
			# send_mail("Inteer Opportunity Application Approved", 'Inteer','ainteer72@gmail.com',['maitrayee.bhaumik@cbnits.com'],html_message=msg_html2 )
		
				
		rtn_obj['msg'] = "status Updated successfully!"
		rtn_obj['Ack'] = "1"
		rtn_obj['status'] = status
		data = json.dumps(rtn_obj)
		return HttpResponse(data)
	
	def get(self,request):

		rtn_obj = {}
		rtn_obj['msg'] = "input data by post method"
		rtn_obj['Ack'] = "0"
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

# def fetch_all_oppo(request):
class fetch_all_oppo(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		data = {}
		user_id = request.POST.get('user_id', None)
		filter_type = request.POST.get('filter_type', None)
		
		cursor = connection.cursor()
		if filter_type == "saved" :
			cursor = connection.cursor()
			sql = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" ORDER BY date(adminpanel_opportunities.start_date) ASC"

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)
			coordinate_data = []
			
			
			if opportunity:
				for i in range(len(opportunity)):
					if opportunity[i]['start_date'] not in coordinate_data: 
						coordinate_data.append(opportunity[i]['start_date'])

				data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			else :
				data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
		elif filter_type == "shared" :
			cursor = connection.cursor();
			org_id =0
			# return HttpResponse(user_id)
			sql_orgid = "SELECT * FROM adminpanel_cordinatorrequest WHERE user_id_id="+str(user_id)+" AND is_request !='' AND status='Approved'"
			cursor.execute(sql_orgid)
			org_id_details = dictfetchall(cursor)
			# return HttpResponse(len(org_id_details)) 
			if len(org_id_details) > 1:
				# print('hereif')
				full_data = []
				idArray = []
				idArray2=[]
				coordinate_data=[]
				k=0
				for j in range(len(org_id_details)):
					org_id = org_id_details[j]['org_id_id']
					
					if org_id not in idArray:
						idArray.append(org_id)
				for org in idArray:
					sql_shared = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunities LEFT JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" ORDER BY date(adminpanel_opportunities.start_date) DESC"
				
					cursor.execute(sql_shared)
					opportunity = dictfetchall(cursor)
					print(opportunity)
					v_data={}
					
					if opportunity:
						for i in range(len(opportunity)):
							if opportunity[i]['start_date'] not in coordinate_data:
								coordinate_data.append(opportunity[i]['start_date'])
								
				data = json.dumps({"Ack": 1,"event_data":coordinate_data}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			elif len(org_id_details) == 1:
				print('hereelif')
				org_id = org_id_details[0]['org_id_id']
				sql = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org_id)+" ORDER BY date(adminpanel_opportunities.start_date) ASC"
			

				cursor.execute(sql)
				opportunity = dictfetchall(cursor)
				# return HttpResponse(len(opportunity)) 
				v_data={}
				coordinate_data=[]
				if opportunity:
					for i in range(len(opportunity)):
						if opportunity[i]['start_date'] not in coordinate_data:
							coordinate_data.append(opportunity[i]['start_date'])
							

					data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
				else :
					data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
					return HttpResponse(data) 
			else:
				print('hereelse')  
				cursor = connection.cursor()
				sql_shared = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.user_id ="+str(user_id)+" ORDER BY date(adminpanel_opportunities.start_date) DESC"
			
				cursor.execute(sql_shared)
				opportunity = dictfetchall(cursor)
				coordinate_data = {}
				
				
				if opportunity:
					for i in range(len(opportunity)):
						if opportunity[i]['start_date'] not in coordinate_data:
							coordinate_data.append(opportunity[i]['start_date'])

					data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
				else :
					data = json.dumps({"Ack": 0, "event_data":coordinate_data, "is_saved": is_saved, "is_shared": is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
		elif filter_type == "applied" :
			
			cursor = connection.cursor()
			sql = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" ORDER BY date(adminpanel_opportunities.start_date) ASC"

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)

			v_data={}
			coordinate_data=[]
			if opportunity:
				for i in range(len(opportunity)):
					if opportunity[i]['start_date'] not in coordinate_data: 
						coordinate_data.append(opportunity[i]['start_date'])

			data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else:
			cursor = connection.cursor(); 
			sql = "SELECT DISTINCT(date(adminpanel_opportunities.start_date)) as start_date FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" ORDER BY date(adminpanel_opportunities.start_date) ASC"
			

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)

			v_data={}
			coordinate_data=[]
			idArray1 = []
			if opportunity:
				for i in range(len(opportunity)):
					if opportunity[i]['start_date'] not in coordinate_data: 
						coordinate_data.append(opportunity[i]['start_date'])
					# oppurtunityArr=opportunity[i]
					# coordinate_data[i]=oppurtunityArr	

				data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			else :
				data = json.dumps({"Ack": 1, "event_data":coordinate_data}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
	def get(self,request) :
		data = {}
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class get_opportunity_volunteer_by_user_date(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request,offset):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		from datetime import date
		import time
		from time import gmtime, strftime
		import pytz

		data = {}
		user_id = request.POST.get('user_id', None)
		filter_type = request.POST.get('filter_type', None)
		is_saved = request.POST.get('is_saved', None)
		is_shared = request.POST.get('is_shared', None)
		type_val = request.POST.get('type_val',None)
		time_zone = request.POST.get('time_zone',None)
		# return HttpResponse(type_val)
		limit = 12
		cursor = connection.cursor()
		user_details_sql = "SELECT role_id,user_id_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
		cursor.execute(user_details_sql)
		user_details = dictfetchall(cursor)
		tz_NY = pytz.timezone(time_zone)
		# tz_NY = pytz.timezone('Pacific/Honolulu') 
		# tz_NY = pytz.timezone('Etc/Greenwich') 
		#tz_NY = pytz.timezone('America/Rainy_River')
		# currentDate = datetime.datetime.now(tz_NY).date()
		currentDate = datetime.datetime.now(tz_NY).isoformat()
		currentDate = currentDate.rsplit(':', 2)[0]
		# return HttpResponse(currentDate)
		
		if filter_type == "date" : 
			event_date = request.POST.get('event_date', None)
			
			sql = ""
			if is_saved == '0':

				if type_val== "past":
					if is_shared == '0':
						sql = "select adminpanel_opportunities.id as opt_id,adminpanel_opportunities.*,adminpanel_cordinatorrequest.status from adminpanel_opportunities inner join adminpanel_cordinatorrequest on (adminpanel_cordinatorrequest.oppurtunity_id=adminpanel_opportunities.id and adminpanel_cordinatorrequest.user_id_id != adminpanel_opportunities.user_id_id) WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" and adminpanel_opportunities.user_id_id !="+str(user_id)+ " and date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC"
					if is_shared == '1':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunityshared LEFT JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id = adminpanel_opportunities.id WHERE date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' GROUP BY adminpanel_opportunities.id ORDER BY adminpanel_opportunities.start_date DESC"
				elif type_val== "upcoming":
					if is_shared == '0':
						sql = "select adminpanel_opportunities.id as opt_id,adminpanel_opportunities.*,adminpanel_cordinatorrequest.status from adminpanel_opportunities inner join adminpanel_cordinatorrequest on (adminpanel_cordinatorrequest.oppurtunity_id=adminpanel_opportunities.id and adminpanel_cordinatorrequest.user_id_id != adminpanel_opportunities.user_id_id) WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" and adminpanel_opportunities.user_id_id !="+str(user_id)+ " and date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
					if is_shared == '1':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunityshared LEFT JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id = adminpanel_opportunities.id WHERE date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' GROUP BY adminpanel_opportunities.id ORDER BY adminpanel_opportunities.start_date ASC"
				elif type_val== "default":
					if is_shared == '0':
						sql = "select adminpanel_opportunities.id as opt_id,adminpanel_opportunities.*,adminpanel_cordinatorrequest.status from adminpanel_opportunities inner join adminpanel_cordinatorrequest on (adminpanel_cordinatorrequest.oppurtunity_id=adminpanel_opportunities.id and adminpanel_cordinatorrequest.user_id_id != adminpanel_opportunities.user_id_id) WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" and adminpanel_opportunities.user_id_id !="+str(user_id)+ " and date(adminpanel_opportunities.start_date)='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
					if is_shared == '1':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunityshared LEFT JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id = adminpanel_opportunities.id WHERE date(adminpanel_opportunities.start_date)='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' GROUP BY adminpanel_opportunities.id ORDER BY adminpanel_opportunities.start_date ASC"
				else:
					if is_shared == '0':
						sql = "select adminpanel_opportunities.id as opt_id,adminpanel_opportunities.*,adminpanel_cordinatorrequest.status from adminpanel_opportunities inner join adminpanel_cordinatorrequest on (adminpanel_cordinatorrequest.oppurtunity_id=adminpanel_opportunities.id and adminpanel_cordinatorrequest.user_id_id != adminpanel_opportunities.user_id_id) WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" and adminpanel_opportunities.user_id_id !="+str(user_id)+ " and date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC"
					if is_shared == '1':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunityshared LEFT JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id = adminpanel_opportunities.id WHERE date(adminpanel_opportunities.start_date)='"+event_date +"' GROUP BY adminpanel_opportunities.id ORDER BY adminpanel_opportunities.start_date ASC"

			elif is_saved == '1':
				if type_val== "past":
					if is_shared == '0':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+ " AND adminpanel_opportunities.start_date='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC"
					else:
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+" AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC"
				elif type_val== "upcoming":
					if is_shared == '0':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+ " AND adminpanel_opportunities.start_date='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
					else:
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+" AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
				elif type_val== "default":
					if is_shared == '0':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+ " AND adminpanel_opportunities.start_date='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
					else:
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+" AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
				else:
					if is_shared == '0':
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+ " AND adminpanel_opportunities.start_date='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC"
					else:
						sql = "SELECT adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* FROM adminpanel_opportunitysaved LEFT JOIN adminpanel_opportunities ON  adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id WHERE adminpanel_opportunitysaved.user_id!="+str(user_id)+" ORDER BY adminpanel_opportunities.start_date ASC"
			else:
				if type_val== "past":
					sql = "select adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* from adminpanel_opportunities WHERE user_id_id="+str(user_id)+ " and date(start_date)='"+ event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC"
					# return HttpResponse(1)
				elif type_val== "upcoming":
					sql = "select adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* from adminpanel_opportunities WHERE user_id_id="+str(user_id)+ " and date(start_date)='"+ event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
				else:
					sql = "select adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* from adminpanel_opportunities WHERE user_id_id="+str(user_id)+ " and date(start_date)='"+ event_date +"' ORDER BY adminpanel_opportunities.start_date ASC"
			
			cursor.execute(sql)
			opportunity = dictfetchall(cursor)
			if (len(opportunity)==0):
				sql3 = "select adminpanel_opportunities.id as opt_id, adminpanel_opportunities.* from adminpanel_opportunities WHERE user_id_id="+str(user_id)+ " and date(start_date)='"+ event_date +"' ORDER BY adminpanel_opportunities.start_date ASC"
				cursor3 = connection.cursor()
				cursor3.execute(sql3)
				opportunity= dictfetchall(cursor3)	
			v_data={}
			coordinate_data={}
			if opportunity:
				for i in range(len(opportunity)):
					o_id=opportunity[i]['id']
					opportunity[i]['opportunity_id']=opportunity[i]['id']
					cursor3 = connection.cursor()
					sql3 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'" 
					cursor3.execute(sql3)
					pendinglist = dictfetchall(cursor3)
					opportunity[i]['pending']=pendinglist

					cursor4 = connection.cursor()
					sql4 = "select adminpanel_userprofile.profile_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
					cursor4.execute(sql4)
					approvelist = dictfetchall(cursor4)
					opportunity[i]['approve']=approvelist


					cursor2 = connection.cursor()
					sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
					cursor2.execute(sql2)
					volunteers = dictfetchall(cursor2)
					opportunity[i]['volunteers']=volunteers
					oppurtunityArr=opportunity[i]
					coordinate_data[i]=oppurtunityArr	

				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved":is_saved, "is_shared" : is_shared, "event_date" : event_date, "user_details":user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			else :
				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved":is_saved, "is_shared" : is_shared, "event_date" : event_date,"user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
		elif filter_type == "saved" :
			cursor = connection.cursor()
			event_date = request.POST.get('event_date', None)
			if type_val == "past":
				if event_date:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
			elif type_val == "upcoming":
				if event_date:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
			elif type_val == "default":
				if event_date:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
			else:

				if event_date:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitysaved ON adminpanel_opportunities.id = adminpanel_opportunitysaved.opportunity_id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)
			coordinate_data = {}
			
			
			if opportunity:
				for i in range(len(opportunity)):
					o_id=opportunity[i]['id']
					opportunity[i]['opportunity_id']=opportunity[i]['id']
					cursor3 = connection.cursor()
					sql3 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
					cursor3.execute(sql3)
					pendinglist = dictfetchall(cursor3)
					opportunity[i]['pending']=pendinglist

					cursor4 = connection.cursor()
					sql4 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
					cursor4.execute(sql4)
					approvelist = dictfetchall(cursor4)
					opportunity[i]['approve']=approvelist



					cursor2 = connection.cursor()
					sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
					cursor2.execute(sql2)
					volunteers = dictfetchall(cursor2)
					opportunity[i]['volunteers']=volunteers
					oppurtunityArr=opportunity[i]
					coordinate_data[i]=oppurtunityArr

				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved" : is_saved, "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			else :
				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved": is_saved, "is_shared": is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
		elif filter_type == "shared" :
			cursor = connection.cursor()
			event_date = request.POST.get('event_date', None)
			org_id =0
			# return HttpResponse(user_id)
			sql_orgid = "SELECT * FROM adminpanel_cordinatorrequest WHERE user_id_id="+str(user_id)+" AND is_request !='' AND status='Approved'"
			cursor.execute(sql_orgid)
			org_id_details = dictfetchall(cursor)
			# return HttpResponse(len(org_id_details)) 
			if len(org_id_details) > 1:
				full_data = []
				idArray = []
				idArray2=[]
				coordinate_data={}
				k=0
				for j in range(len(org_id_details)):
					org_id = org_id_details[j]['org_id_id']
					
					if org_id not in idArray:
						idArray.append(org_id)
				for org in idArray:
					# return HttpResponse(org)
					if type_val == "past":
						if event_date:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset) 
						else:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
					elif type_val == "upcoming":
						if event_date:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset) 
						else:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
					elif type_val == "default":
						if event_date:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset) 
						else:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
					else:
						if event_date:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset) 
						else:
							sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.org_id_id ="+str(org)+" ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
				
					cursor.execute(sql_shared)
					opportunity = dictfetchall(cursor)
					# data = json.dumps({"Ack": 1,"event_data":sql_shared}, cls=DjangoJSONEncoder)
					# return HttpResponse(data)
					# return HttpResponse(len(opportunity)) 
					# print(sql_shared)
					# return HttpResponse(1)
					v_data={}
					# coordinate_data={}
					if opportunity:
						for i in range(len(opportunity)):
							o_id=opportunity[i]['id']
							opportunity[i]['opportunity_id']=opportunity[i]['id']

							cursor3 = connection.cursor()
							sql3 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
							cursor3.execute(sql3)
							pendinglist = dictfetchall(cursor3)
							opportunity[i]['pending']=pendinglist

							cursor4 = connection.cursor()
							sql4 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
							cursor4.execute(sql4)
							approvelist = dictfetchall(cursor4)
							opportunity[i]['approve']=approvelist

							cursor2 = connection.cursor()
							sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.user_id_id="+str(user_id)
							cursor2.execute(sql2)
							volunteers = dictfetchall(cursor2)
							opportunity[i]['volunteers']=volunteers

							if opportunity[i]['opportunity_id'] not in idArray2:
								idArray2.append(opportunity[i]['opportunity_id'])
								oppurtunityArr=opportunity[i]
								coordinate_data[k]=oppurtunityArr
								k+=1
					# Merge(full_data, coordinate_data) 
						full_data.append(oppurtunityArr)
				data = json.dumps({"Ack": 1,"idArray":idArray, "event_data":coordinate_data, "is_saved": is_saved , "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			elif len(org_id_details) == 1:
				org_id = org_id_details[0]['org_id_id']
				if event_date:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org_id)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC"
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.org_id_id ="+str(org_id)+" ORDER BY adminpanel_opportunities.start_date ASC"
			

				cursor.execute(sql)
				opportunity = dictfetchall(cursor)
				# return HttpResponse(len(opportunity)) 
				v_data={}
				coordinate_data={}
				if opportunity:
					for i in range(len(opportunity)):
						o_id=opportunity[i]['id']
						opportunity[i]['opportunity_id']=opportunity[i]['id']

						cursor3 = connection.cursor()
						sql3 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
						cursor3.execute(sql3)
						pendinglist = dictfetchall(cursor3)
						opportunity[i]['pending']=pendinglist

						cursor4 = connection.cursor()
						sql4 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
						cursor4.execute(sql4)
						approvelist = dictfetchall(cursor4)
						opportunity[i]['approve']=approvelist

						cursor2 = connection.cursor()
						sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.user_id_id="+str(user_id)
						cursor2.execute(sql2)
						volunteers = dictfetchall(cursor2)
						opportunity[i]['volunteers']=volunteers
						oppurtunityArr=opportunity[i]
						coordinate_data[i]=oppurtunityArr	

					data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved": is_saved , "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
				else :
					data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved" : is_saved, "is_shared" : is_shared, "event_date": event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
					return HttpResponse(data) 
			else:
				cursor = connection.cursor()
				event_date = request.POST.get('event_date', None)
				if event_date:
					sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.user_id ="+str(user_id)+" AND date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC" 
				else:
					sql_shared = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunityshared INNER JOIN adminpanel_opportunities ON adminpanel_opportunityshared.opportunity_id= adminpanel_opportunities.id WHERE adminpanel_opportunityshared.user_id ="+str(user_id)+" ORDER BY adminpanel_opportunities.start_date DESC"
			
				cursor.execute(sql_shared)
				opportunity = dictfetchall(cursor)
				coordinate_data = {}
				
				
				if opportunity:
					for i in range(len(opportunity)):
						o_id=opportunity[i]['id']
						opportunity[i]['opportunity_id']=opportunity[i]['id']
						cursor3 = connection.cursor()
						sql3 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
						cursor3.execute(sql3)
						pendinglist = dictfetchall(cursor3)
						opportunity[i]['pending']=pendinglist

						cursor4 = connection.cursor()
						sql4 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
						cursor4.execute(sql4)
						approvelist = dictfetchall(cursor4)
						opportunity[i]['approve']=approvelist



						cursor2 = connection.cursor()
						sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
						cursor2.execute(sql2)
						volunteers = dictfetchall(cursor2)
						opportunity[i]['volunteers']=volunteers
						oppurtunityArr=opportunity[i]
						coordinate_data[i]=oppurtunityArr

					data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved" : is_saved, "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
				else :
					data = json.dumps({"Ack": 0, "event_data":coordinate_data, "is_saved": is_saved, "is_shared": is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
					return HttpResponse(data)
		elif filter_type == "applied" :
			
			cursor = connection.cursor()
			event_date = request.POST.get('event_date', None)
			if type_val == "past":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date+"' AND adminpanel_opportunities.start_date < current_date ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
			elif type_val == "upcoming":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date+"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC"
					# return HttpResponse(sql)
			elif type_val == "default":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date+"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
			else:
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_cordinatorrequest ON adminpanel_opportunities.id = adminpanel_cordinatorrequest.oppurtunity_id WHERE adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)

			v_data={}
			coordinate_data={}
			if opportunity:
				for i in range(len(opportunity)):
					o_id=opportunity[i]['id']
					opportunity[i]['opportunity_id']=opportunity[i]['id']
					cursor3 = connection.cursor()
					sql3 = "select adminpanel_userprofile.profile_image as pending_user_image,auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
					cursor3.execute(sql3)
					pendinglist = dictfetchall(cursor3)
					opportunity[i]['pending']=pendinglist

					cursor4 = connection.cursor()
					sql4 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
					cursor4.execute(sql4)
					approvelist = dictfetchall(cursor4)
					opportunity[i]['approve']=approvelist
					cursor2 = connection.cursor()
					sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
					cursor2.execute(sql2)
					volunteers = dictfetchall(cursor2)
					opportunity[i]['volunteers']=volunteers
					oppurtunityArr=opportunity[i]
					coordinate_data[i]=oppurtunityArr	

			# data = json.dumps({"Ack": 1, "event_data":opportunity, "is_saved": is_saved , "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details, "message": "you haven't signed up for any opportunities yet. Go ahead and search them in Feed page"}, cls=DjangoJSONEncoder)
			data = json.dumps({"Ack": 1, "event_data":opportunity, "is_saved": is_saved , "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details, "message": "you haven't signed up for any opportunities yet. Go ahead and search them in Feed page"}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else:
			cursor = connection.cursor()
			event_date = request.POST.get('event_date', None)
			if type_val == "past":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities  WHERE adminpanel_opportunities.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC"
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" AND adminpanel_opportunities.start_date < '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date DESC LIMIT "+str(limit)+" OFFSET "+str(offset)
					# return HttpResponse(sql)
			elif type_val == "upcoming":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities  WHERE adminpanel_opportunities.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND adminpanel_opportunities.start_date >= '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" AND adminpanel_opportunities.start_date >= '"+ str(currentDate)  +"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
					# return HttpResponse(sql)
			elif type_val == "default":
				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities  WHERE adminpanel_opportunities.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" AND date(adminpanel_opportunities.start_date) = '"+str(currentDate)+"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
			else:

				if event_date != None:
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities  WHERE adminpanel_opportunities.user_id_id="+str(user_id)+ " AND date(adminpanel_opportunities.start_date)='"+event_date +"' ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)
				else:	
					sql = "SELECT adminpanel_opportunities.id as opt_id,adminpanel_opportunities.* FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" ORDER BY adminpanel_opportunities.start_date ASC LIMIT "+str(limit)+" OFFSET "+str(offset)

			cursor.execute(sql)
			opportunity = dictfetchall(cursor)

			v_data={}
			coordinate_data={}
			if opportunity:
				for i in range(len(opportunity)):
					o_id=opportunity[i]['id']
					opportunity[i]['opportunity_id']=opportunity[i]['id']
					cursor3 = connection.cursor()
					sql3 = "select adminpanel_userprofile.profile_image as pending_user_image,auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Pending'"
					cursor3.execute(sql3)
					pendinglist = dictfetchall(cursor3)
					opportunity[i]['pending']=pendinglist

					cursor4 = connection.cursor()
					sql4 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
					cursor4.execute(sql4)
					approvelist = dictfetchall(cursor4)
					opportunity[i]['approve']=approvelist
					cursor2 = connection.cursor()
					sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
					cursor2.execute(sql2)
					volunteers = dictfetchall(cursor2)
					opportunity[i]['volunteers']=volunteers
					oppurtunityArr=opportunity[i]
					coordinate_data[i]=oppurtunityArr	

				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved": is_saved , "is_shared" : is_shared, "event_date" : event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
			else :
				data = json.dumps({"Ack": 1, "event_data":coordinate_data, "is_saved" : is_saved, "is_shared" : is_shared, "event_date": event_date, "user_details": user_details}, cls=DjangoJSONEncoder)
				return HttpResponse(data)
	def get(self,request) :
		data = {}
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

##########Done3#############
# def fetch_all_answers(request):
class fetch_all_answers(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		opportunity_id = request.POST.get("opportunity_id",None)
		user_id = request.POST.get("user_id",None)

		fetch_all_answers= Opportunityanswers.objects.select_related('opportunity','user','question').filter(opportunity_id=opportunity_id,user_id = user_id)
		user_details_get = User.objects.select_related('userprofile').filter(id = user_id).values('email','username')

		all_answer = []
		for answr in fetch_all_answers:
			answer_array= {}
			answer_array['id'] = answr.id
			answer_array['user_id'] = answr.user_id
			answer_array['opportunity_id'] = answr.opportunity_id
			answer_array['answer'] = answr.answer

			question_value = OpportunityQuestions.objects.filter(id= answr.question_id)
			print(question_value[0].question)
			answer_array['question'] = question_value[0].question
			all_answer.append(answer_array)
		# print(fetch_all_answers)  

		data= json.dumps({"Ack":1,"all_answer":all_answer,"user_details_email":user_details_get[0]['email'],"user_details_username":user_details_get[0]['username'], "message":"Successfully fetched"})
		return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack":0},cls=DjangoJSONEncoder)	
		return HttpResponse(data)

##########done#############	
class delete_volunteer(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		request_id = request.POST.get("request_id",None)
		CordinatorRequest.objects.filter(id=str(request_id)).delete()
		data= json.dumps({"Ack":1, "message":"Successfully Deleted"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack":0},cls=DjangoJSONEncoder)	
		return HttpResponse(data)

##########Done3#############		
# def save_opportunity(request):
class save_opportunity(APIView):
	permission_classes = (IsAuthenticated,)
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		opportunity_id = request.POST.get('opportunity_id', None)
		user_id = request.POST.get('user_id', None)
		save_opportunity = OpportunitySaved(
			opportunity_id = opportunity_id,
			user_id = user_id
		)
		save_opportunity.save()
		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved"})
		return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)

# def opportunity_apply(request):
class opportunity_apply(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		org_id = request.POST.get('org_id', None)
		
		opportunity_id = request.POST.get('opportunity_id', None)
		# return HttpResponse(opportunity_id)

		user_id = request.POST.get('user_id', None)
		url = request.POST.get('url', None)
		
		if 'answer' in request.POST and request.POST.get('answer') != '':
			answers = request.POST.get('answer')
			answers = json.loads(answers)
		
		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_opportunities WHERE id="+str(opportunity_id); 
		cursor.execute(sql)
		opportunity_details = dictfetchall(cursor)
		# return HttpResponse(opportunity_details)
		owner_role = 0

		if 'answer' in request.POST and request.POST.get('answer') == '':
			opportunity_owner = opportunity_details[0]['user_id_id']
			opportunity_owner_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+str(opportunity_owner)
			cursor.execute(opportunity_owner_sql)
			opportunity_owner_details = dictfetchall(cursor)
			owner_role = opportunity_owner_details[0]['role_id']
		
		if 	owner_role == 2:
			organisation_details = Organization.objects.get(pk = org_id)
			organisation_name = organisation_details.organization_name
			userprofiles= User.objects.get(pk=opportunity_owner)

			# adminEmail = 'kher.nachiket@gmail.com'
			adminEmail = 'ainteer72@gmail.com'

			msg_html = '<p>Hi Admin,</p> <p> '+userprofiles.first_name+' '+userprofiles.last_name+' applied for an opportunity related to '+organisation_name+'.</p><p>Thanks,</p><p>Inteer Team</p>'

			send_mail('Notification For Admin', 'InterApp', 'ainteer72@gmail.com', ['ainteer72@gmail.com'], html_message=msg_html)
			data = json.dumps({"Ack": 3,"web_url": organisation_details.web_url, "msg": "once the coordination got access tour application will be applied"})
			return HttpResponse(data)
		else:

			data={}
			data['det'] = opportunity_details
			opparray = {}
			
			if opportunity_details:
				i=0
				for o_details in opportunity_details:
					if opportunity_details[i]['start_date'] != "":
						opparray['start_date']= o_details['start_date']
						opparray['end_date']= o_details['end_date']
					else:
						opparray['start_date'] = ""
						opparray['end_date'] = ""
					i+=1
			else:
				opparray['start_date'] = ""
				opparray['end_date'] = ""
			
			cursor = connection.cursor()
			sql_presence = "SELECT * FROM adminpanel_volunteersactivities WHERE opportunity_id="+str(opportunity_id)+" AND user_id="+str(user_id)
			cursor.execute(sql_presence)
			volunteer_details = dictfetchall(cursor)
			

			if volunteer_details:
				# return HttpResponse(1)
				row_id = volunteer_details[0]['id']
				msg ="Already present"
				sql_remove = "DELETE FROM adminpanel_volunteersactivities WHERE id="+str(row_id)
				cursor.execute(sql_remove)
				data = json.dumps({"Ack": 1, "msg": "Application Removed"})
			else:
				# return HttpResponse(2)
				start_date = str(opparray['start_date'])
				end_date = str(opparray['end_date'])
				sql_last_row = "SELECT * FROM adminpanel_volunteersactivities ORDER BY id DESC LIMIT 1"
				cursor.execute(sql_last_row)
				last_row_details = dictfetchall(cursor)
				if last_row_details:
					last_row_id = last_row_details[0]['id']
				else:
					last_row_id = 0
				new_row_id = last_row_id+1
				activities = Activities.objects.filter(id = 2).values('start_date','end_date','id','activity_name','description')
				if activities:
					ist = 0
				else:
					activitiesset =Activities(
						id = 2,
						activity_name = 'abc',
						author_name ='Gaurang Kher',
						address = '',
						start_date ='2019-08-21',
						end_date = '2019-08-30',
						description = '',
						opportunity_id = int(opportunity_id)
					)
					activitiesset.save()
				sql_save_volunteerof = "INSERT INTO adminpanel_volunteersactivities(id,start_date,end_date,activity_id,opportunity_id,user_id) VALUES("+str(new_row_id)+",'"+start_date+"','"+end_date+"',2,"+opportunity_id+","+user_id+")"
				cursor.execute(sql_save_volunteerof)
				data = json.dumps({"Ack": 1, "msg": "Successfully applied for this opportunity"})
			# return HttpResponse(org_id) 
			cordinatorRequest = CordinatorRequest(
				user_id_id = user_id,
				org_id_id = org_id,
				status = 'Pending', 
				role = int(2),
				oppurtunity_id = int(opportunity_id),
				is_url = False if url == 'false' else True, 
				is_request = "volunteer"
			)
			cordinatorRequest.save()
			data = json.dumps({"Ack": 1, "msg": "Successfully applied for this opportunity"})
		
			get_questions = OpportunityQuestions.objects.filter(opportunity_id=opportunity_id)
		
			if 'answer' in request.POST and request.POST.get('answer') != '': 
				i=0
				for answer in answers:
					# print(get_questions[i].id)
					opportunityanswers = Opportunityanswers(
						opportunity = Opportunities.objects.get(id=opportunity_id),
						user = User.objects.get(id=user_id),
						question = OpportunityQuestions.objects.get(id=get_questions[i].id),
						# question= get_questions[i].id,
						answer = answer['answer']
					)
					opportunityanswers.save()
					i= i+1
			return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0,"web_url":"", "msg": "Only post method applied"})
		return HttpResponse(data)
		
###############Done4################
class save_opp(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id', None)
		opportunity_id = request.POST.get('opportunity_id', None)
		cursor = connection.cursor()
		
		opportunity_present = OpportunitySaved.objects.filter(opportunity_id=opportunity_id,user_id=user_id)

		saved_data = {}

		if opportunity_present:
			
			OpportunitySaved.objects.filter(opportunity_id=opportunity_id,user_id=user_id).delete()
			data = json.dumps({"Ack": 1,"msg": "removed successfully"}, cls=DjangoJSONEncoder)
		else:
			save_opportunity = OpportunitySaved(
				opportunity_id = opportunity_id,
				user_id = user_id
			)
			save_opportunity.save()

			data = json.dumps({"Ack": 1,"msg": "saved successfully"}, cls=DjangoJSONEncoder)
		
		return HttpResponse(data)		
	def get(self,request):
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class save_opp_list(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id', None)
		
		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_opportunitysaved INNER JOIN adminpanel_opportunities ON adminpanel_opportunitysaved.opportunity_id = adminpanel_opportunities.id  WHERE adminpanel_opportunitysaved.user_id="+str(user_id)
		cursor.execute(sql)
		opportunity_present = dictfetchall(cursor)
		data_list = {}
		if opportunity_present:
			for i in range(len(opportunity_present)):
				data_list[i] = opportunity_present[i]

			data = json.dumps({"Ack": 1, "list_data":data_list}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else:
			data = json.dumps({"Ack": 0, "list_data":data_list}, cls=DjangoJSONEncoder)
			return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)


# def get_saved_opportunity(request):
class get_saved_opportunity(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		user_id = request.POST.get('user_id', None)
		event_date = request.POST.get('event_date', None)
		cursor = connection.cursor()
		
		user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
		cursor.execute(user_details_sql)
		user_details = dictfetchall(cursor)
		sql = "select adminpanel_opportunities.* from adminpanel_opportunities inner join adminpanel_opportunitysaved on adminpanel_opportunities.id=adminpanel_opportunitysaved.opportunity_id WHERE adminpanel_opportunitysaved.user_id="+str(user_id)+" and date(adminpanel_opportunities.start_date)='"+ event_date +"'"
		cursor.execute(sql) 
		opportunity = dictfetchall(cursor)
		v_data={}
		coordinate_data={}
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['id']
				opportunity[i]['opportunity_id']=opportunity[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
				cursor2.execute(sql2)
				volunteers = dictfetchall(cursor2)
				opportunity[i]['volunteers']=volunteers
				oppurtunityArr=opportunity[i]
				coordinate_data[i]=oppurtunityArr
				v_data=coordinate_data
			data = json.dumps({"Ack": 1, "event_data":v_data,"user_details":user_details}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0,"user_details":user_details}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0, "user_details":user_details}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

##########Done4#############
def remove_save_opportunity(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		id = request.POST.get('id', None)
		
		OpportunitySaved.objects.filter(id=str(id)).delete()
		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully removed"})
	else:
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
	return HttpResponse(data) 

class get_opportunities_savedstatus(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request,offset) :
		import json
		user_id = request.POST.get('user_id', None)
		
		limit = 12
		cursor = connection.cursor()
		user_details = []
		related_organization = []
		# return HttpResponse(user_id)
		if user_id != "":
			user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
			cursor.execute(user_details_sql)
			user_details = dictfetchall(cursor)
			cursor13 = connection.cursor()
			sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_cordinatorrequest.user_id_id='"+user_id+"' and adminpanel_cordinatorrequest.status='Approved' and adminpanel_organization.affiliated_org = 0 "
			cursor13.execute(sql3)
			related_organization = dictfetchall(cursor13)
		cursor = connection.cursor()
		changed_limit = int(offset)+limit

		sql = "select AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and AO.start_date > current_date order by AO.start_date asc LIMIT "+str(limit)+" OFFSET "+str(offset)
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['id']
				cursor2 = connection.cursor()
				if user_id != "":
					sql2 = "select count(*) as save_count from adminpanel_opportunitysaved WHERE opportunity_id="+str(o_id)+" and user_id="+str(user_id)
				else:
					sql2 = "select count(*) as save_count from adminpanel_opportunitysaved WHERE opportunity_id="+str(o_id)
				cursor2.execute(sql2)
				savedopportunities = dictfetchall(cursor2)
				opportunity[i]['is_saved']=savedopportunities

				question_cursor = connection.cursor()
				question_sql = "select * from adminpanel_opportunityquestions where opportunity_id ="+str(o_id)
				question_cursor.execute(question_sql)
				questions = dictfetchall(question_cursor)
				opportunity[i]['questions'] = questions



				shared_conn = connection.cursor()
				shared_sql = "select count(*) from adminpanel_opportunityshared where opportunity_id = " + str(o_id)
				shared_conn.execute(shared_sql)
				shared_res = dictfetchall(shared_conn)
				# print(shared_res)
				if shared_res[0]['count'] > 0:
					opportunity[i]['shared'] = 1
				else:
					opportunity[i]['shared'] = 0

				if opportunity[i]['user_id_id'] == str(user_id):
					opportunity[i]['is_apply'] = 0
				else:
					opportunity[i]['is_apply'] = 1
					apply_cursor = connection.cursor()
					if user_id != "":
						apply_sql = "select * from adminpanel_cordinatorrequest where oppurtunity_id ="+str(o_id)+" and user_id_id = "+str(user_id)
					else:
						apply_sql = "select * from adminpanel_cordinatorrequest where oppurtunity_id ="+str(o_id)
					apply_cursor.execute(apply_sql)
					applystatus = dictfetchall(apply_cursor)
					if applystatus:
						opportunity[i]['is_apply'] = 2
						opportunity[i]['application_status'] = applystatus[0]['status']
				opportunity[i]['secret_message'] =''
		
		data = json.dumps({"Ack": 1, "opportunities": opportunity, "image_url": settings.BASE_URL + "/media/","user_details":user_details,"related_organization":related_organization})
		return HttpResponse(data)

# def get_myapplied_opportunity(request):
class get_myapplied_opportunity(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id', None)
		cursor = connection.cursor()

		user_details_sql = "SELECT role_id FROM adminpanel_userprofile WHERE user_id_id="+user_id
		cursor.execute(user_details_sql)
		user_details = dictfetchall(cursor)
		sql = "select * from adminpanel_cordinatorrequest WHERE user_id_id="+str(user_id)
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['oppurtunity_id']
				cursor2 = connection.cursor()
				sql2 = "select * from adminpanel_opportunities WHERE id="+str(o_id)
				cursor2.execute(sql2)
				opportunities = dictfetchall(cursor2)
				opportunity[i]['opportunities']=opportunities
			data = json.dumps({"Ack": 1, "event_data":opportunity, "user_details":user_details}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0,"user_details":user_details}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0, "user_details": user_details}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

###############Done3################
class share_opportunity_mail(ObtainAuthToken):
	
	# permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		# return HttpResponse(request)
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		msg_html = ""
		opportunity_id = request.POST.get('opportunity_id', None)
		# return HttpResponse(opportunity_id)
		user_id = request.POST.get('user_id', None)
		if user_id == '0':
			userid = 67
		else:
			userid = user_id
		# return HttpResponse(userid)
		org_id_id = request.POST.get('organization_id', None)
		# return HttpResponse(13)
		if org_id_id != '':
			org_id_id = org_id_id
		else:
			org_id_id = 0

		mail = request.POST.get('mail_id', None)
		# link = request.POST.get('link', None)
		if opportunity_id:
			link = settings.PATH_URL+"/opportunity-details/"+opportunity_id

		save_opportunity = OpportunityShared(
			opportunity_id = opportunity_id,
			user_id = userid,
			facebook = False,
			twitter = False,
			linkedin = False,
			google = True,
			org_id_id = int(org_id_id)
		)
		save_opportunity.save()
		cursor = connection.cursor()
		sql = "select AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and AO.id ="+str(opportunity_id)
		cursor.execute(sql)
		opportunity_details = dictfetchall(cursor)
		
		print(opportunity_details)
		if userid != 0:
			# return HttpResponse(0)
			userprofiles = User.objects.filter(id=userid)
			msg_html += '<p>Hi,</p> <p> '+userprofiles[0].first_name+' '+userprofiles[0].last_name+' shared this opportunity with you.</p><p>Go ahead and apply to this opportunity!</p> <a href="'+link+'"></p><div style="width: 300px; border: 1px solid #d8d8d8;"><div style="width: 300px; height: 180px;">'
			if opportunity_details[0]['image']: 
				msg_html += '<img src="'+settings.UPLOAD_IMAGES_URL_ROOT+opportunity_details[0]['image']+'" style="width: 100%; height: 100%;">'
				# msg_html += '<img src="'+settings.IMAGE_URL+opportunity_details[0]['image']+'" style="width: 100%; height: 100%;">'
				print(msg_html)
				print('kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
			else:
				msg_html +='<img src="'+settings.APP_URL+'app/assets/img/default.png" style="width: 100%; height: 100%;">' 
			
			msg_html +='</div><div style="padding: 0px 20px 20px;"><div style="display: flex; align-items: center;"><div style="margin-right: 10px;"><p style="line-height: 35px;background: #fff;border: 1px solid #d8d8d8; width: 35px; height: 35px; text-align: center; border-radius: 50%;">TA</p></div><div><p style="margin: 0px; font-family: "Open Sans", sans-serif; ">'+opportunity_details[0]['organization_name']+'</p><small>'+opportunity_details[0]['first_name']+' '+opportunity_details[0]['last_name']+'</small></div></div><div><h3 style="font-family: Open Sans, sans-serif; margin: 0px; font-size: 13px;">'+opportunity_details[0]['opportunity_name']+'</h3><p style="font-family: Open Sans, sans-serif; margin: 0px; margin-top: 5px; font-size: 13px; color: #676767;">'+opportunity_details[0]['description']+'</p></div></div></div></a><p>Thanks,</p><p>Inteer Team</p>' 
		else:
			# return HttpResponse(1)
			msg_html = '<p>Hello,</p><p><a href="'+link+'">Shared Link</a></p><p>Thanks,</p><p>Inteer Team</p>'
		if send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', [mail], html_message=msg_html):
			data = json.dumps({"Ack": 1, "msg": "Opportunity successfully shared"})
			# return HttpResponse(1)
		else:
			data = json.dumps({"Ack": 1, "msg": "Mail not sent"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)
	def put(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)
	def patch(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)
	def delete(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)

###############Done2################
# def share_opportunity(request):

class share_opportunity(ObtainAuthToken):
	# permission_classes = (IsAuthenticated,)
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		opportunity_id = request.POST.get('opportunity_id', None)
		user_id = request.POST.get('user_id', None)
		facebook = request.POST.get('facebook', None)
		twitter = request.POST.get('twitter', None)
		linkedin = request.POST.get('linkedin', None)
		google = request.POST.get('google', None)
		org_id_id = request.POST.get('organization_id', None)
		 
		# print(request.POST)

		if org_id_id != '':
			org_id_id = org_id_id
		else:
			opportunity_details = Opportunities.objects.filter(id=str(opportunity_id)).values('org_id')
			org_id_id = opportunity_details[0]['org_id']

		# return HttpResponse(org_id_id)
		save_opportunity = OpportunityShared(
			opportunity_id = opportunity_id,
			user_id = user_id,
			facebook = True if facebook == 1 else False,
			twitter = True if twitter == 1 else False,
			linkedin = True if linkedin == 1 else False,
			google = True if google == 1 else False,
			org_id_id = int(org_id_id)
		)
		save_opportunity.save()
		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully shared"})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
		return HttpResponse(data)

###############Done4################
def share_opportunity_update(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		opportunity_id = request.POST.get('opportunity_id', None)
		user_id = request.POST.get('user_id', None)
		facebook = request.POST.get('facebook', None)
		twitter = request.POST.get('twitter', None)
		linkedin = request.POST.get('linkedin', None)
		google = request.POST.get('google', None)
		id = request.POST.get('id', None)

		OpportunityShared.objects.filter(id=str(id)).update(facebook=facebook,twitter=twitter,linkedin=linkedin,google=google)
		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully shared"})
	else:
		data = json.dumps({"Ack": 0, "msg": "Please try again"})
	return HttpResponse(data)

# def get_shared_opportunity(request):

class get_shared_opportunity(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		user_id = request.POST.get('user_id', None)
		event_date = request.POST.get('event_date', None)
		cursor = connection.cursor()
		sql = "select adminpanel_opportunities.* from adminpanel_opportunities inner join adminpanel_opportunityshared on adminpanel_opportunities.id=adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.user_id="+str(user_id)+"  and date(adminpanel_opportunities.start_date)='"+ event_date +"'"
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		v_data={}
		coordinate_data={}
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['id']
				opportunity[i]['opportunity_id']=opportunity[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
				cursor2.execute(sql2)
				volunteers = dictfetchall(cursor2)
				opportunity[i]['volunteers']=volunteers
				oppurtunityArr=opportunity[i]
				coordinate_data[i]=oppurtunityArr
				v_data=coordinate_data
			data = json.dumps({"Ack": 1, "event_data":v_data}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class my_inteer(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		data = {}
		user_id = request.POST.get('user_id', None)
		cursor = connection.cursor()
		sql = "select adminpanel_opportunities.* from adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		c_data={}
		coordinate_data={}
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['id']
				opportunity[i]['opportunity_id']=opportunity[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
				cursor2.execute(sql2)
				cordinators = dictfetchall(cursor2)
				opportunity[i]['cordinators']=cordinators

				oppurtunityArr=opportunity[i]
				coordinate_data[i]=oppurtunityArr
				c_data=coordinate_data
		else :
			c_data = [];
		sql1 = "select auth_user.*,adminpanel_volunteersactivities.* from adminpanel_volunteersactivities inner join auth_user on adminpanel_volunteersactivities.user_id=auth_user.id where adminpanel_volunteersactivities.user_id="+str(user_id)
		cursor.execute(sql1)
		volunteer_list = dictfetchall(cursor)
		v_data={}
		volunteer_data={}
		if volunteer_list:
			for i in range(len(volunteer_list)):
				
				volunteer_data[i]=volunteer_list[i]
				v_data=volunteer_data
		else :
			v_data = []
		data = json.dumps({"Ack": 1,"coordinator_data":c_data,"volunteer_data": v_data}, cls=DjangoJSONEncoder)
		return HttpResponse(data)	
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)	

class applicant_sorting(APIView):

	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id', None)
		c_sort = request.POST.get('sort_condition', None)
		data = {}
		cursor = connection.cursor()
		if c_sort:
			if c_sort == "approved":
				sql = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitiesapplied ON adminpanel_opportunities.id = adminpanel_opportunitiesapplied.opportunity_id WHERE adminpanel_opportunitiesapplied.is_url != false AND adminpanel_opportunitiesapplied.user_id="+str(user_id)
			else:
				sql = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitiesapplied ON adminpanel_opportunities.id = adminpanel_opportunitiesapplied.opportunity_id WHERE adminpanel_opportunitiesapplied.is_url = false AND adminpanel_opportunitiesapplied.user_id="+str(user_id)
		else:
			sql = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunitiesapplied ON adminpanel_opportunities.id = adminpanel_opportunitiesapplied.opportunity_id WHERE adminpanel_opportunitiesapplied.user_id="+str(user_id)

		cursor.execute(sql)
		list_application ={}
		all_application = dictfetchall(cursor)
		if all_application:
			for i in range(len(all_application)):
				list_application[i]= all_application[i]
			data = json.dumps({"Ack": 1, "list_application":list_application}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

# def get_coordinator(request):
class get_coordinator(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		user_id = request.POST.get('user_id', None)
		
		cursor = connection.cursor()
		sql = "select adminpanel_opportunities.* from adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)
		v_data={}
		coordinate_data={}
		if opportunity:
			for i in range(len(opportunity)):
				o_id=opportunity[i]['id']
				opportunity[i]['opportunity_id']=opportunity[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2"
				cursor2.execute(sql2)
				cordinators = dictfetchall(cursor2)
				opportunity[i]['cordinators']=cordinators

				oppurtunityArr=opportunity[i]
				coordinate_data[i]=oppurtunityArr
				v_data=coordinate_data
			data = json.dumps({"Ack": 1, "event_data":v_data}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

# def get_volunteersof(request):
class get_volunteersof(APIView):
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		import datetime
		user_id = request.POST.get('user_id', None)
		
		cursor = connection.cursor()
		
		sql = "select auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id where adminpanel_cordinatorrequest.user_id_id="+str(user_id)+" and adminpanel_cordinatorrequest.role=2"
		
		cursor.execute(sql)
		volunteer_list = dictfetchall(cursor)
		v_data={}
		volunteer_data={}
		if volunteer_list:
			for i in range(len(volunteer_list)):
				
				volunteer_data[i]=volunteer_list[i]
				
			data = json.dumps({"Ack": 1, "volunteer_data":volunteer_data}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

# def get_myshared_opportunity(request):
class get_myshared_opportunity(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id',None)
		data = {}
		cursor = connection.cursor();
		sql = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunityshared.user_id="+str(user_id);
		cursor.execute(sql)
		opportunity = dictfetchall(cursor)

		v_data={}
		coordinate_data={}
		if opportunity:
			for i in range(len(opportunity)):
				data[i] = opportunity[i]
				opportunities=data
			data = json.dumps({"Ack": 1, "sharedopportunities":opportunities}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class get_myposted_opportunity(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def get(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id',None)
		event_date = request.POST.get('event_date', None)
		data = {}
	
		cursor = connection.cursor();
		sql1 = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities WHERE adminpanel_opportunities.user_id_id="+str(user_id)+" and date(adminpanel_opportunities.start_date)='"+event_date +"'"
		cursor.execute(sql1)
		opportunity = dictfetchall(cursor)
		
		if opportunity:
			for i in range(len(opportunity)):
				data[i] = opportunity[i]
				opportunities=data
			data = json.dumps({"Ack": 1, "mypostedopportunities":opportunities}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class shared_opportunity_organization(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder

		user_id = request.POST.get('user_id',None)
		opportunity_id = request.POST.get('opportunity_id',None)
		data = {}
		cursor = connection.cursor();
		sql = "SELECT * FROM adminpanel_cordinatorrequest WHERE user_id_id ="+str(user_id)
		cursor.execute(sql)
		user_deatils = dictfetchall(cursor)
		org_id = user_deatils[0]['org_id_id']

		fetch_users_query = "SELECT * FROM adminpanel_cordinatorrequest WHERE org_id_id="+str(org_id)
		cursor.execute(fetch_users_query)
		all_user_deatils = dictfetchall(cursor)
		insert_user = {}
		if all_user_deatils:
			for m in range(len(all_user_deatils)):
				insert_user[m] = all_user_deatils[m]['id']
				if (all_user_deatils[m]['id'] != 24 and all_user_deatils[m]['id'] !=54 and all_user_deatils[m]['id'] !=57 and all_user_deatils[m]['id'] !=58 and all_user_deatils[m]['id'] !=64 and all_user_deatils[m]['id'] !=62 and all_user_deatils[m]['id'] !=63 and all_user_deatils[m]['id'] !=71 and all_user_deatils[m]['id'] !=72 and all_user_deatils[m]['id'] !=73 and all_user_deatils[m]['id'] !=74 and all_user_deatils[m]['id'] !=75 and all_user_deatils[m]['id'] !=76 and all_user_deatils[m]['id'] !=77):
					sql_presence = "SELECT * FROM adminpanel_opportunityshared WHERE opportunity_id="+str(opportunity_id)+" AND user_id="+str(insert_user[m])
					cursor.execute(sql_presence)
					present_details = dictfetchall(cursor)
					
					if present_details:
						msg = "already present"
					else:
						insert_sql = "INSERT INTO adminpanel_opportunityshared(facebook, twitter, linkedin, google, opportunity_id, user_id) VALUES(true, true, true, true, "+str(opportunity_id)+", "+str(insert_user[m])+")"
						
						cursor.execute(insert_sql)


			data = json.dumps({"Ack": 1,"msg":"shared"}, cls=DjangoJSONEncoder)
			
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			
		return HttpResponse(data)
	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

class get_shared_opportunity_organization(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request) :
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('user_id',None)
		org_id = request.POST.get('organization_id',None)
		data = {}
	
		cursor = connection.cursor();
		sql1 = "SELECT adminpanel_opportunities.* FROM adminpanel_opportunities INNER JOIN adminpanel_opportunityshared ON adminpanel_opportunities.id = adminpanel_opportunityshared.opportunity_id WHERE adminpanel_opportunities.org_id_id = "+str(org_id)+" AND adminpanel_opportunityshared.user_id="+str(user_id)
		cursor.execute(sql1)
		opportunity = dictfetchall(cursor)
		
		if opportunity:
			for i in range(len(opportunity)):
				data[i] = opportunity[i]
				opportunities=data
			data = json.dumps({"Ack": 1, "myOrganizationSharedOpportunities":opportunities}, cls=DjangoJSONEncoder)
			return HttpResponse(data)
		else :
			data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
			return HttpResponse(data)

	def get(self,request) :
		data = json.dumps({"Ack": 0}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

###############Done2################
def update_opportunity(request):
	import json

	if request.method == 'POST':
		id = request.POST.get('id', None)
		opportunity_name = request.POST.get('opportunity_name', None)
		description = request.POST.get('description', None)
		no_of_volunteers = request.POST.get('no_of_volunteers', None)
		address = request.POST.get('address', None)

		parent_id = request.POST.get('parent_id', None)
		no_ofyear = request.POST.get('no_ofyear', None)
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		questions = request.POST.get('questions')
		questions = json.loads(questions)
		
		if len(request.FILES) != 0:
			image_file = request.FILES['file']
		else :
			image_file = ''

		opportunitiy = Opportunities.objects.get(id = id)
		opportunitiy.opportunity_name = opportunity_name
		opportunitiy.description = description
		opportunitiy.no_of_volunteers = no_of_volunteers
		opportunitiy.address = address
		opportunitiy.parent_id = parent_id
		opportunitiy.no_ofyear = no_ofyear
		opportunitiy.start_date = start_date
		opportunitiy.end_date = end_date
		opportunitiy.save()

		data = json.dumps({"Ack": 1, "msg": "Opportunity successfully saved"})
	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})

	return HttpResponse(data)

# def addvolunteerbyowner(request):
class addvolunteerbyowner(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		import base64
		import time
		activate_no = int(time.time())
		o_id=request.POST.get('opportunity_id', None)
		org_id = request.POST.get('org_id', None)
		password= '123456'
		
		manual_volunteer = request.POST.get('manual_volunteer', None)
		manual_volunteer = json.loads(manual_volunteer)
		
		if request.POST.get('manual_volunteer') != '':
			i=0
			approvelist = {}
			for volunteer in manual_volunteer:
				
				volunteer_fetch = User.objects.filter(email=str(volunteer['email'])).values('id')
				if volunteer_fetch:
					user_id = volunteer_fetch[0]['id']
				else:
					user = User.objects.create_user(
						password=password,
						is_superuser=False,
						username=volunteer['email'],
						first_name=volunteer['first_name'],
						last_name=volunteer['last_name'],
						email=volunteer['email'],
						is_staff=False,
						is_active=True
					)
					user.save()
					user_id=user.id
					user_info = UserProfile(
						address='',
						phone_number='',
						latitude=22.57,
						longitude=88.36,
						is_verified=0,
						interest_id_id=1,
						role_id=2,
						user_id_id=user_id,
						activate_token=activate_no
					)
					user_info.save()
				check_applied = CordinatorRequest.objects.filter(user_id_id=user_id,org_id_id=org_id,is_request='volunteer',oppurtunity_id=o_id,status="Approved").count()
				if check_applied >0 :
					apply_coordinator = CordinatorRequest.objects.filter(user_id_id=user_id,org_id_id=org_id,is_request='volunteer',oppurtunity_id=o_id,status="Approved").last()
					data = json.dumps({"Ack": 1, "msg": "Already is in this list"})
				else:
					apply_coordinator = CordinatorRequest(
						user_id_id=user_id,
						org_id_id=org_id,
						status="Approved",
						address='N/A',
						employee_number=0,
						oppurtunity_id=o_id,
						role=2,
						is_request='volunteer'
						)
					apply_coordinator.save()
					cursor4 = connection.cursor()
					sql4 = "select DISTINCT(RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name))))) AS Full_Name, adminpanel_userprofile.profile_image as pending_user_image, auth_user.first_name,auth_user.last_name,auth_user.email,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.id="+str(apply_coordinator.id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
					cursor4.execute(sql4)
					approvelist_details = dictfetchall(cursor4)
					approvelist[i] = approvelist_details
					i+=1


			data = json.dumps({"Ack": 1, "msg": "Volunteer Added for opportunitiy","approvelist":approvelist})
			return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

###############Done2################
class upload_image(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	
	def post(self,request):
		import json
		data = {}
		user_id = request.POST.get('user_id', None)
		cursor = connection.cursor()
		logo_path = settings.PROFILE_IMAGE_ROOT

		if request.FILES:

			folder=settings.PROFILE_IMAGE_URL
			logo_image = request.FILES['user_image']
			

			fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 

			splitted_value = logo_image.name.split(".")
			img_type = splitted_value[len(splitted_value)-1]

			image_name_get = ''
			for i in range(len(splitted_value)-1):
				image_name_get+=str(splitted_value[i]).replace(" ", "_")+'_'
			
			img_name =image_name_get+ str( time.time()).split('.')[0]
			final_image = img_name+'.'+img_type

			filename = fs.save(final_image, logo_image)

			imgthumb = Image.open(settings.PROFILE_IMAGE_URL+final_image)
			imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
			imgthumb.save(settings.PROFILE_IMAGE_URL+final_image)

			# filename = fs.save(logo_image.name, logo_image)
			file_url_image = fs.url(filename)
			file_url = final_image
		else:
			file_url = ""	

		image = logo_path+file_url
		if file_url != "" :
			UserProfile.objects.filter(user_id_id=str(user_id)).update(profile_image = str(file_url))
		data = json.dumps({"Ack": 1, "image": image})
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

def test_upload_image(request) :
	import json
	data = {}
	user_id = request.POST.get('user_id', None)
	# userImage =  request.FILES['user_image']
	cursor = connection.cursor()
	logo_path = settings.PROFILE_IMAGE_ROOT
	# print(userImage)
	print(request.FILES)
	print('dfgfffr')
	if request.FILES:

		folder=settings.PROFILE_IMAGE_URL
		logo_image = request.FILES['user_image']
		

		fs = FileSystemStorage(location=folder) #defaults to MEDIA_ROOT 

		splitted_value = logo_image.name.split(".")
		img_type = splitted_value[len(splitted_value)-1]

		image_name_get = ''
		for i in range(len(splitted_value)-1):
			image_name_get+=str(splitted_value[i]).replace(" ", "_")+'_'
		
		img_name =image_name_get+ str( time.time()).split('.')[0]
		final_image = img_name+'.'+img_type

		filename = fs.save(final_image, logo_image)

		imgthumb = Image.open(settings.PROFILE_IMAGE_URL+final_image)
		imgthumb = imgthumb.resize((820,464), PIL.Image.ANTIALIAS)
		imgthumb.save(settings.PROFILE_IMAGE_URL+final_image)

		# filename = fs.save(logo_image.name, logo_image)
		file_url_image = fs.url(filename)
		file_url = final_image
	else:
		file_url = ""	

	image = logo_path+file_url
	if file_url != "" :
		UserProfile.objects.filter(user_id_id=str(user_id)).update(profile_image = str(file_url))
	data = json.dumps({"Ack": 1, "image": image})
	return HttpResponse(data)

def getengage_by_userdate2(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		user_id = request.POST.get('volunteer', None) 
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		user_set_id = request.POST.get('user_set_id',None)
		
		sql2 = "SELECT cr.org_id_id FROM adminpanel_cordinatorrequest cr WHERE cr.status='Approved' AND cr.user_id_id= "+str(user_set_id)+" GROUP BY cr.org_id_id"
		
		cursor = connection.cursor()
		cursor.execute(sql2)
		orgs = dictfetchall(cursor)
		opportunities = []
		newone12 = {}
		k=0
		if orgs:
			for i in range(len(orgs)):
				
				org_id = orgs[i]['org_id_id']
				sql = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, au.first_name, au.last_name FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id WHERE cr.user_id_id = "+str(user_id)+" AND o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"' AND cr.status='Approved' AND cr.is_request='volunteer' AND cr.org_id_id="+str(org_id)
		
				cursor = connection.cursor()
				cursor.execute(sql)
				opportunities = dictfetchall(cursor)
				for op in opportunities:
					newone12[k] = op
					k+=1



		data = json.dumps({"Ack": 1, "opportunities": newone12,"total":len(opportunities),"user_name": "yes"}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
	return HttpResponse(data)


class getengage_by_userdate(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		user_id = request.POST.get('volunteer', None) 
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		user_set_id = request.POST.get('user_set_id',None)


	
		sql = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, au.first_name, au.last_name, cr.oppurtunity_id as optu_id FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id WHERE cr.user_id_id = "+str(user_id)+" AND o.start_date >= '"+start_date+"' AND o.start_date <= '"+end_date+"' AND cr.status='Approved' AND cr.is_request='volunteer' AND o.user_id_id= "+str(user_set_id)
		
		cursor = connection.cursor()
		cursor.execute(sql)
		opportunities = dictfetchall(cursor) 

		if opportunities:
			for i in range(len(opportunities)):
				oppurtunityArr ={}
				o_id=opportunities[i]['id']
				opportunities[i]['opportunity_id']=opportunities[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select adminpanel_cordinatorrequest.user_id_id from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 GROUP BY adminpanel_cordinatorrequest.user_id_id"
				cursor2.execute(sql2)
				volunteer_ids = dictfetchall(cursor2) 

				vol_id = opportunities[i]['user_id_id']
				sql_vol = "select adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" AND adminpanel_cordinatorrequest.user_id_id ="+str(vol_id)
				cursor2.execute(sql_vol) 
				volunteers = dictfetchall(cursor2)
				if volunteers: 
					is_present=1
				else:
					is_present=0
				opportunities[i]['employee_number']	= is_present

		data = json.dumps({"Ack": 1, "opportunities": opportunities,"total":len(opportunities),"user_name": "yes"}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)


class getengage_by_date(APIView):
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		user_id_set = request.POST.get('user_id', None)
		

		getdetails = UserProfile.objects.filter(user_id_id=user_id_set).values('role_id')
		# return HttpResponse(getdetails[0]['role_id'])

		if getdetails[0]['role_id'] == 3:
			sql = "SELECT *,o.end_date, o.start_date, EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, cr.oppurtunity_id as optu_id, au.first_name, au.last_name, aup.profile_image as profilepic  FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id LEFT JOIN adminpanel_userprofile aup on cr.user_id_id = aup.user_id_id WHERE o.start_date >= '"+start_date+"' AND o.start_date <= '"+end_date+"' AND cr.status='Approved' AND cr.is_request='volunteer' AND o.user_id_id= "+str(user_id_set)

		else:


			sql = "SELECT *, EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, cr.oppurtunity_id as optu_id, au.first_name, au.last_name, aup.profile_image as profilepic FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id LEFT JOIN adminpanel_userprofile aup on cr.user_id_id = aup.user_id_id WHERE o.start_date >= '"+start_date+"' AND o.start_date <= '"+end_date+"' AND cr.status='Approved' AND cr.is_request='volunteer' AND cr.user_id_id= "+str(user_id_set)

		cursor = connection.cursor()
		cursor.execute(sql)
		opportunities = dictfetchall(cursor)

		if opportunities:
			for i in range(len(opportunities)):
				oppurtunityArr ={}
				o_id=opportunities[i]['optu_id']
				opportunities[i]['opportunity_id']=opportunities[i]['optu_id']
				# opportunities[i]['time_calculation'] = (opportunities[i]['end_date']-opportunities[i]['start_date'])/3600
				# return HttpResponse(opportunities[i]['end_date'])
				cursor2 = connection.cursor()
				sql2 = "select adminpanel_cordinatorrequest.user_id_id from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 GROUP BY adminpanel_cordinatorrequest.user_id_id"
				cursor2.execute(sql2)
				volunteer_ids = dictfetchall(cursor2) 

				vol_id = opportunities[i]['userid'] 
				sql_vol = "select adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" AND adminpanel_cordinatorrequest.user_id_id ="+str(vol_id)
				cursor2.execute(sql_vol) 
				volunteers = dictfetchall(cursor2)
				if volunteers:
					is_present=1
				else:
					is_present=0
				opportunities[i]['employee_number']	= is_present


		data = json.dumps({"Ack": 1, "opportunities":opportunities, "total":len(opportunities),"user_name": ""}, cls=DjangoJSONEncoder)
		return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)






def getengage_by_date2(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		user_id_set = request.POST.get('user_id', None)	
		sql2 = "SELECT cr.org_id_id FROM adminpanel_cordinatorrequest cr WHERE cr.status='Approved' AND cr.user_id_id= "+str(user_id_set)+" GROUP BY cr.org_id_id"
		
		cursor = connection.cursor()
		cursor.execute(sql2)
		orgs = dictfetchall(cursor)
		opportunities = []
		newone = []
		newone12 = {}
		k=0
		if orgs: 
			for i in range(len(orgs)):
				
				org_id = orgs[i]['org_id_id']
				sql = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, au.first_name, au.last_name, aup.profile_image as profilepic FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id LEFT JOIN adminpanel_userprofile aup on cr.user_id_id = aup.user_id_id WHERE o.start_date >= '"+start_date+"' AND o.start_date <= '"+end_date+"' AND cr.status='Approved' AND cr.is_request='volunteer' AND cr.org_id_id="+str(org_id)
		
				cursor = connection.cursor()
				cursor.execute(sql)
				opportunities = dictfetchall(cursor)
				for op in opportunities:
					newone12[k] = op
					k+=1
				# newone[i] = newone12


		v_data={}
		coordinate_data={}
		k=0
		volunteer_ids = []

		another_array = {}
		idArray = []
		k=0
		if opportunities:
			for i in range(len(opportunities)):
				another_array11 = {}
				another_array11['name']= opportunities[i]['first_name']+' '+opportunities[i]['last_name']
				another_array11['opportunity'] = opportunities[i]['id']
				another_array11['opportunity_name'] = opportunities[i]['opportunity_name']

				# oppurtunityArr=another_array11[i]
				if opportunities[i]['id'] not in idArray:
						idArray.append(opportunities[i]['id'])
						another_array[k]=another_array11
				
				k+=1
		if opportunities:
			for i in range(len(opportunities)):
				oppurtunityArr ={}
				o_id=opportunities[i]['id']
				opportunities[i]['opportunity_id']=opportunities[i]['id']
				cursor2 = connection.cursor()
				sql2 = "select adminpanel_cordinatorrequest.user_id_id from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 GROUP BY adminpanel_cordinatorrequest.user_id_id"
				cursor2.execute(sql2)
				volunteer_ids = dictfetchall(cursor2) 

				vol_id = opportunities[i]['user_id_id']
				sql_vol = "select adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" AND adminpanel_cordinatorrequest.user_id_id ="+str(vol_id)
				cursor2.execute(sql_vol) 
				volunteers = dictfetchall(cursor2)
				if volunteers:
					is_present=1
				else:
					is_present=0
				opportunities[i]['employee_number']	= is_present 

		data = json.dumps({"Ack": 1, "opportunities":newone12, "volunteer_ids": volunteer_ids,"another_array":another_array,"newone":newone12,"orgs":orgs, "total":len(opportunities)}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
	return HttpResponse(data)



class getengage_by_date_new(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		
		
		sql = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, o.user_id_id as given_user_id FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id WHERE o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"'"
		# print(sql)
		cursor = connection.cursor()
		cursor.execute(sql)
		opportunities = dictfetchall(cursor)
		
		data_array = []
		opt_array = []
		user_array = []
		for index, value in enumerate(opportunities):
			
			user_array.append(value['given_user_id'])

			data_array.append(value['oppurtunity_id'])
		data_array_set=[]
		for j in data_array:
			if j not in data_array_set:
				data_array_set.append(j)

		
		for all_opp in data_array_set:
			
			query = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id WHERE cr.oppurtunity_id ='"+ str(all_opp) +"' AND o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"'"
			print (1)
			cursor.execute(query)
			new_array = dictfetchall(cursor)
			opt_array.append(new_array)

		opportunity_list1 = {}
		new_dict_list3 = []
		for value_op in opt_array:
			new_dict_list2 = []
			for value_in in value_op:
				new_dict_list = {}
				user_id_id = value_in["user_id_id"]
				flag = 0
				for k,v in new_dict_list.items():
					if str(user_id_id) == str(k):
						flag = 1
						break
				if flag == 0:
					new_dict_list[user_id_id] = list(value_in)
				else:
					new_dict_list[user_id_id].append(value_in)
				new_dict_list2.append(new_dict_list)
			new_dict_list3.append(new_dict_list2)
		

		after_unique = list(set(user_array))
		
		user_array_add = []
		for all_user in user_array:
			query_user = "SELECT * FROM adminpanel_opportunities WHERE user_id_id ='"+str(all_user)+"' AND start_date >= '"+start_date+"' AND end_date <= '"+end_date+"'"
			
			cursor.execute(query_user)
			new_user_array = dictfetchall(cursor)
			user_array_add.append(new_user_array)
			
		new_user_array_test = {}
		i = 0
		viewlist = {}
		for after_uniq in after_unique:
			
			query_user = "SELECT *,EXTRACT(EPOCH FROM (end_date - start_date)/3600) as time_cal FROM adminpanel_opportunities WHERE user_id_id ='"+str(after_uniq)+"' AND start_date >= '"+start_date+"' AND end_date <= '"+end_date+"'"
			
			cursor.execute(query_user)
			new_user_array_test[after_uniq]=(dictfetchall(cursor))
			
		data = json.dumps({"Ack": 1, "opportunities": opportunities, "opt_array": new_dict_list3,"user_array": new_user_array_test}, cls=DjangoJSONEncoder)
		return HttpResponse(data)

	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

class get_user_volunteer(APIView):
	
	def post(self,request):
		import json
		import datetime
		cursor = connection.cursor()
		sql = "SELECT adu.id,au.first_name,au.last_name FROM adminpanel_userprofile adu LEFT JOIN auth_user au ON adu.user_id_id = au.id WHERE adu.role_id='2'"
		cursor.execute(sql)
		vuser = dictfetchall(cursor)
		data = json.dumps({"Ack": 1, "volunteeruser": vuser})
		return HttpResponse(data)

class get_volunteers_suggestion(APIView):

	def post(self,request):
		import json
		cursor = connection.cursor()
		user_id = request.POST.get('user_id', None)
		# return HttpResponse(user_id)
		if user_id!=0:
			# return HttpResponse(0);
			sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org=0 and adminpanel_cordinatorrequest.user_id_id='"+str(user_id)+"' and adminpanel_cordinatorrequest.is_request='coordinator'"
		else:
			# return HttpResponse(1);
			opportunity_id = request.POST.get('organization_id', None)

			opportunity_details = Opportunities.objects.filter(id=str(id)).values('org_id_id')
			org_id = opportunity_details[0]['org_id_id']
			sql3 = "select adminpanel_cordinatorrequest.*,adminpanel_cordinatorrequest.id as cid, adminpanel_cordinatorrequest.org_id_id as org_id,adminpanel_cordinatorrequest.status as astatus, adminpanel_cordinatorrequest.user_id_id as user_id, adminpanel_organization.*, adminpanel_organization.id as org_id, adminpanel_organization.organization_name as org_name, adminpanel_organization.address as address from adminpanel_cordinatorrequest inner join adminpanel_organization on adminpanel_cordinatorrequest.org_id_id=adminpanel_organization.id where adminpanel_organization.affiliated_org=0 and adminpanel_cordinatorrequest.org_id_id="+str(org_id)+" and adminpanel_cordinatorrequest.is_request='coordinator'"

		cursor.execute(sql3)
		coordinatorexist = dictfetchall(cursor)

		idArray1 = []
		for key, value in enumerate(coordinatorexist):
			if value['org_id'] not in idArray1: 
				idArray1.append(value['org_id'])

		sql = ""
		volunteers ={}
		emailArray =[]
		i=0
		if idArray1:

			for idGet in idArray1:
				# print(idGet)
				sql = "SELECT DISTINCT(RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name))))) AS Full_Name,auth_user.first_name,auth_user.last_name ,auth_user.email,adminpanel_cordinatorrequest.user_id_id, adminpanel_userprofile.profile_image FROM auth_user INNER JOIN adminpanel_cordinatorrequest ON auth_user.id = adminpanel_cordinatorrequest.user_id_id AND adminpanel_cordinatorrequest.role=2 LEFT JOIN adminpanel_userprofile ON auth_user.id = adminpanel_userprofile.user_id_id WHERE adminpanel_cordinatorrequest.org_id_id="+str(idGet)
				cursor.execute(sql)
				volunteers_data = dictfetchall(cursor)
				if volunteers_data:
					for v_data in volunteers_data:
						# print(v_data['email']) 
						if v_data['email'] not in emailArray: 
							emailArray.append(v_data['email'])
							volunteers[i] = v_data
							i=i+1

		
		else:
			sql = "SELECT DISTINCT(RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name))))) AS Full_Name,auth_user.first_name,auth_user.last_name ,auth_user.email,adminpanel_cordinatorrequest.user_id_id, adminpanel_userprofile.profile_image FROM auth_user INNER JOIN adminpanel_cordinatorrequest ON auth_user.id = adminpanel_cordinatorrequest.user_id_id AND adminpanel_cordinatorrequest.role=2 LEFT JOIN adminpanel_userprofile ON auth_user.id = adminpanel_userprofile.user_id_id"
			cursor.execute(sql)
			volunteers = dictfetchall(cursor)

		data = json.dumps({"Ack": 1, "volunteers": volunteers,"sql":idArray1}) 
		return HttpResponse(data)
	def get(self,request):
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
		return HttpResponse(data)

def  upload_organization(request):
	import csv
	import json
	import io
	import openpyxl
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		data = []
		data_result = []
		all_volunteer ={}
		user = User.objects.get(id=42)
		getfile = request.FILES['csv_file']
		splitted_value = getfile.content_type.split('/')[1].split(".")
		sheet_type = splitted_value[len(splitted_value)-1]
		 
		wb = openpyxl.load_workbook(getfile)
		for sheetName in wb.sheetnames:
			if sheetName == 'Organisation':
				ws = wb.get_sheet_by_name('Organisation')
				for index, row in enumerate(ws.iter_rows(),start=2):
					c=1 
					cell_dict={}
					for cell in row:
						cell_dict[ws.cell(row=1, column=c).value] = ws.cell(row=index, column=c).value
						c+=1 
					# print(cell_dict)  
					if 'Organization Name(Youth development)' in cell_dict and 'Location' in cell_dict :
						if cell_dict['Organization Name(Youth development)'] != None and cell_dict['Location'] != None :
							organization_name = cell_dict['Organization Name(Youth development)']
							address 		  = cell_dict['Location']
							try:  
								org_obj = Organization.objects.filter(organization_name = organization_name,address=address).first()
								print(org_obj.id)
								print(org_obj.organization_name)
								excel_org_save(org_obj,cell_dict,0)
							except:
								new_org_obj = Organization.objects.create(organization_name = organization_name,address=address,user_id =7,parent_id=0)
								print('lllllllllllllllllll')
								print(org_obj.id)

								excel_org_save(new_org_obj,cell_dict,1)
								
			if sheetName == 'Opportunity':
				org_obj = None
				ws = wb.get_sheet_by_name('Opportunity')
				for index, row in enumerate(ws.iter_rows(),start=2):
					c=1 
					cell_dict={}
					for cell in row:
						cell_dict[ws.cell(row=1, column=c).value] = ws.cell(row=index, column=c).value
						c+=1
					if 'Organization Name(Youth development)' in cell_dict :
						if cell_dict['Organization Name(Youth development)'] != None :
							organization_name = cell_dict['Organization Name(Youth development)']
							address 		  = cell_dict['Location']
							try:  
								org_obj = Organization.objects.filter(organization_name = organization_name,address=address).first()
								print(org_obj.id)
								print(org_obj.organization_name)
								excel_org_save(org_obj,cell_dict,0)
							except:
								org_obj = Organization.objects.create(organization_name = organization_name,address=address,user_id =user.id,parent_id=0)
								cordinatorRequest = CordinatorRequest(
									user_id = user,
									org_id = org_obj,
									status = 'Pending',
									role = int(2),
									oppurtunity_id = 0,
									is_url = False,
									is_request = "coordinator"
								)  
								cordinatorRequest.save()
							org_id_id =116
							if org_obj != None :
								getorgId = CordinatorRequest.objects.filter(user_id_id = user.id,is_request = 'coordinator',org_id=org_obj).values('org_id_id')
								if len(getorgId) > 0:
									org_id_id = getorgId[0]['org_id_id']
								excel_op_save(org_id_id,user,org_obj,cell_dict)

					

	return HttpResponse(1)

def excel_op_save(org_id_id,user,org_obj,cell_dict):
	no_of_volunteers = 1
	if cell_dict['Number of Volunteers Needed'] != None and cell_dict['Number of Volunteers Needed'] != 'NA' :
		no_of_volunteers = no_of_volunteers

	opportunitiy = Opportunities(
		user_id = user,
		org_id = org_obj,
		opportunity_name = cell_dict['Opportunity Name'],
		parent_id = 0,
		lat = 22.2345,
		lon = 88.4567,
		parent_opportunity = 0,
		start_date = timezone.now(),
		end_date = timezone.now(),
		no_of_volunteers= no_of_volunteers
	)
	opportunitiy.save()
	catObj = ActivityCategory.objects.get(id =1)
	opportunityCategories = OpportunityCategories(
		opportunity_id = opportunitiy,
		category_id = catObj
	)
	opportunityCategories.save()

def excel_org_save(org_obj,cell_dict,new_reqest):
	# org_obj.parent_id = org_obj.parent_id

	if cell_dict['Cause'] != None and cell_dict['Cause'] != 'NA' :
		org_obj.cause = cell_dict['Cause']

	if cell_dict['Location'] != None and cell_dict['Location'] != 'NA' :
		org_obj.address = cell_dict['Location']
	
	if cell_dict['Email'] != None and cell_dict['Email'] != 'NA' :
		org_obj.email = cell_dict['Email']
	
	if cell_dict['IRS rank'] != None and cell_dict['IRS rank'] != 'NA' :
		org_obj.irs_rank = cell_dict['IRS rank']
	
	if cell_dict['Website Link'] != None and cell_dict['Website Link'] != 'NA' :
		org_obj.web_url = cell_dict['Website Link']

	if cell_dict['Location1'] != None and cell_dict['Location1'] != 'NA' :
		org_obj.address1 = cell_dict['Location1']

	if cell_dict['Contact Number2'] != None and cell_dict['Contact Number2'] != 'NA' :
		org_obj.phone2 = cell_dict['Contact Number2']
	
	if cell_dict['FB Page'] != None and cell_dict['FB Page'] != 'NA' :
		org_obj.fb_url = cell_dict['FB Page']

	if cell_dict['EIN'] != None and cell_dict['EIN'] != 'NA' :
		org_obj.tax_id = cell_dict['EIN']
	
	if cell_dict['Who we are?'] != None and cell_dict['Who we are?'] != 'NA' :
		org_obj.about_us = cell_dict['Who we are?']
	
	if cell_dict['Event 1'] != None and cell_dict['Event 1'] != 'NA' :
		org_obj.event1 = cell_dict['Event 1']
	
	if cell_dict['Event 2'] != None and cell_dict['Event 2'] != 'NA' :
		org_obj.event2 = cell_dict['Event 2']
	
	if cell_dict['Event 3'] != None and cell_dict['Event 3'] != 'NA' :
		org_obj.event3 = cell_dict['Event 3']
	
	if cell_dict['Event 3'] != None and cell_dict['Event 3'] != 'NA' :
		org_obj.event3 = cell_dict['Event 3']

	if cell_dict['Submission Date'] != None and cell_dict['Submission Date'] != 'NA' :
		org_obj.submission_date = cell_dict['Submission Date']

	org_obj.why_us = cell_dict['What we do?']
	# org_obj.why_us = cell_dict['What we do?'] + cell_dict['why should we valunteer with you?']

	org_obj.save() 

	if new_reqest == 1 :
		cordinatorRequest = CordinatorRequest(
			user_id = User.objects.get(id=7),
			org_id = org_obj,
			status = 'Pending',
			role = int(2),
			oppurtunity_id = 0,
			is_url = False,
			is_request = "coordinator"
		)  
		cordinatorRequest.save()


def  upload_opportunity(request):
	import csv
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		data = []
		data_result = []
		all_volunteer ={}
		getfile = request.FILES['csv_file']

		# print(getfile)
		after_decode = getfile.read().decode('utf-8').splitlines()
		reader = csv.DictReader(after_decode)

		cursor = connection.cursor()
		for fetchrow in reader:
			data.append(fetchrow)
			if fetchrow['Organization Name']:
				ogranization_search_sql = "SELECT * FROM adminpanel_organization WHERE affiliated_org=0 and organization_name ='"+fetchrow['Organization Name']+"'"
				cursor.execute(ogranization_search_sql)
				fetch_organization = dictfetchall(cursor)
				org_is_present = len(fetch_organization)
				if org_is_present > 0:
					org_id_id = fetch_organization[0]['id']
				else:
					organization_insert_sql = "INSERT INTO adminpanel_organization(organization_name, parent_id, address, tax_id, email, web_url, phone, cause, about_us, our_activity, why_us, photo, address1, event1, event2, event3, fb_url, irs_rank, phone2, submission_date,affiliated_org) values ('"+fetchrow['Organization Name']+"',0,'"+fetchrow['Location']+"',0,'','', null ,'"+fetchrow['Cause']+"','','','','','','','','','','','','"+fetchrow['Submission Date']+"','0')"
					cursor.execute(organization_insert_sql)
					org_id_id = cursor.lastrowid
					
			if fetchrow['Opportunity Name']:
				opportunity_search_sql = "SELECT * FROM auth_user WHERE username = '"+fetchrow['Coordinator Name']+"'"
				cursor.execute(opportunity_search_sql)
				fetch_opportunity = dictfetchall(cursor)
				opportunity_is_present = len(fetch_opportunity)
				if opportunity_is_present > 0:
					opportunity_id_id = fetch_opportunity[0]['id']
				else:

					if fetchrow['Opportunity Description'] == 'NA':
						Opportunity_Description = ''
					else:
						Opportunity_Description = fetchrow['Opportunity Description']
					
					if fetchrow['Number of Volunteers Needed'] == 'NA':
						Number_of_Volunteers_Needed = 0
					else:
						Number_of_Volunteers_Needed = fetchrow['Number of Volunteers Needed']
					# return HttpResponse(Number_of_Volunteers_Needed)	
					opportunity_insert_sql = "INSERT INTO adminpanel_opportunities(opportunity_name,author_name,address,parent_id,start_date,end_date,org_id_id,user_id_id,description,no_of_volunteers) VALUES('"+str(fetchrow['Opportunity Name'].replace("'", ""))+"','"+str(fetchrow['Coordinator Name'])+"','"+str(fetchrow['Location'])+"',0,'2019-07-19 12:00:00','2019-07-19 12:00:00',"+str(org_id_id)+",67,'"+str(Opportunity_Description)+"',"+str(Number_of_Volunteers_Needed)+")"
					cursor.execute(opportunity_insert_sql)
					print(opportunity_insert_sql)

	data_result = json.dumps({"Ack": 1}, cls=DjangoJSONEncoder)
		

	return HttpResponse(data_result)
########################Done7 ###########################
def upload_csv(request):
	import csv
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	
	if request.method == 'POST':
		data = []
		data_result = []
		all_volunteer ={}
		if request.FILES:
			 
			getfile = request.FILES['csv_file']
			opportunity_id = request.POST.get('opt_id',None)
			user_id = 0 
			try:
				opportunity_Obj = Opportunities.objects.get(id = opportunity_id )
				organization_id = opportunity_Obj.org_id.id
				after_decode = getfile.read().decode('utf-8').splitlines()
				reader = csv.DictReader(after_decode, fieldnames=['Name','Email','Phoneno','Organization Name','Activity Hours'])

				cursor = connection.cursor()
				approvelist ={}
				i=0
				for fetchrow in reader:
					data.append(fetchrow)
					
					if fetchrow['Name'] and fetchrow['Email'] and fetchrow['Phoneno'] and fetchrow['Organization Name'] and fetchrow['Activity Hours']:
						
						total = Volunteer.objects.count()
						if fetchrow['Organization Name'] == '':
							fetchrow['Organization Name'] = 0

						if fetchrow['Activity Hours'] == '':
							fetchrow['Activity Hours'] = 0

						total_number = User.objects.filter(email=str(fetchrow['Email'])).count()
		
						nam_split = fetchrow['Name'].split(' ')

						if len(nam_split) > 1:
							fname = nam_split[0]
							lname = nam_split[1]
						elif len(nam_split) == 1:
							fname = nam_split[0]
							lname = ''
						else:
							fname = ''
							lname = ''
						if total_number == 0:
							user = User.objects.create_user(
								password='123456',
								is_superuser=False,
								username=fetchrow['Email'],
								first_name=fname,
								last_name=lname,
								email=fetchrow['Email'],
								is_staff=False,
								is_active=True
							)
							user.save()
							user_id=user.id

							user_info = UserProfile(
								address='',
								latitude=22.57,
								longitude=88.36,
								is_verified=0,
								interest_id_id=1,
								role_id=2,
								user_id_id=user_id,
								activate_token='123456',
								phone_number = fetchrow['Phoneno']
							)
							user_info.save() 

							## Affiliate organization ##

							affiliate_organization_exist = Organization.objects.filter(affiliated_org=1,user_id=user_id).count()
							if affiliate_organization_exist > 0 :
								affiliate_organization = Organization.objects.filter(affiliated_org=1, organization_name=str(fetchrow['Organization Name']),user_id = user_id).last()
								affiliate_organization.organisation_name = str(fetchrow['Organization Name'])
								affiliate_organization.affiliated_activity_hours = fetchrow['Activity Hours']
								affiliate_organization.save()
							else:
								affiliate_organization = Organization(affiliated_org=1, organization_name=str(fetchrow['Organization Name']),affiliated_activity_hours=fetchrow['Activity Hours'] , user_id=user_id).save()
							
							## Affiliate organization ##
							
							apply_coordinator = CordinatorRequest(
								user_id_id=str(user_id),
								org_id_id=str(organization_id),
								status="Approved",
								address='N/A',
								employee_number=0,
								oppurtunity_id=str(opportunity_id),
								role=2,
								is_request = "volunteer"
							) 
							apply_coordinator.save()
							cursor4 = connection.cursor() 

							sql4 = "select DISTINCT(RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name))))) AS Full_Name, adminpanel_userprofile.profile_image as pending_user_image, auth_user.first_name,auth_user.last_name,auth_user.email,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.id="+str(apply_coordinator.id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
							cursor4.execute(sql4)
							approvelist_details = dictfetchall(cursor4)
							if i > 0:
								approvelist[i] = approvelist_details
							i+=1
						else:
							user_details = User.objects.filter(email=str(fetchrow['Email'])).values('id')
							user_id = user_details[0]['id']

							## Affiliate organization ##

							affiliate_organization_exist = Organization.objects.filter(affiliated_org=1,user_id=user_id).count()
							if affiliate_organization_exist > 0 :
								affiliate_organization = Organization.objects.filter(affiliated_org=1, organization_name=str(fetchrow['Organization Name']),user_id = user_id).last()
								affiliate_organization.organisation_name = str(fetchrow['Organization Name'])
								affiliate_organization.affiliated_activity_hours = fetchrow['Activity Hours']
								affiliate_organization.save()
							else:
								affiliate_organization = Organization(affiliated_org=1, organization_name=str(fetchrow['Organization Name']),affiliated_activity_hours=fetchrow['Activity Hours'] , user_id=user_id).save()
							
							## Affiliate organization ##
						

							## chk existance of co-ordinator in Approved status
							chk_approved_coordinate_req = CordinatorRequest.objects.filter(user_id_id = str(user_id), org_id_id = str(organization_id), status = "Approved", oppurtunity_id = str(opportunity_id), is_request = "volunteer").count()
	
							if chk_approved_coordinate_req == 0 :
								apply_coordinator = CordinatorRequest(
									user_id_id=str(user_id),
									org_id_id=str(organization_id),
									status="Approved",
									address='N/A',
									employee_number=0,
									oppurtunity_id=str(opportunity_id),
									role=2,
									is_request = "volunteer"
								) 
								apply_coordinator.save()
							else: 
								apply_coordinator = CordinatorRequest.objects.filter(user_id_id = str(user_id), org_id_id = str(organization_id), status = "Approved", oppurtunity_id = str(opportunity_id), is_request = "volunteer").last()
							## chk existance of co-ordinator in Approved status

							cursor4 = connection.cursor()
							sql4 = "select DISTINCT(RTRIM(CONCAT(LTRIM(RTRIM(auth_user.first_name)) , ' ' , LTRIM(RTRIM(auth_user.last_name))))) AS Full_Name, adminpanel_userprofile.profile_image as pending_user_image, auth_user.first_name,auth_user.last_name,auth_user.email,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id left join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.id="+str(apply_coordinator.id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
							cursor4.execute(sql4)
							approvelist_details = dictfetchall(cursor4)
							if i >= 0:
								approvelist[i]= approvelist_details
							i+=1
							
							data_result = json.dumps({"Ack": 0, "msg": "already present","approvelist":approvelist}, cls=DjangoJSONEncoder)
					else:
						data_result = json.dumps({"Ack": 0, "msg": "format doesnot not matched"}, cls=DjangoJSONEncoder)

				data_result = json.dumps({"Ack": 1, "msg": "Successfully uploaded","user_id":user_id, "organization_id":organization_id,"opportunity_id": opportunity_id,"approvelist":approvelist}, cls=DjangoJSONEncoder)
			except ObjectDoesNotExist :
				return JsonResponse({'ACK':0,'msg':'Invalid Opportunity'})
		else:
			data_result = json.dumps({"Ack": 0, "msg": "File missig"}, cls=DjangoJSONEncoder)
	
	return HttpResponse(data_result)


def get_oppoptunity_volunteertime(request):
	userarr=[]
	usrarr1=[]
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	cursor = connection.cursor()
	start_date = request.POST.get('start_date', None)
	end_date = request.POST.get('end_date', None)
	sql = "select * from adminpanel_opportunities where adminpanel_opportunities.start_date >= '"+start_date+"' AND adminpanel_opportunities.end_date <= '"+end_date+"'"
	
	cursor.execute(sql)
	alluserDetails = dictfetchall(cursor)
	for userDetails in alluserDetails:
		usrdt={}
		usrdt['id']=int(userDetails['id'])
		usrdt['user_id']=int(userDetails['user_id_id'])
		usrdt['opportunity_name']=str(userDetails['opportunity_name'])
		usrdt['start_date']=str(userDetails['start_date'])
		usrdt['end_date']=str(userDetails['end_date'])
		sql1 = "SELECT *,EXTRACT(EPOCH from (adminpanel_opportunities.end_date - adminpanel_opportunities.start_date)/3600) as time_cal from adminpanel_opportunities where adminpanel_opportunities.id >= '"+str(userDetails['id'])+"'"
		cursor.execute(sql1)
		alluserinfo = dictfetchall(cursor)
		for allusr in alluserinfo:
			usrdt1={}
			usrdt1['time']=str(allusr['time_cal'])
			usrarr1.append(usrdt1)
		usrdt['time']=usrarr1
		userarr.append(usrdt)
	data = json.dumps({"Ack": 1, "volunteeruser": userarr}, cls=DjangoJSONEncoder)
	return HttpResponse(data)

###############Done2################
def excel_export(request):
	import json
	import wget
	
	url = settings.UPLOAD_URL_ROOT_CSV+'/opportunity.xlsx'
	wget.download(url)
	
	data = json.dumps({"Ack":"1", "text":"Suucessfully downloaded"})
	return HttpResponse(data)

###############Done2################
def excel_export_test(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	import wget
	import os

	cursor = connection.cursor()
	sql = "SELECT * FROM auth_user"
	cursor.execute(sql)
	result_get = dictfetchall(cursor)
	
	url = settings.UPLOAD_URL_ROOT_CSV+'/new_file.xlsx'
	
	workbook = Workbook(os.path.abspath('media/all_excel/new_file.xlsx'))
	worksheet = workbook.add_worksheet()
	value ={}
	
	all_keys = result_get[0].keys()
	for key_len in range (len(all_keys)):
		worksheet.write(0, key_len, list(all_keys)[key_len])
	for row in range (len(result_get)):
		count = 0
		for c, col in result_get[row].items():
			worksheet.write(row+1, count, col)
			count += 1

	workbook.close()

	if result_get:
		data_array = json.dumps({"Ack":"1", "text":"file write done","url":url})
	else:
		data_array = json.dumps({"Ack":"1", "text":"file write fail"})
	
	
	return HttpResponse(data_array)

def export_volunteer_report(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	import wget
	import os

	if request.method == 'POST':
		
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)

		sql2 = "SELECT cr.org_id_id FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id WHERE o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"' AND cr.status='Approved' GROUP BY cr.org_id_id"
		
		cursor = connection.cursor()
		cursor.execute(sql2)
		orgs = dictfetchall(cursor)
		opportunities = []
		if orgs:
			for i in range(len(orgs)):
				
				org_id = orgs[i]['org_id_id']
				sql = "SELECT o.opportunity_name, COALESCE(to_char(o.start_date, 'MM-DD-YYYY HH24:MI:SS'), '') AS start, COALESCE(to_char(o.end_date, 'MM-DD-YYYY HH24:MI:SS'), '') AS end, EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, au.first_name, au.last_name FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id WHERE o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"' AND cr.status='Approved' AND cr.org_id_id="+str(org_id)
		
				cursor = connection.cursor()
				cursor.execute(sql)
				opportunities = dictfetchall(cursor)
		
		url = settings.UPLOAD_URL_ROOT_CSV+'/volunteer_report.xlsx'

		workbook = Workbook(os.path.abspath('media/all_excel/volunteer_report.xlsx'))
		worksheet = workbook.add_worksheet()
		value ={}

		all_keys = opportunities[0].keys()
		for key_len in range (len(all_keys)):
			worksheet.write(0, key_len, list(all_keys)[key_len])
		for row in range (len(opportunities)):
			count = 0
			for c, col in opportunities[row].items():
				worksheet.write(row+1, count, col)
				count += 1

		workbook.close()	
		if opportunities:
			data = json.dumps({"Ack":"1", "text":"file write done","url":url})
		else:
			data = json.dumps({"Ack":"1", "text":"file write fail"})
	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
	return HttpResponse(data)

def export_engagethree_report(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	import wget
	import os

	if request.method == "POST":
		start_date = request.POST.get('start_date', None)
		end_date = request.POST.get('end_date', None)
		user_id = request.POST.get('volunteer_id', None)
		cursor = connection.cursor()
		sql = "SELECT *,EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id WHERE cr.user_id_id = "+user_id+" AND o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"'"
		cursor.execute(sql)
		opportunities = dictfetchall(cursor)


		sql2 = "SELECT cr.org_id_id FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id WHERE cr.user_id_id = "+user_id+" AND o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"' AND cr.status='Approved' GROUP BY cr.org_id_id"
		
		cursor = connection.cursor()
		cursor.execute(sql2)
		orgs = dictfetchall(cursor)
		opportunities = []
		if orgs:
			for i in range(len(orgs)):
				
				org_id = orgs[i]['org_id_id']
				sql = "SELECT o.opportunity_name, COALESCE(to_char(o.start_date, 'MM-DD-YYYY HH24:MI:SS'), '') AS start, COALESCE(to_char(o.end_date, 'MM-DD-YYYY HH24:MI:SS'), '') AS end, EXTRACT(EPOCH FROM (o.end_date - o.start_date)/3600) as time_cal, cr.user_id_id as userid, au.first_name, au.last_name FROM adminpanel_cordinatorrequest cr LEFT JOIN adminpanel_opportunities o ON cr.oppurtunity_id = o.id LEFT JOIN auth_user au ON cr.user_id_id = au.id WHERE cr.user_id_id = "+user_id+" AND o.start_date >= '"+start_date+"' AND o.end_date <= '"+end_date+"' AND cr.status='Approved' AND cr.org_id_id="+str(org_id)
		
				cursor = connection.cursor()
				cursor.execute(sql)
				opportunities = dictfetchall(cursor)

		url = settings.UPLOAD_URL_ROOT_CSV+'/engagethree_report.xlsx'

		workbook = Workbook(os.path.abspath('media/all_excel/engagethree_report.xlsx'))
		worksheet = workbook.add_worksheet()
		value ={}

		all_keys = opportunities[0].keys()
		for key_len in range (len(all_keys)):
			worksheet.write(0, key_len, list(all_keys)[key_len])
		for row in range (len(opportunities)):
			count = 0
			for c, col in opportunities[row].items():
				worksheet.write(row+1, count, col)
				count += 1

		workbook.close()	
		if opportunities:
			data = json.dumps({"Ack":"1", "text":"file write done","url":url})
		else:
			data = json.dumps({"Ack":"1", "text":"file write fail"})

	else:
		data = json.dumps({"Ack": 0, "msg": "Post method supported only"})
	return HttpResponse(data)	

##########Done1#############
def export_organization_report(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	import wget
	import os

	result_get =[]
	organization_details = Organization.objects.all()
	result_get = fetch_row_organization(organization_details)

	url = settings.UPLOAD_URL_ROOT_CSV+'/organization_report.xlsx'
	
	workbook = Workbook(os.path.abspath('media/all_excel/organization_report.xlsx'))
	worksheet = workbook.add_worksheet()
	value ={}
	
	all_keys = result_get[0].keys()
	for key_len in range (len(all_keys)):
		worksheet.write(0, key_len, list(all_keys)[key_len])
	for row in range (len(result_get)):
		count = 0
		for c, col in result_get[row].items():
			worksheet.write(row+1, count, col)
			count += 1

	workbook.close()

	if result_get:
		data_array = json.dumps({"Ack":"1", "text":"file write done","url":url})
	else:
		data_array = json.dumps({"Ack":"1", "text":"file write fail"})
	
	return HttpResponse(data_array)

###############Done2################	
def excel_export_organization(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	import wget
	import os

	result_get =[]
	organization_details = Organization.objects.all()
	result_get = fetch_row_organization(organization_details)

	url = settings.UPLOAD_URL_ROOT_CSV+'/organization_report.xlsx'

	workbook = Workbook(os.path.abspath('media/all_excel/organization_report.xlsx'))
	worksheet = workbook.add_worksheet()
	value ={}

	for key_len in range (len(all_keys)):
		return HttpResponse(list(all_keys))
		worksheet.write(0, key_len, list(all_keys)[key_len])
	for row in range (len(result_get)):
		count = 0
		for c, col in result_get[row].items():
			worksheet.write(row+1, count, col)
			count += 1

	workbo
	ok.close()
	if result_get:
		data_array = json.dumps({"Ack":"1", "text":"file write done","url":url})
	else:
		data_array = json.dumps({"Ack":"1", "text":"file write fail"})
	
	
	return HttpResponse(data_array)


###############Done2 difference################						
def gettest(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	# return HttpResponse(1)
	if request.method == 'POST':
		
		cursor = connection.cursor()
		cursor.execute("SELECT * FROM adminpanel_volunteer")
		all_value = dictfetchall(cursor)
		data = json.dumps({"ACK": 0, "msg": "Get Method supported executed!!!","all_value": all_value})
		return HttpResponse(data)
		

##########Done#############

def get_organizationlist(request):
	import json
	org_details =[]
	organization_details = Organization.objects.all()
	org_details = fetch_row_organization(organization_details)

	data = json.dumps({"Ack": 1, "organizations": org_details})
	return HttpResponse(data)

##########Done1#############
def get_organization_by_name(request):
	import json
	import datetime
	if request.method == 'POST':
		org_name_case = request.POST.get('org_name', None)
		org_name = org_name_case.lower()
		
		org_details = []
		organization_details = Organization.objects.filter(organization_name__icontains=org_name)
		org_details = fetch_row_organization(organization_details)

		data = json.dumps({"Ack": 1, "organization": org_details}) 
	return HttpResponse(data)

##########Done#############

class get_organization_by_id(APIView):
	
	permission_classes = (IsAuthenticated,)
	
	def post(self,request):
		import json
		import datetime
		id = request.POST.get('id', None)
		
		org_details = []
		
		organization_details = Organization.objects.filter(id=str(id))
		org_details = fetch_row_organization(organization_details)
			
		data = json.dumps({"Ack": 1, "organization": org_details})
		return HttpResponse(data)

##########Done1#############
def get_opportunity_by_organization(request,offset):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		id = request.POST.get('org_id', None)
		
		oppor_details = []
		offset =  int(offset)
		limit = offset+12

		# print(int(offset,2)) 
		# return HttpResponse(offset)
		# new_limit = str(offset)+str(limit)
		opportunity_details = Opportunities.objects.filter(org_id_id=str(id))[offset:limit] 
		if opportunity_details:
			oppor_details = fetch_row_opportunity(opportunity_details) 
		data = json.dumps({"Ack": 1, "opportunities": oppor_details }, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

##########Done1#############
def get_test(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		id = request.POST.get('id', None)
		oppor_details = []

		opportunity_details = Opportunities.objects.filter(id=str(id))
		oppor_details = fetch_row_opportunity(opportunity_details) 
		data = json.dumps({"Ack": 1, "opportunities": oppor_details }, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

##########Done############# 

# def update_role(request):
class update_role(APIView):
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		import json
		from django.core.serializers.json import DjangoJSONEncoder
		data = {}
		user_id = request.POST.get('user_id',None)
		request_id = request.POST.get("request_id",None)
		org_id_id = request.POST.get("org_id",None)
		CordinatorRequest.objects.filter(user_id_id=str(user_id),org_id_id=str(org_id_id)).delete()
		UserProfile.objects.filter(user_id_id=str(user_id)).update(role_id=2)
		data = json.dumps({"Ack": 1, "msg": "Successfully Updated"})
		return HttpResponse(data)

	def get(self,request):
		data = {}
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed","time": current})
		return HttpResponse(data)

	def put(self,request):
		data = {}
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed","time": current})
		return HttpResponse(data)

	def patch(self,request):
		data = {}
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed","time": current})
		return HttpResponse(data)
	
	def delete(self,request):
		data = {}
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed","time": current})
		return HttpResponse(data)

##########Done#############

def save_contactus(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	current = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
	data = {}
	msg_html = ""
	if request.method == 'POST':
		contactUsEmail = request.POST.get('contactUsEmail',None)
		contactUsDescription = request.POST.get('contactUsDescription', None)

		contactUs = ContactUs(
			fromemail = contactUsEmail,
			message = contactUsDescription,
			is_replied = 0,
			entry_date = current
		) 
		contactUs.save()
		adminEmail = 'ainteer72@gmail.com'
		msg_html = '<p>Hi Admin,</p><p>'+contactUsDescription+'</p><p><p>Thanks,</p><p>'+contactUsEmail+'</p> '
		if send_mail('Contact Us Email', 'InterApp', 'kher.nachiket@gmail.com', ['ainteer72@gmail.com'], html_message=msg_html):
			return HttpResponse(1)
		else:
			return HttpResponse(0)

	data = json.dumps({"Ack": 0, "msg": "Only post method allowed","time": current})
	return HttpResponse(data)

##########Done5#############
def fetch_row_opportunity(opportunity_details):
	oppor_details = []
	if opportunity_details :
		for oppor in opportunity_details:
				opportunity = {}
				opportunity['id'] = oppor.id
				opportunity['opportunity_name'] = str(oppor.opportunity_name)
				opportunity['user_id'] = str(oppor.user_id)
				opportunity['parent_id'] = int(oppor.parent_id) 
				opportunity['author_name'] = str(oppor.author_name)
				opportunity['address'] = str(oppor.address)
				opportunity['org_id'] = str(oppor.org_id)
				opportunity['image'] = str(oppor.image)
				opportunity['no_ofyear'] = str(oppor.no_ofyear)
				opportunity['parent_opportunity'] = int(oppor.parent_opportunity)
				opportunity['start_date'] = str(oppor.start_date)
				opportunity['end_date'] = str(oppor.end_date)
				opportunity['description'] = str(oppor.description)
				opportunity['no_of_volunteers'] = int(oppor.no_of_volunteers)
				opportunity['repeat_number'] = int(oppor.repeat_number)
				opportunity['lat'] = oppor.lat
				opportunity['lon'] = oppor.lon 
				oppor_details.append(opportunity)
	return oppor_details

##########Done5#############

def fetch_row_organization(organization_details):
	org_details = []
	if organization_details :
		for org in organization_details:
				organization = {}
				organization['id'] = org.id
				organization['organization_name'] = str(org.organization_name)
				organization['parent_id'] = org.parent_id
				organization['address'] = str(org.address)
				organization['tax_id'] = str(org.tax_id)
				organization['email'] = str(org.email) 
				organization['web_url'] = str(org.web_url)
				organization['phone'] = str(org.phone)
				organization['cause'] = str(org.cause)
				organization['about_us'] = str(org.about_us)
				organization['our_activity'] = str(org.our_activity)
				organization['why_us'] = str(org.why_us)
				organization['photo'] = str(org.photo)
				organization['address1'] = str(org.address1)
				organization['phone2'] = str(org.phone2)
				organization['irs_rank'] = str(org.irs_rank)
				organization['fb_url'] = str(org.fb_url)
				organization['event1'] = str(org.event1)
				organization['event2'] = str(org.event2)
				organization['event3'] = str(org.event3)
				org_details.append(organization)
	return org_details

##########Done5#############
def vol_cor_details(sql4):
	volunteerexist = []
	if sql4 : 
		for query4 in sql4:
			coorninator = {}
			coorninator['cid'] = query4.id
			coorninator['org_id'] = str(query4.org_id_id)
			coorninator['user_id'] = query4.user_id_id
			coorninator['org_name'] = str(query4.organization_name)
			coorninator['address'] = str(query4.address)
			coorninator['status'] = str(query4.status) 
			coorninator['role'] = str(query4.role)
			coorninator['oppurtunity_id'] = str(query4.oppurtunity_id)
			coorninator['is_url'] = str(query4.is_url)
			coorninator['parent_id'] = query4.parent_id
			coorninator['tax_id'] = str(query4.tax_id)
			coorninator['email'] = str(query4.email) 
			coorninator['web_url'] = str(query4.web_url)
			coorninator['phone'] = str(query4.phone)
			coorninator['cause'] = str(query4.cause)
			coorninator['about_us'] = str(query4.about_us)
			coorninator['our_activity'] = str(query4.our_activity)
			coorninator['why_us'] = str(query4.why_us)
			coorninator['photo'] = str(query4.photo)
			coorninator['address1'] = str(query4.address1)
			coorninator['phone2'] = str(query4.phone2)
			coorninator['irs_rank'] = str(query4.irs_rank)
			coorninator['fb_url'] = str(query4.fb_url)
			coorninator['event1'] = str(query4.event1)
			coorninator['event2'] = str(query4.event2)
			coorninator['event3'] = str(query4.event3)
			volunteerexist.append(coorninator)
	return volunteerexist

#########Done5#############
def get_organization_by_user11(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
		organizationObjs = Opportunities.objects.all().select_related()
		opportunities = []
		for i in range(len(organizationObjs)):
			data1 = {}
			# print(organizationObjs[i])
			data1[i] = organizationObjs[i]
			opportunities=data1

		data = json.dumps({"Ack": 1,'opportunities':opportunities}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"})
	return HttpResponse(data)

#########Done5#############
def cronset(request):
	import json
	import datetime

	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'GET':
		cursor = connection.cursor()
		msg_html2 = "" 
		msg_html = "" 

		dt = date.today() + timedelta(2)
		# dt = datetime.datetime.strptime(dt, "%Y-%m-%d")
		# return HttpResponse(dt)
		sql = "select current_date, AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and date(AO.start_date) = current_date"
		cursor.execute(sql)
		opportunity_details_today = dictfetchall(cursor) 

		coordinate_data1={}
		if opportunity_details_today:
			for i in range(len(opportunity_details_today)):
				o_id1=opportunity_details_today[i]['id']

				opportunity_details_today[i]['opportunity_id']=opportunity_details_today[i]['id']
				cursor4 = connection.cursor()
				sql4 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id1)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'" 
				cursor4.execute(sql4)
				approvelist1 = dictfetchall(cursor4)
				opportunity_details_today[i]['is_volunteer']=approvelist1
				if o_id1:
					link = settings.PATH_URL+"/opportunity-details/"+str(o_id1)
				# print(len(pendinglist))
				if len(approvelist1) > 0:
					for j in range(len(approvelist1)):
						email1 = approvelist1[j]['email']

						msg_html += '<p> Hi there, you have a volunteering opportunity coming up today. Make preparations for the same.</p> <a href="'+link+'"></p><div style="width: 300px; border: 1px solid #d8d8d8;"><div style="width: 300px; height: 180px;">'
						if opportunity_details_today[i]['image']: 
							msg_html += '<img src="'+settings.IMAGE_URL+opportunity_details_today[i]['image']+'" style="width: 100%; height: 100%;">'
						else:
							 msg_html +='<img src="'+settings.APP_URL+'app/assets/img/default.png" style="width: 100%; height: 100%;">' 
						
						msg_html +='</div><div style="padding: 0px 20px 20px;"><div style="display: flex; align-items: center;"><div style="margin-right: 10px;"><p style="line-height: 35px;background: #fff;border: 1px solid #d8d8d8; width: 35px; height: 35px; text-align: center; border-radius: 50%;">TA</p></div><div><p style="margin: 0px; font-family: "Open Sans", sans-serif; ">'+opportunity_details_today[i]['organization_name']+'</p><small>'+opportunity_details_today[i]['first_name']+' '+opportunity_details_today[i]['last_name']+'</small></div></div><div><h3 style="font-family: Open Sans, sans-serif; margin: 0px; font-size: 13px;">'+opportunity_details_today[i]['opportunity_name']+'</h3><p style="font-family: Open Sans, sans-serif; margin: 0px; margin-top: 5px; font-size: 13px; color: #676767;">'+opportunity_details_today[i]['description']+'</p></div></div></div></a><p>Thanks,</p><p>Inteer Team</p>'

						send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', [email1], html_message=msg_html)


				
				oppurtunityArr1=opportunity_details_today[i]
				coordinate_data1[i]=oppurtunityArr1	

		sql = "select AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and AO.start_date < '"+str(dt)+"' AND AO.start_date > current_date"
		cursor.execute(sql)
		opportunity_details_next_two_days = dictfetchall(cursor)

		coordinate_data={}
		if opportunity_details_next_two_days:
				for i in range(len(opportunity_details_next_two_days)):
					o_id=opportunity_details_next_two_days[i]['id']
					opportunity_details_next_two_days[i]['opportunity_id']=opportunity_details_next_two_days[i]['id']
					cursor3 = connection.cursor()
					sql3 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.oppurtunity_id="+str(o_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'" 
					cursor3.execute(sql3)
					approvelist = dictfetchall(cursor3)
					opportunity_details_next_two_days[i]['is_volunteer']=approvelist
					if o_id:
						link = settings.PATH_URL+"/opportunity-details/"+str(o_id)
					# print(len(pendinglist))
					if len(approvelist) > 0:
						for j in range(len(approvelist)):
							name = approvelist[j]['first_name']+' '+approvelist[j]['last_name']
							email = approvelist[j]['email']

							msg_html2 += '<p>Hi there, We noticed you have a volunteering opportunity coming up in the next 2 days. Let us know if you need any information.</p> <a href="'+link+'"></p><div style="width: 300px; border: 1px solid #d8d8d8;"><div style="width: 300px; height: 180px;">'
							if opportunity_details_next_two_days[i]['image']: 
								msg_html2 += '<img src="'+settings.IMAGE_URL+opportunity_details_next_two_days[i]['image']+'" style="width: 100%; height: 100%;">'
							else:
								msg_html2 +='<img src="'+settings.APP_URL+'app/assets/img/default.png" style="width: 100%; height: 100%;">' 
							
							msg_html2 +='</div><div style="padding: 0px 20px 20px;"><div style="display: flex; align-items: center;"><div style="margin-right: 10px;"><p style="line-height: 35px;background: #fff;border: 1px solid #d8d8d8; width: 35px; height: 35px; text-align: center; border-radius: 50%;">TA</p></div><div><p style="margin: 0px; font-family: "Open Sans", sans-serif; ">'+opportunity_details_next_two_days[i]['organization_name']+'</p><small>'+opportunity_details_next_two_days[i]['first_name']+' '+opportunity_details_next_two_days[i]['last_name']+'</small></div></div><div><h3 style="font-family: Open Sans, sans-serif; margin: 0px; font-size: 13px;">'+opportunity_details_next_two_days[i]['opportunity_name']+'</h3><p style="font-family: Open Sans, sans-serif; margin: 0px; margin-top: 5px; font-size: 13px; color: #676767;">'+opportunity_details_next_two_days[i]['description']+'</p></div></div></div></a><p>Thanks,</p><p>Inteer Team</p>'

							send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', [email], html_message=msg_html2)


					
					oppurtunityArr=opportunity_details_next_two_days[i]
					coordinate_data[i]=oppurtunityArr	

		# msg_html2 = ""
		# send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', ['maitrayee.bhaumik@cbnits.com'], html_message=msg_html2)
		data = json.dumps({"Ack": 1,"msg": "success","dt":dt,"coordinate_data":coordinate_data,"opportunity_details_today":coordinate_data1}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"}, cls=DjangoJSONEncoder)
	return HttpResponse(data)

#########Done5#############

def new_export_csv(request):

	import json
	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'POST':
	
		start_date = request.POST.get('start_date')
		
		end_date = request.POST.get('end_date')

		start_date_get_list = start_date.split(' ')[0]
		end_date_get_list = end_date.split(' ')[0]

		# print(start_date)
		# print(end_date)
		# print(start_date_get_list)
		start_date_obj = datetime.strptime(start_date_get_list, "%Y-%m-%d")
		end_date_obj   = datetime.strptime(end_date_get_list, "%Y-%m-%d")
		# print( dir(my_date) )
		#     print( start_date_obj.strftime('%b') )

		# start_date_list = start_date_get_list.split('-')
		# final_start_date = start_date_list[1]+'/'+start_date_list[2]+'/'+start_date_list[0]  # 2 for date, 1 for month, 0 for year
		# end_date_list = end_date_get_list.split('-')
		# final_end_date = end_date_list[1]+'/'+end_date_list[2]+'/'+end_date_list[0]  # 2 for date, 1 for month, 0 for year
 
 
		final_start_date = str( start_date_obj.strftime('%b') ) + ' ' + str(start_date_obj.strftime('%y')) + ',' + str(start_date_obj.strftime('%d')) + ' ' +str(start_date.split(' ')[1])
		final_end_date   = str( end_date_obj.strftime('%b') ) + ' ' + str(end_date_obj.strftime('%y')) + ',' + str(end_date_obj.strftime('%d')) + ' ' + str(end_date.split(' ')[1])
		

		category_ids = request.POST.get('abc') 		
			
		opportunity_name = request.POST.get('opportunity_name')
		opportunity_name = json.loads(opportunity_name)
 
		export_data = request.POST.get('exportData')
		export_data = json.loads(export_data)

		summary = request.POST.get('summary')
		summary = json.loads(summary)

		url = settings.UPLOAD_URL_ROOT_CSV+'/volunteer_report.xlsx'

		workbook = Workbook(os.path.abspath('media/all_excel/volunteer_report.xlsx'))
		worksheet = workbook.add_worksheet()
		value ={}     
		start_end = "Records for time period: "+final_start_date+" to "+final_end_date
		worksheet.write(0, 1, str(start_end)) 

		i= 0
		for opname in opportunity_name:
			# print(list(opportunity_name)[i])
			val = list(opportunity_name)[i]
			i=i+1
			worksheet.write(1, i, str(val))

		j= 0
		for row in export_data:
			count = 0
			print(row)

			val2= list(export_data)[j]
			for c, col in enumerate(val2):
				# print(col)
				worksheet.write(j+2, count, col)
				count += 1
			
			j+=1
			# count += 1
		a =j+2
		b = 0
		worksheet.write(a, 0, "Summary")
		for summ in summary:
			b=b+1
			worksheet.write(a, b, summ)	

		workbook.close()
				
		data = json.dumps({"Ack": 1,"msg": "success","url":url,"summary":summary}, cls=DjangoJSONEncoder)
	else:
		data = json.dumps({"Ack": 0, "msg": "Only post method allowed"}, cls=DjangoJSONEncoder)
	return HttpResponse(data)

def test_table(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	cursor4 = connection.cursor()
	# sql4 = "CREATE TABLE adminpanel_countries (id TYPE column_constraint, table_constraint table_constraint) INHERITS existing_table_name;" 
	# cursor4.execute(sql4)

	data = json.dumps({"Ack": 1,"msg": "success"}, cls=DjangoJSONEncoder)
	
	# print('here')
	return HttpResponse(data)

def fetch_country(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	cursor = connection.cursor()
	sql = "SELECT * FROM adminpanel_countries"
	cursor.execute(sql)
	countries_list = dictfetchall(cursor) 
	
	data = json.dumps({"Ack": 1,"msg": "success","all_value": countries_list}, cls=DjangoJSONEncoder)
	# print('here')
	return HttpResponse(data)

def fetch_state(request):
	import json
	from django.core.serializers.json import DjangoJSONEncoder

	if request.method == 'POST':
		country_id = request.POST['country_id']
		
		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_states WHERE country_id_id="+str(country_id)
		cursor.execute(sql)
		state_list = dictfetchall(cursor) 
	
		data = json.dumps({"Ack": 1,"msg": "success","all_state": state_list}, cls=DjangoJSONEncoder)
		# print('here')
		return HttpResponse(data)

def feedback_send(request):
	import json
	import datetime
	import base64

	from django.core.serializers.json import DjangoJSONEncoder
	if request.method == 'GET':
		cursor = connection.cursor()
		msg_html2 = "" 
		msg_html = "" 

		dt = date.today() + timedelta(2)
		# dt = datetime.datetime.strptime(dt, "%Y-%m-%d")
		# return HttpResponse(dt)
		sql = "select current_date, AO.id, AO.parent_id, AO.opportunity_name, AO.description,AO.user_id_id, AO.image, AUO.first_name, AUO.last_name, AO.org_id_id, AOR.organization_name, AOR.web_url, AOR.photo from adminpanel_opportunities AO left join auth_user AUO on AO.user_id_id = AUO.id left join adminpanel_organization AOR on AO.org_id_id = AOR.id WHERE AOR.affiliated_org=0 and date(AO.start_date) = current_date-1"
		cursor.execute(sql)
		opportunity_details_today = dictfetchall(cursor) 

		coordinate_data1={}
		if opportunity_details_today:
			for i in range(len(opportunity_details_today)):
				o_id=opportunity_details_today[i]['id']
				org_id_id = opportunity_details_today[i]['org_id_id']
				user_id_id = opportunity_details_today[i]['user_id_id']

				opportunity_details_today[i]['opportunity_id']=opportunity_details_today[i]['id']
				cursor4 = connection.cursor()
				sql4 = "select adminpanel_userprofile.profile_image as pending_user_image, auth_user.*,auth_user.email, adminpanel_cordinatorrequest.* from adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id where adminpanel_cordinatorrequest.org_id_id="+str(org_id_id)+" and adminpanel_cordinatorrequest.user_id_id="+str(user_id_id)+" and adminpanel_cordinatorrequest.role=3 and adminpanel_cordinatorrequest.status='Approved'" 
				cursor4.execute(sql4)
				approvelist1 = dictfetchall(cursor4)
				opportunity_details_today[i]['is_coordinator']=approvelist1
				link = settings.PATH_URL+"/feedback_submit/"+str(o_id)+"/"+str(opportunity_details_today[i]['user_id_id'])

				opportunity_details_today[i]['link'] = link 
				if len(approvelist1) > 0:
					for j in range(len(approvelist1)):
						email1 = approvelist1[j]['email']

						msg_html += '<p> Hi, please go through the link to give feedback about that opportunity named '+opportunity_details_today[i]['opportunity_name']+'.</p> <a href="'+link+'"></p><div style="width: 300px; border: 1px solid #d8d8d8;"><div style="width: 300px; height: 180px;">'
						
						
						msg_html +='</div><p>Thanks,</p><p>Inteer Team</p>'

						send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', [email1], html_message=msg_html)


				
				# oppurtunityArr1=opportunity_details_today[i]
				# coordinate_data1[i]=oppurtunityArr1 

		data = json.dumps({"Ack": 1,"msg": "success","dt":dt,"opportunity_details_today":opportunity_details_today}, cls=DjangoJSONEncoder)
		# print('here')
		return HttpResponse(data)

class feedback_submit(APIView):
	permission_classes = (IsAuthenticated,)
	

	def post(self,request):
		import json
		import datetime
		import base64

		from django.core.serializers.json import DjangoJSONEncoder
		cursor = connection.cursor()
		sql = "SELECT * FROM adminpanel_feedback AF LEFT JOIN adminpanel_opportunities AO ON AO.id = AF.opportunity_id LEFT JOIN auth_user AU ON AU.id = AF.user_id"
		cursor.execute(sql)
		fetch_all = dictfetchall(cursor)

		questions = request.POST.get('questions')  
		questions = json.loads(questions)
		answers = request.POST.get('answer')
		answers = json.loads(answers)

		opportunity_id = request.get.POST('opportunity_id',None)
		user_id = request.get.POST('user_id',None)

		if request.POST.get('answer') != '':
			i=0
			for answer in answers:
				sql_feedback_insert ="INSERT INTO adminpanel_feedback(opportunity_id, user_id, answer, question) VALUES ("+str(opportunity_id)+","+str(user_id)+", '"+str(answers)+"', '"+str(questions[i])+"')"
				cursor.execute(sql_feedback_insert)
				i= i+1

		volunteer_details = "SELECT * FROM adminpanel_cordinatorrequest inner join auth_user on adminpanel_cordinatorrequest.user_id_id=auth_user.id inner join adminpanel_userprofile on adminpanel_cordinatorrequest.user_id_id = adminpanel_userprofile.user_id_id WHERE adminpanel_cordinatorrequest.opportunity_id = "+str(opportunity_id)+" and adminpanel_cordinatorrequest.user_id_id = "+str(user_id)+" and adminpanel_cordinatorrequest.role=2 and adminpanel_cordinatorrequest.status='Approved'"
		cursor.execute(volunteer_details)
		fetch_all_volunteer = dictfetchall(cursor)

		
		for volunteer in fetch_all_volunteer:
			volunteer_email = volunteer['email']
			msg_html = ''
			msg_html += '<p> Hi,</p>'
			for answer in answers:
				msg_html += '<p><b> question:'+str(questions[i])+'.</b></p> <p>'+str(answer)+'</p>'
								
			msg_html +='</div><p>Thanks,</p><p>Inteer Team</p>'
			send_mail('Link shared Email', 'InterApp', 'ainteer72@gmail.com', [volunteer_email], html_message=msg_html)
		
		data = json.dumps({"Ack": 1,"msg": "success","feedback":fetch_all}, cls=DjangoJSONEncoder)
		return HttpResponse(data)  

def stripehome(request):
	key = settings.STRIPE_PUBLISHABLE_KEY
	data = json.dumps({"key": key})
	return HttpResponse(data)

class charge(APIView): # new
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		print('testbbbbbbbbbbbbbbbbbbbb')
		rtn_obj={}
		subscription_status = 0
		payload = request.body
		my_json = payload.decode('utf8')
		data = json.loads(my_json)['formdata']
		print(data)
		package = data['package']
		userObj = User.objects.get(id=data['user_id'])
		# if request.user.is_authenticated:
		if data['user_id']:
			try:
				# check user  already exist or not in the paymentTable
				user = PayementInformations.objects.filter(user=userObj,subscription_ended=False).last()
				print('jlnjnjnjlnjn')
				print(user)
				# user.subscription_ended = True
				# user.subscribed = False
				# user.save()
			except PayementInformations.DoesNotExist:
				user = None

			if user == None:
				if package == 'basic' : # BASIC
					rtn_obj  = payment_processing_execution_basic(request,userObj,package)
					return HttpResponse(json.dumps(rtn_obj))
				elif package == 'premium' : # PREMIUM
					# if not exist in DB STRIPE SUBSCRIPTION and DB INSERTION GOING
					return_dict = payment_processing_execution(request,userObj,package)  # STRIPE SUBSCRIPTION 
					return JsonResponse(return_dict,safe=False)
					
				elif package == 'enterprise' : # ENTERPRISE
					return HttpResponse(json.dumps({'ACK':0,'msg':'Discuss with client'}))
			else:
				rtn_obj['ACK']=0
				rtn_obj['msg']= 'You have already subscribe to the plan'
				rtn_obj['package'] = user.package
				rtn_obj['volunteer_number'] =  user.volunteer_number
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
		else:
			rtn_obj['ACK']=0
			rtn_obj['msg']= 'Please Login!'
			data = json.dumps(rtn_obj)
			messages.error(request, 'Please Login!') 
			return HttpResponse(data)
	def get(self,request):
		rtn_obj={}
		rtn_obj['ACK']=0
		rtn_obj['msg']= 'Post Method Allowed!'
		data = json.dumps(rtn_obj)
		return HttpResponse(data)

def payment_processing_execution(request,user,package):
		payload = request.body
		my_json = payload.decode('utf8')
		data = json.loads(my_json)['formdata']
		print(data)
		package = data['package']
		plan = data['plan']
		volunteer_number = data['volunteer_number']
		payment_environment = data['payment_environment']
		payment_mode = data['payment_mode']
		paymentMethod = data['payment_method']

		user_plan = SubscriptionPlanDetails.objects.filter(package=package,plan=plan,volunteer_number=volunteer_number,payment_environment=payment_environment,payment_mode=payment_mode).last()
		print(package)
		print(plan)
		print(volunteer_number)
		print(payment_environment)
		print(payment_mode)
		print(user_plan)
		country = data['country']
		return_dict={}

		trial_period_days = int(data['trial_period_days'])
		# check this user alredy take trial period facility or not
		chk_exist_payments = PayementInformations.objects.filter(user=user)
		try:
			if len(chk_exist_payments) >= 1 :
				trial_period_days = None
				print('kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk')
				existing_payment_obj =  chk_exist_payments.last()
				cust = stripe.Customer.retrieve(existing_payment_obj.customer_id)
				paymethod_attach = stripe.PaymentMethod.attach(
					paymentMethod,
					customer=cust['id'],
				)
				cust_modify = stripe.Customer.modify(
					cust['id'],
					# payment_method=paymentMethod,
					name = str(data['first_name'])+' '+str(data['last_name']),
					address = {
						'line1': data['state'],
						'postal_code': data['zip'],
						'city': data['state'],
						'state': data['state'],
						'country': data['country'],
					},
					invoice_settings={
						'default_payment_method': paymentMethod
					},
				)
				print('cccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccc')
				print(cust_modify)
				print('cusssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
				print(cust)
			else:
				cust = stripe.Customer.create(
					payment_method=paymentMethod,
					email=data['email'],
					invoice_settings={
						'default_payment_method': paymentMethod
					},
					name = str(data['first_name'])+' '+str(data['last_name']),
					address = {
						'line1': data['state'],
						'postal_code': data['zip'],
						'city': data['state'],
						'state': data['state'],
						'country': data['country'],
					},
					shipping={
						'name': str(data['first_name'])+' '+str(data['last_name']),
						'address': {
							'line1': data['state'],
							'postal_code': data['zip'],
							'city': data['state'],
							'state': data['state'],
							'country': data['country'],
						}
					},
				)

			if(trial_period_days==1):
				trial_period_days=settings.TRIAL_PERIOD
			else:
				trial_period_days=None

			subscription = stripe.Subscription.create(
				customer=cust['id'],
				items=[{"plan": user_plan.plan_id } ],
				trial_period_days = trial_period_days,
				expand=["latest_invoice.payment_intent"],
				default_payment_method = paymentMethod
			)
			return_dict['subscription']=subscription
			return_dict['setup_intent']=None
			if subscription['pending_setup_intent'] != None :
				print(subscription['pending_setup_intent'])
				setup_intent = stripe.SetupIntent.retrieve(subscription['pending_setup_intent'])
				return_dict['setup_intent'] = setup_intent
			return return_dict
		except Exception as e:
			return_dict['ACK'] = 0 
			return_dict['msg'] = str(e) 
			return return_dict

def payment_processing_execution_basic(request,userObj,package):
	#DB INSERTION
	rtn_obj = {}
	current_time =str( time.time()).split('.')[0]
	insert_to_payment_table = PayementInformations(
		user             = userObj,
		userprofile 	 = UserProfile.objects.get(user_id=userObj),
		volunteer_number = settings.MANUALADDVOL,
		customer_id  	 = userObj.id,
		email 			 = userObj.email,
		package 		 = package,
		subscription_start_date =current_time,
		subscribed = True,
	)
	insert_to_payment_table.save()

	# insert to USERPROFILE table (subscription end date and Subscribed= TRUE)
	user_profile = UserProfile.objects.get(user_id=userObj)
	user_profile.subscribed = True
	user_profile.save()

	messages.add_message(request, messages.SUCCESS, 'you have paid')
	subscription_start_date = datetime.fromtimestamp(int(current_time))
	rtn_obj['ACK']=1
	rtn_obj['package']= package
	rtn_obj['package_name']= switch_to_strings(package)
	rtn_obj['volunteer_number']= settings.MANUALADDVOL
	rtn_obj['subscription_start_date']= '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
	rtn_obj['subscribed']= True
	rtn_obj['msg']= 'You are successfully subscribed to this plan'
	return rtn_obj

def get_subscription(self):
	import json
	cursor13 = connection.cursor()
	sql3 = "select * from adminpanel_subscription"
	cursor13.execute(sql3)
	subscription = dictfetchall(cursor13)

	data = json.dumps({"Ack": 1, "subscription": subscription})

	return HttpResponse(data)	   

class paymenthistory(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self, request):
		rtn_dict={}
		resultArray = []
		user_obj = User.objects.get(id=request.POST['user_id'])
		user_profile_obj = UserProfile.objects.filter(user_id = user_obj).first()
		# print(user_profile_obj.role.id)

		# user_payment_history =  PayementInformations.objects.filter(user = user_obj).order_by('-id')
		if user_profile_obj.role.id == 3 :
			user_payment_history =  PayementInformations.objects.filter(user = user_obj).order_by('id')
			if len(user_payment_history) >= 1 : 
				index=0
				for obj in user_payment_history :
					rtn_dict ={}
					# rtn_dict[index] = []
					# rtn_dict[index].append('k')
					subscription_start_date = datetime.fromtimestamp(int(obj.subscription_start_date))
					subscription_end_date = datetime.fromtimestamp(int(obj.subscription_end_date))
					rtn_dict['volunter'] = obj.volunteer_number
					rtn_dict['email'] = obj.email
					rtn_dict['name'] = '{} {}'.format(obj.user.first_name, obj.user.last_name)
					rtn_dict['subscription_start'] ='{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
					rtn_dict['subscription_end'] = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
					print(obj.id)
					# rtn_dict['subscription_amount'] = json.loads(obj.plan_response)['amount']/100
					try:
						str_to_dict_result = ast.literal_eval(obj.plan_response)
					except Exception as ex:
						# if type(ex).__name__ == 'ValueError':
						str_to_dict_result = json.loads(obj.plan_response)

					rtn_dict['subscription_amount'] = str_to_dict_result['amount']/100
					rtn_dict['trial_period'] = obj.trial_period
					rtn_dict['subscription_ended'] = obj.subscription_ended
					rtn_dict['plan_type'] = obj.plan_type
					resultArray.append(rtn_dict)
					index+=1
			print(type(json.dumps(resultArray)))
		return HttpResponse(json.dumps(resultArray))

class user_subscription_details(APIView):

	permission_classes = (IsAuthenticated,)

	def post(self,request):
		user_obj = User.objects.get(id=request.POST['user_id'])
		user_profile_obj  = UserProfile.objects.get(user_id=user_obj)
		rtn_dict={}
		rtn_dict['subscribed'] = user_profile_obj.subscribed
		rtn_dict['co_ordinator'] = 0
		if user_profile_obj.role.id == 3 :
			rtn_dict['co_ordinator'] = 3
			paymentobj = PayementInformations.objects.filter(user=user_obj).last()
			if paymentobj:
				subscription_status = 0
				if paymentobj.subscription_deleted == SubcriptionDeletion.ACTIVE:
					subscription_status = 0
				elif paymentobj.subscription_deleted == SubcriptionDeletion.DELETED:
					subscription_status = 1
				elif paymentobj.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
					subscription_status = 3
				subscription_start_date = datetime.fromtimestamp(int(paymentobj.subscription_start_date))
				subscription_end_date = datetime.fromtimestamp(int(paymentobj.subscription_end_date))
				rtn_dict['package'] = paymentobj.package
				rtn_dict['package_name'] =  paymentobj.package
				rtn_dict['volunteer_number'] =  paymentobj.volunteer_number
				rtn_dict['plan_type'] = paymentobj.plan_type
				rtn_dict['trial_period'] = paymentobj.trial_period
				rtn_dict['subscription_ended'] = paymentobj.subscription_ended
				rtn_dict['subscription_deleted'] = subscription_status
				rtn_dict['subscription_start_date'] ='{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
				rtn_dict['subscription_end_date'] = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
		return HttpResponse(json.dumps(rtn_dict))

def switch_to_strings(argument): 
	package_name = settings.PACKAGE_NAME
	switcher = { 
		0: package_name[0], 
		1: package_name[1], 
		2: package_name[2]
	} 
	return switcher.get(argument,'nothing') 

class subscriptiondelete(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		rtn_dict = {}
		userObj = User.objects.get(id=request.POST['user_id'])
		last_subscribed_payment_obj = PayementInformations.objects.filter(user=userObj).last()
		subscription_status = 0
		# subscription delete
		stripe.Subscription.delete(last_subscribed_payment_obj.subscription_id)

		all_subs_collections  = PayementInformations.objects.filter(user=userObj,subscription_id=last_subscribed_payment_obj.subscription_id)
		if len(all_subs_collections) >= 1 :
			for obj in all_subs_collections :
				obj.subscription_deleted = 3
				obj.subscription_ended = True
				obj.subscribed = False
				obj.save()
				print('ok')
		new_last_subscribed_payment_obj = PayementInformations.objects.filter(user=userObj).last()
		print(new_last_subscribed_payment_obj.subscription_deleted)
		print('qweqweqweqweqrwfdscvdsvdfgtrhtyjyujy')
		
		# subscription_deleted = PayementInformations(
		# 	user        = userObj,
		# 	userprofile = UserProfile.objects.get(user_id=userObj),
		# 	volunteer_number    = last_subscribed_payment_obj.volunteer_number,
		# 	customer_id  = last_subscribed_payment_obj.customer_id,
		# 	email =userObj.email,
		# 	token =last_subscribed_payment_obj.token,
		# 	plan_id =last_subscribed_payment_obj.plan_id,
		# 	plan_type = last_subscribed_payment_obj.plan_type,
		# 	package = last_subscribed_payment_obj.package,
		# 	subscription_id = last_subscribed_payment_obj.subscription_id,
		# 	subscription_start_date = last_subscribed_payment_obj.subscription_start_date,
		# 	subscription_end_date = last_subscribed_payment_obj.subscription_end_date,
		# 	subscribed = False,
		# 	subscription_response = last_subscribed_payment_obj.subscription_response,
		# 	plan_response = last_subscribed_payment_obj.plan_response,
		# 	product_response = last_subscribed_payment_obj.product_response,
		# 	subscription_plan_details_id = last_subscribed_payment_obj.subscription_plan_details_id,
		# 	subscription_ended = True,
		# 	subscription_deleted = 3,
		# )
		# subscription_deleted.save()
		subscription_start_date = datetime.fromtimestamp(int(last_subscribed_payment_obj.subscription_start_date))
		subscription_end_date = datetime.fromtimestamp(int(last_subscribed_payment_obj.subscription_end_date))
		user_profile = UserProfile.objects.get(user_id=userObj)
		user_profile.subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
		user_profile.subscription_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
		user_profile.subscribed = False
		user_profile.save()
		print('jjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjjj')
		print(user_profile.subscribed)
		print(last_subscribed_payment_obj.subscription_deleted)
		print('jkgigjkghkjgkjgu')
		if new_last_subscribed_payment_obj.subscription_deleted == SubcriptionDeletion.ACTIVE:
			subscription_status = 0
		elif new_last_subscribed_payment_obj.subscription_deleted == SubcriptionDeletion.DELETED:
			subscription_status = 1
		elif new_last_subscribed_payment_obj.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
			subscription_status = 3
		subscription_start_date = datetime.fromtimestamp(int(new_last_subscribed_payment_obj.subscription_start_date))
		subscription_end_date = datetime.fromtimestamp(int(new_last_subscribed_payment_obj.subscription_end_date))
		request.session['package'] = new_last_subscribed_payment_obj.package
		request.session['plan_type'] = new_last_subscribed_payment_obj.plan_type
		request.session['trial_period'] = new_last_subscribed_payment_obj.trial_period
		request.session['subscription_ended'] = new_last_subscribed_payment_obj.subscription_ended
		request.session['subscription_deleted'] = subscription_status
		rtn_dict['ACK'] =1
		rtn_dict['msg'] ='You are unsubscribed successfully'
		rtn_dict['package'] =new_last_subscribed_payment_obj.package
		rtn_dict['plan_type'] =new_last_subscribed_payment_obj.plan_type
		rtn_dict['trial_period'] =new_last_subscribed_payment_obj.trial_period
		rtn_dict['subscription_ended'] =new_last_subscribed_payment_obj.subscription_ended
		rtn_dict['subscription_deleted'] =subscription_status
		rtn_dict['subscription_start_date'] ='{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
		rtn_dict['subscription_end_date'] = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
			
		return HttpResponse(json.dumps(rtn_dict))
	
def sdc(request):
	p = SubscriptionPlanDetails(package='premium',plan='year',plan_id='plan_GV3OcXBvepIYm8',volunteer_number='300-999',payment_mode='stripe',payment_environment='live')
	p.save()
	return HttpResponse(1)


class customUserSubscription(APIView):
	def post(self,request):
		amount = request.POST['amount']
		nickname = request.POST['nickname']
		currency = request.POST['currency']
		interval = request.POST['interval']
		product_response  = stripe.Product.retrieve(settings.STRIPE_PRODUCT_ID)
		plan = stripe.Plan.create(
			amount=amount,
			currency=currency,
			interval=interval,
			product=product_response,
			billing_scheme='per_unit',
			nickname = nickname,
			usage_type='licensed',
			interval_count=1
		)
		return JsonResponse({'ACK':1,'msg':'plan is successfully created'})
	
class unlockPackageSubscription(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		print(request.POST)
		print('HIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIIII')
		rtn_obj = {}
		try:
			userObj = User.objects.get(id=request.POST['user_id'])
			return_dict = unlock_subscription_execution(request,userObj)  # STRIPE SUBSCRIPTION 
			print(return_dict)
			if 'pay' in return_dict.keys():
				pay = return_dict['pay']
				if pay['status'] == 'paid' and pay['paid']==True:
					#DB INSERTION
					vol_number = SubscriptionPlanDetails.objects.filter(plan_id = return_dict['plan_response']['id']).first()
					insert_to_payment_table = UnlockPayementInformations(
						user        = userObj,
						userprofile = UserProfile.objects.get(user_id=userObj),
						customer_id  = return_dict['cust']['id'],
						email =userObj.email,
						token =request.POST['stripeToken'],
						plan_id =return_dict['plan_response']['id'],
						plan_type = request.POST['plan'],
						package = request.POST['package'],
						subscription_id = return_dict['subs']['id'],
						subscription_start_date = return_dict['subs']['current_period_start'],
						subscription_end_date = return_dict['subs']['current_period_end'],
						subscribed = True,
						subscription_response = return_dict['subs'],
						plan_response = return_dict['plan_response'],
						product_response = return_dict['product_response'],
						subscription_plan_details_id = return_dict['subscription_plan_details_id']
					)
					insert_to_payment_table.save()

					subscription_start_date = datetime.fromtimestamp(int(return_dict['subs']['current_period_start']))
					subscription_end_date = datetime.fromtimestamp(int(return_dict['subs']['current_period_end']))
					user_profile = UserProfile.objects.get(user_id=userObj)
					insert_to_payment_table.unlock_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
					insert_to_payment_table.unlock_end_date   = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
					insert_to_payment_table.save()

					if insert_to_payment_table.subscription_deleted == SubcriptionDeletion.ACTIVE:
						subscription_status = 0
					elif insert_to_payment_table.subscription_deleted == SubcriptionDeletion.DELETED:
						subscription_status = 1
					elif insert_to_payment_table.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
						subscription_status = 3

					request.session['package'] = insert_to_payment_table.package
					request.session['plan_type'] = insert_to_payment_table.plan_type
					request.session['subscription_ended'] = insert_to_payment_table.subscription_ended
					request.session['subscription_deleted'] = subscription_status

					messages.add_message(request, messages.SUCCESS, 'you have paid')
					
					rtn_obj['ACK']=1
					rtn_obj['customer_id']= return_dict['cust']['id']
					rtn_obj['token']= request.POST['stripeToken']
					rtn_obj['plan_type']= request.POST['plan']
					rtn_obj['package']= request.POST['package']
					rtn_obj['package_name']= request.POST['package']
					rtn_obj['plan_id']= return_dict['plan_response']['id']
					rtn_obj['subscription_id']= return_dict['subs']['id']
					rtn_obj['subscription_start_date']= '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
					rtn_obj['subscription_end_date']= '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
					rtn_obj['subscribed']= True
					rtn_obj['subscription_response']= return_dict['subs']
					rtn_obj['plan_response']= return_dict['plan_response']
					rtn_obj['product_response']= return_dict['product_response']
					rtn_obj['subscription_ended']=  insert_to_payment_table.subscription_ended
					rtn_obj['subscription_deleted']=  subscription_status

					rtn_obj['msg']= 'You are successfully subscribed to this plan'
					data = json.dumps(rtn_obj)
					return HttpResponse(data)
					# return HttpResponseRedirect(reverse('stripehome'))
				else:
					rtn_obj['ACK']=0
					rtn_obj['msg']= 'Your payment not done'
					data = json.dumps(rtn_obj)
					return HttpResponse(data)
			else:
				return JsonResponse(return_dict)
		except User.DoesNotExist:
			return JsonResponse({'ACK':0,'msg':'User does not exist'})

def unlock_subscription_execution(request,userObj):
	country = request.POST['country']
	stripeToken = request.POST['stripeToken']
	addr_line1 = request.POST['addr_line1']
	state = request.POST['state']
	user_id = request.POST['user_id']
	first_name = request.POST['first_name']
	last_name = request.POST['last_name']
	package = request.POST['package']
	plan = request.POST['plan']
	payment_environment = request.POST['payment_environment']
	payment_mode = request.POST['payment_mode']
	interval_count = request.POST['interval_count']
	zip_code = request.POST['zip_code']
	return_dict = {}
	user_plan = SubscriptionPlanDetails.objects.filter(package=package,plan=plan,payment_environment=payment_environment,payment_mode=payment_mode,interval_count=interval_count).last()
	if user_plan != None:
		try:
			# check user  already exist or not in the paymentTable
			user_payment_obj = UnlockPayementInformations.objects.filter(user=userObj,subscription_ended=False).last()
		except UnlockPayementInformations.DoesNotExist:
			user_payment_obj = None
		if user_payment_obj == None :
			chk_exist_payments = UnlockPayementInformations.objects.filter(user=userObj)
			print(chk_exist_payments)
			if len(chk_exist_payments) >= 1 :
				existing_payment_obj =  chk_exist_payments.last()
				cust = stripe.Customer.retrieve(existing_payment_obj.customer_id)
				try:
					
					cust_card_source = stripe.Customer.create_source(
						existing_payment_obj.customer_id,
						source = stripeToken,
					)
					
					cust_modify = stripe.Customer.modify(
						existing_payment_obj.customer_id,
						default_source=cust_card_source['id']
					)

					cust_modify = stripe.Customer.modify_source(
						existing_payment_obj.customer_id,
						cust_card_source['id'],
						name = first_name+' '+last_name,
						address_city= country,
						address_country= country,
						address_line1= country,
						address_line2= state,
						address_state=state,
					)
				except:
					return {'ACK':0,'msg':'Same Stripe token can be used only for one time'}
				
				print('cccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccc')
				print(cust_modify)
				print('cusssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
				print(cust)
			else:
				cust = stripe.Customer.create(
					email = userObj.email,
					source  = stripeToken,
				)
				cust_modify = stripe.Customer.modify_source(
					cust['id'],
					cust['default_source'],
					name = first_name+' '+last_name,
					address_city= country,
					address_country= country,
					address_line1= country,
					address_line2= state,
					address_state= state,
				)
			subs = stripe.Subscription.create(
				customer=cust['id'],
				items=[{"plan": user_plan.plan_id } ],
			)
			invoice_payment_info = stripe.Invoice.retrieve(
				subs['latest_invoice']
			)
			
			pay=invoice_payment_info
			if invoice_payment_info['status'] == 'open' and country=='India':
				country = 'Germany' # non INR charges outside india, shipping / billing address should be outside INDIA
			
				payment_intent = stripe.PaymentIntent.modify(
					invoice_payment_info['payment_intent'], #payment_intent(invoice)
					shipping={
						'name': first_name+' '+last_name,
						'address': {
							'line1': state,
							'postal_code': zip_code,
							'city': state,
							'state': state,
							'country': country,
						}
					},
				)
				pay = stripe.Invoice.pay(subs['latest_invoice'])
			plan_response     = stripe.Plan.retrieve(user_plan.plan_id)
			product_response  = stripe.Product.retrieve(plan_response['product'])
			return_dict['pay']=pay
			return_dict['cust']=cust
			return_dict['subs']=subs
			return_dict['invoice_payment_info']=invoice_payment_info
			return_dict['plan_response']=plan_response
			return_dict['product_response']=product_response
			return_dict['subscription_plan_details_id']=user_plan
			return return_dict
		else:
			return {'ACK':1,'msg':'user is already subscribed to a plan'}
	else:
		return {'ACK':1,'msg':'Plan does not exist'}


@csrf_exempt
def stripewebhook(request):
	import json
	payload = request.body
	print('bbbbbbbbbbbbbbbbb')
	my_json = payload.decode('utf8')
	event = None
	try:
		event = stripe.Event.construct_from(json.loads(my_json), stripe.api_key)
		print(event.type)
		if event.type == 'customer.subscription.updated':
			sub_evt = event.data.object
			sub_evt_json =  json.loads(json.dumps(sub_evt))
			subscription_status = 0

			# product_response  = stripe.Product.retrieve('prod_GCtzKIadGhIQjp')
			# package = SubscriptionPlanDetails.objects.filter(plan_id = 'plan_GWNlbw93SD4iOg').first()
			# subscribedObjList  = PayementInformations.objects.filter(subscription_id='sub_GXV7ZQPY6FgHya')
			# unlock_subscribed_obj_list  = UnlockPayementInformations.objects.filter(subscription_id='sub_GFsluQIc40lKkK')

			product_response  = stripe.Product.retrieve(sub_evt_json['plan']['product'])
			package = SubscriptionPlanDetails.objects.filter(plan_id = sub_evt_json['plan']['id']).first()
			subscribedObjList  = PayementInformations.objects.filter(subscription_id=sub_evt_json['id'])

			print('package---------------')

			print(sub_evt_json['plan']['id'])
			print(package)

		

			print(len(subscribedObjList))
			print(subscribedObjList)

			if len(subscribedObjList) >= 1 and (package)!=None:
				chk_subscribed_exist = PayementInformations.objects.filter(subscription_id=sub_evt_json['id'],subscription_start_date=sub_evt_json['current_period_start'],subscription_end_date=sub_evt_json['current_period_end'])
				print(len(chk_subscribed_exist))
				if len(chk_subscribed_exist) == 0:
					userObj = 0
					token = 0
					plan_type = 0
					volunteer_number = 0
					for obj in subscribedObjList :
						obj.subscription_deleted = 1
						obj.subscription_ended = True
						obj.subscribed = False
						obj.save()
						userObj = obj.user
						token = obj.token
						plan_type = obj.plan_type
						volunteer_number = obj.volunteer_number
						print('ok')
					
					subscription_renew = PayementInformations(
						user        = userObj,
						userprofile = UserProfile.objects.get(user_id=userObj),
						volunteer_number    = volunteer_number,
						customer_id  = sub_evt_json['customer'],
						email =userObj.email,
						token =token,
						plan_id =sub_evt_json['plan']['id'],
						# plan_id ='plan_GHnoTOwhdFrBMp',
						plan_type = package.plan,
						package = package.package,
						subscription_id = sub_evt_json['id'],
						subscription_start_date = sub_evt_json['current_period_start'],
						subscription_end_date = sub_evt_json['current_period_end'],
						subscribed = True,
						subscription_response = sub_evt_json,
						plan_response = sub_evt_json['plan'],
						product_response = product_response,
						subscription_plan_details_id = package
					)
					subscription_renew.save()

					subscription_start_date = datetime.fromtimestamp(int(sub_evt_json['current_period_start']))
					subscription_end_date = datetime.fromtimestamp(int(sub_evt_json['current_period_end']))
					user_profile = UserProfile.objects.get(user_id=userObj)
					user_profile.subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
					user_profile.subscription_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
					user_profile.subscribed = True
					user_profile.save()

					if subscription_renew.subscription_deleted == SubcriptionDeletion.ACTIVE:
						subscription_status = 0
					elif subscription_renew.subscription_deleted == SubcriptionDeletion.DELETED:
						subscription_status = 1
					elif subscription_renew.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
						subscription_status = 3
					print('eeeeeeeeeeeee')
					return JsonResponse({'ACK':1,'msg':'Subscription is successfully updated'})
			
		return JsonResponse({'ACK':0,'msg':' subscription is not updated'})
		# return HttpResponse(status=200)
	except ValueError as e:
		print('errorrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr')
		print(e)
		return JsonResponse({'ACK':0,'msg':'Invalid value'})

def ttt(request):
	ppp = PayementInformations.objects.get(id=79)
	userObj = User.objects.get(id=42)
	insert_to_payment_table = PayementInformations(
		user        = userObj,
		userprofile = UserProfile.objects.get(user_id=userObj),
		volunteer_number    = 99,
		customer_id  = ppp.customer_id,
		email =userObj.email,
		token =ppp.token,
		plan_id =ppp.plan_id,
		plan_type = ppp.plan_type,
		package = ppp.package,
		trial_period =None,
		subscription_id = ppp.subscription_id,
		subscription_start_date = ppp.subscription_start_date,
		subscription_end_date = ppp.subscription_end_date,
		subscribed = False,
		subscription_response = ppp.subscription_response,
		plan_response = ppp.plan_response,
		product_response = ppp.product_response,
		subscription_plan_details_id = SubscriptionPlanDetails.objects.get(id=2),
		subscription_ended = True,
		subscription_deleted = SubcriptionDeletion.DELETED
	)
	insert_to_payment_table.save()
	return HttpResponse(1)

def public_key(request):
	return JsonResponse({'publicKey': settings.STRIPE_PUBLISHABLE_KEY })

def create_customer(request):
	# print(request.POST)
	print(request.body)
	# print(dir(request))
	payload = request.body
	my_json = payload.decode('utf8')
	data = json.loads(my_json)['formdata']
	print(data)
	paymentMethod = data['payment_method']
	# print(paymentMethod)
	user =0
	setup_intent = {}
	if user == 0 :
		customer = stripe.Customer.create(
			payment_method=paymentMethod,
			email=data['email'],
			invoice_settings={
				'default_payment_method': paymentMethod
			},
			shipping={
				'name': 'a',
				'address': {
					'line1': 'wb',
					'postal_code': '712258',
					'city': 'kolkata',
					'state': 'wb',
					# 'country': 'india',
					'country': 'Germany',
				}
			},
		)


		# paymethod_attach = stripe.PaymentMethod.attach(
		#     paymentMethod,
		#     customer="cus_GNPc3K1mVkUppm",
		# )
		# customer = stripe.Customer.modify(
		#     "cus_GNPc3K1mVkUppm",
		#     # payment_method=paymentMethod,
		#     invoice_settings={
		#         'default_payment_method': paymentMethod
		#     },
		# )
		# At this point, associate the ID of the Customer object with your
		# own internal representation of a customer, if you have one.
		# print(customer)

		# Subscribe the user to the subscription created
		subscription = stripe.Subscription.create(
			customer=customer.id,
			items=[
				{
					# "plan": 'plan_GtANVbPECdrLId',
					"plan": 'plan_GtdZDrhLY5iNpH', #one day subscription'''
				},
			],
			expand=["latest_invoice.payment_intent"],
			# trial_period_days=5,
			default_payment_method = paymentMethod
		)
		if subscription['pending_setup_intent'] != None :
			print(subscription['pending_setup_intent'])
			setup_intent = stripe.SetupIntent.retrieve(subscription['pending_setup_intent'])
			print(setup_intent)
		data = json.dumps({'ACK':1,'subscription':subscription,'setup_intent':setup_intent})
		# data = 1
		return HttpResponse(data)
	else :
		rtn_obj={}
		rtn_obj['ACK']=0
		rtn_obj['msg']= 'You have already subscribe to the plan'
		data = json.dumps(rtn_obj)
		print(data)
		return HttpResponse(data)

class subscription(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		payload = request.body
		my_json = payload.decode('utf8')
		
		data = json.loads(my_json)
		print('zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz')
		print(data)
		formData = data['formdata']
		stripeToken = formData['stripeToken']
		try:
			subscription = stripe.Subscription.retrieve(data['subscriptionId'])
			plan_response     = stripe.Plan.retrieve(subscription['plan']['id'])
			product_response  = stripe.Product.retrieve(plan_response['product'])
			user_plan = SubscriptionPlanDetails.objects.filter(plan_id=subscription['plan']['id']).last()
			trial_period_days = None
			rtn_obj = {}
			if subscription['status'] == 'active' or subscription['status'] == 'trialing':
				#DB INSERTION
				vol_number = SubscriptionPlanDetails.objects.filter(plan_id = subscription['plan']['id']).first()
				userObj = User.objects.get(id=data['user_id'])

				if subscription['trial_end'] != None :
					trial_period_days = subscription['plan']['trial_period_days']

				insert_to_payment_table = PayementInformations(
					user        = userObj,
					userprofile = UserProfile.objects.get(user_id=userObj),
					volunteer_number    = vol_number.volunteer_number.split('-')[1],
					customer_id  = subscription['customer'],
					email =userObj.email,
					token =stripeToken,
					plan_id =plan_response['id'],
					plan_type = formData['plan'],
					package = formData['package'],
					trial_period =trial_period_days,
					subscription_id = data['subscriptionId'],
					subscription_start_date = subscription['current_period_start'],
					subscription_end_date = subscription['current_period_end'],
					subscribed = True,
					subscription_response = subscription,
					plan_response =plan_response,
					product_response = product_response,
					subscription_plan_details_id = user_plan
				)
				print(';kkballlllllllllllll')
				print(insert_to_payment_table)
				insert_to_payment_table.save()

				# insert to USERPROFILE table (subscription end date and Subscribed= TRUE)
				subscription_start_date = datetime.fromtimestamp(int(subscription['current_period_start']))
				subscription_end_date = datetime.fromtimestamp(int(subscription['current_period_end']))
				user_profile = UserProfile.objects.get(user_id=userObj)
				user_profile.subscription_start_date = '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
				user_profile.subscription_end_date = '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
				user_profile.subscribed = True
				user_profile.save()

				if insert_to_payment_table.subscription_deleted == SubcriptionDeletion.ACTIVE:
					subscription_status = 0
				elif insert_to_payment_table.subscription_deleted == SubcriptionDeletion.DELETED:
					subscription_status = 1
				elif insert_to_payment_table.subscription_deleted == SubcriptionDeletion.DEACTIVATED:
					subscription_status = 3

				request.session['package'] = insert_to_payment_table.package
				request.session['plan_type'] = insert_to_payment_table.plan_type
				request.session['trial_period'] = insert_to_payment_table.trial_period
				request.session['subscription_ended'] = insert_to_payment_table.subscription_ended
				request.session['subscription_deleted'] = subscription_status

				messages.add_message(request, messages.SUCCESS, 'you have paid')
				
				rtn_obj['ACK']=1
				rtn_obj['customer_id']= subscription['customer']
				rtn_obj['token']= stripeToken
				rtn_obj['plan_type']= formData['plan']
				rtn_obj['package']= formData['package']
				rtn_obj['package_name']= switch_to_strings(formData['package'])
				rtn_obj['plan_id']= plan_response['id']
				rtn_obj['trial_period']= trial_period_days
				rtn_obj['volunteer_number']= vol_number.volunteer_number.split('-')[1]					
				rtn_obj['subscription_id']= data['subscriptionId']
				rtn_obj['subscription_start_date']= '{}-{}-{}'.format(subscription_start_date.year, subscription_start_date.month, subscription_start_date.day)
				rtn_obj['subscription_end_date']= '{}-{}-{}'.format(subscription_end_date.year, subscription_end_date.month, subscription_end_date.day)
				rtn_obj['subscribed']= True
				rtn_obj['subscription_response']= subscription
				rtn_obj['plan_response']= plan_response
				rtn_obj['product_response']= product_response
				rtn_obj['subscription_ended']=  insert_to_payment_table.subscription_ended
				rtn_obj['subscription_deleted']=  subscription_status

				rtn_obj['msg']= 'You are successfully subscribed to this plan'
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			else:
				rtn_obj['ACK']=0
				rtn_obj['msg']= 'Your payment not done'
				data = json.dumps(rtn_obj)
				return HttpResponse(data)
			return JsonResponse({'ACK':1,'subscription':subscription,'rtn_obj':rtn_obj})
		except Exception as e:
			return JsonResponse({'ACK':0,'msg':str(e)},status=status.HTTP_403_FORBIDDEN)

class cancel_subscription(APIView):
	
	permission_classes = (IsAuthenticated,)

	def post(self,request):
		payload = request.body
		my_json = payload.decode('utf8')
		
		data = json.loads(my_json)
		try:
			subscriptionId = data['subscriptionId']
			stripe.Subscription.delete(subscriptionId)
		except :
			pass
		return JsonResponse(1, safe=False)

# Using Django


	# return HttpResponse(1)
	# p = serializers.serialize("json", request.POST)
	# return HttpResponse(p, content_type='application/json')
	# return HttpResponse(p)
	# p = stripe.WebhookEndpoint.create(
	# 	url="nn/api/webhook/",
	# 	enabled_events=[
	# 		"customer.subscription.trial_will_end",
	# 	],
	# )
	# return HttpResponse(json.dumps(dir(stripe.WebhookEndpoint)))
#   payload = request.body
#   event = None

#   try:
#     event = stripe.Event.construct_from(
#       json.loads(payload), stripe.api_key
#     )
#   except ValueError as e:
#     # Invalid payload
#     return HttpResponse(status=400)

#   # Handle the event
#   if event.type == 'payment_intent.succeeded':
#     payment_intent = event.data.object # contains a stripe.PaymentIntent
#     print('PaymentIntent was successful!')
#   elif event.type == 'payment_method.attached':
#     payment_method = event.data.object # contains a stripe.PaymentMethod
#     print('PaymentMethod was attached to a Customer!')
#   # ... handle other event types
#   else:
#     # Unexpected event type
#     return HttpResponse(status=400)

#   return HttpResponse(status=200)

# s = sched.scheduler(time.time, time.sleep)
# def do_something(sc): 
#     print("Doing stuff") 
#     # do your stuff
#     s.enter(60, 1, do_something, (sc,)) 

# s.enter(60, 1, do_something, (s,)) 
# s.run()
# s.start()